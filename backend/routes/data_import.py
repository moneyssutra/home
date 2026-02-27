"""Data import/export routes - Excel sample download and bulk import."""
import io
import uuid
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from server_models import (
    IncomeSource, Expense, Investment, Asset, Loan, CreditCard, Insurance, Account, Goal
)
from routes.auth import get_current_user
from database import db

def get_user_filter(user):
    return {"userId": user.get("user_id")}

router = APIRouter(prefix="/data", tags=["data"])

HEADER_FILL = PatternFill(start_color="14B8A6", end_color="14B8A6", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
EXAMPLE_FILL = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

SHEETS = {
    "Income": {
        "columns": ["name", "type", "expectedAmount", "frequency", "selectedDate", "selectedDay", "selectedMonth", "selectedQuarter", "selectedHalf", "isVariable"],
        "types": "Salary, Business, Self-Employed, Commission, Other",
        "frequencies": "Daily, Weekly, Monthly, Quarterly, Half-Yearly, Yearly",
        "examples": [
            ["Software Engineer Salary", "Salary", 150000, "Monthly", "1", "", "", "", "", "No"],
            ["Freelance Web Dev", "Self-Employed", 40000, "Monthly", "20", "", "", "", "", "Yes"],
            ["Annual Bonus", "Salary", 300000, "Yearly", "", "", "March", "", "", "No"],
        ]
    },
    "Expenses": {
        "columns": ["expenseName", "expenseType", "category", "expectedAmount", "frequency", "selectedDate", "selectedDay", "selectedMonth", "selectedQuarter"],
        "types": "Fixed, Variable",
        "categories": "Housing, Utilities, Food, Travel, Shopping, Medical, Education, Insurance, Subscriptions, EMI, Business Expense, Salary Paid, Investments, Savings",
        "frequencies": "Daily, Weekly, Monthly, Quarterly, Half-Yearly, Yearly, One-Time",
        "examples": [
            ["Rent", "Fixed", "Housing", 28000, "Monthly", "1", "", "", ""],
            ["Groceries", "Variable", "Food", 12000, "Monthly", "1", "", "", ""],
            ["Kids School Fees", "Fixed", "Education", 25000, "Quarterly", "1", "", "", "Q1 (Jan-Mar)"],
        ]
    },
    "Investments": {
        "columns": ["name", "investmentCategory", "investmentMode", "principal", "currentValue", "startDate", "investmentFrequency", "sipAmount", "sipSelectedDate", "returnRate", "maturityDate", "isLiquidAsset"],
        "categories": "Fixed Deposit (FD), Recurring Deposit (RD), Stocks, US Stocks, Mutual Fund, ETF, Bonds, Sovereign Gold Bond (SGB), Digital Gold, P2P Lending, ULIP, Crypto, PPF, NPS, SWP, Other",
        "modes": "Growth Only, Income Generating, Growth with Maturity",
        "examples": [
            ["SBI FD 3yr", "Fixed Deposit (FD)", "Income Generating", 500000, 540000, "2024-01-15", "", "", "", 7.25, "2027-01-15", "No"],
            ["Axis Bluechip MF", "Mutual Fund", "Growth Only", 0, 180000, "2023-01-10", "Monthly", 10000, "10", "", "", "Yes"],
            ["PPF Account", "PPF", "Growth with Maturity", 0, 350000, "2020-04-01", "Monthly", 12500, "5", 7.1, "2035-04-01", "No"],
        ]
    },
    "Assets": {
        "columns": ["assetName", "assetType", "purchaseValue", "currentValue", "purchaseDate", "isFinanced"],
        "types": "Residential Property, Commercial Property, Land, Vehicle, Physical Gold, Physical Silver, Diamonds, Business Asset, Equipment / Machinery, Other",
        "examples": [
            ["2BHK Flat Pune", "Residential Property", 6500000, 8200000, "2020-03-15", "Yes"],
            ["Honda City 2023", "Vehicle", 1500000, 1200000, "2023-08-20", "No"],
            ["Gold Jewelry", "Physical Gold", 450000, 620000, "2018-11-01", "No"],
        ]
    },
    "Loans": {
        "columns": ["loanName", "loanType", "lenderName", "principalAmount", "outstandingAmount", "interestRate", "emiAmount", "emiFrequency", "tenureMonths", "startDate", "endDate", "autoCreateExpense"],
        "types": "Home Loan, Vehicle Loan, Personal Loan, Education Loan, Business Loan, Gold Loan, Credit Card Dues, Hand Loan Taken, Other",
        "examples": [
            ["HDFC Home Loan", "Home Loan", "HDFC Bank", 5000000, 4200000, 8.5, 45000, "Monthly", 240, "2020-03-01", "2040-03-01", "Yes"],
            ["Car Loan", "Vehicle Loan", "ICICI Bank", 1000000, 650000, 9.5, 22000, "Monthly", 60, "2023-08-01", "2028-08-01", "Yes"],
        ]
    },
    "Credit Cards": {
        "columns": ["cardName", "bankName", "creditLimit", "outstandingAmount", "billingDate", "dueDate", "interestRate", "minimumDue"],
        "examples": [
            ["HDFC Regalia", "HDFC Bank", 500000, 45000, 15, 5, 42, 2250],
            ["SBI Simply Click", "SBI", 200000, 12000, 20, 10, 39, 600],
        ]
    },
    "Insurance": {
        "columns": ["policyName", "insuranceType", "coverageAmount", "premiumAmount", "premiumFrequency", "startDate", "endDate", "coveredPerson", "autoCreateExpense"],
        "types": "Term Insurance, Life Insurance, Health Insurance, Vehicle Insurance, Property Insurance, Business Insurance, Travel Insurance, Other",
        "examples": [
            ["ICICI iProtect Smart", "Term Insurance", 10000000, 12000, "Yearly", "2022-04-01", "2052-04-01", "Self", "Yes"],
            ["Star Family Floater", "Health Insurance", 1000000, 25000, "Yearly", "2023-06-01", "2024-06-01", "Self", "Yes"],
        ]
    },
    "Accounts": {
        "columns": ["accountName", "accountType", "currentBalance", "isPrimary"],
        "types": "Bank Account, Cash, Wallet, Others",
        "examples": [
            ["HDFC Savings", "Bank Account", 250000, "Yes"],
            ["Cash in Hand", "Cash", 15000, "No"],
            ["Paytm Wallet", "Wallet", 3500, "No"],
        ]
    },
    "Goals": {
        "columns": ["goalName", "goalType", "targetAmount", "currentAmount", "targetDate"],
        "types": "Wealth Creation, Debt Elimination, Investment Target, Emergency Fund, Other",
        "examples": [
            ["Early Retirement Corpus", "Wealth Creation", 5000000, 800000, "2030-12-31"],
            ["6-Month Safety Net", "Emergency Fund", 600000, 305000, "2027-01-01"],
        ]
    },
}


def _style_sheet(ws, columns, info, examples):
    """Style a worksheet with headers, info row, and example data."""
    # Info rows at top
    info_lines = []
    for key, val in info.items():
        info_lines.append(f"{key}: {val}")
    if info_lines:
        for i, line in enumerate(info_lines):
            ws.cell(row=i + 1, column=1, value=line).font = Font(italic=True, color="666666", size=10)
            ws.merge_cells(start_row=i + 1, start_column=1, end_row=i + 1, end_column=len(columns))

    header_row = len(info_lines) + 2

    # Headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(col_name) + 4)

    # Example rows
    for row_idx, example in enumerate(examples, header_row + 1):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER

    return header_row


@router.get("/sample-excel")
async def download_sample_excel(request: Request):
    """Download a sample Excel template with all data type sheets."""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, config in SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        info = {}
        if "types" in config:
            info["Valid Types"] = config["types"]
        if "categories" in config:
            info["Valid Categories"] = config["categories"]
        if "modes" in config:
            info["Valid Modes"] = config["modes"]
        if "frequencies" in config:
            info["Valid Frequencies"] = config["frequencies"]
        _style_sheet(ws, config["columns"], info, config["examples"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=MoneySSutra_Import_Template.xlsx"}
    )


@router.post("/import-excel")
async def import_excel(request: Request, file: UploadFile = File(...)):
    """Import data from an uploaded Excel file."""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    user_id = user.get("user_id")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))

    results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row (first row with known column names)
        header_row_idx = None
        headers = None
        config = SHEETS.get(sheet_name, {})
        expected_cols = set(config.get("columns", []))

        for i, row in enumerate(rows):
            row_vals = [str(v).strip() if v else "" for v in row]
            if expected_cols and expected_cols.intersection(set(row_vals)):
                header_row_idx = i
                headers = row_vals
                break

        if headers is None:
            continue

        data_rows = rows[header_row_idx + 1:]
        created = 0
        errors = []

        for row_idx, row in enumerate(data_rows):
            row_dict = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row) and header and row[col_idx] is not None:
                    val = row[col_idx]
                    # Convert Yes/No to bool
                    if isinstance(val, str) and val.strip().lower() in ("yes", "true"):
                        val = True
                    elif isinstance(val, str) and val.strip().lower() in ("no", "false"):
                        val = False
                    row_dict[header] = val

            # Skip empty rows
            key_field = config["columns"][0] if config.get("columns") else None
            if not key_field or not row_dict.get(key_field):
                continue

            try:
                if sheet_name == "Income":
                    row_dict["userId"] = user_id
                    row_dict["incomeType"] = "variable" if row_dict.pop("isVariable", False) else "fixed"
                    obj = IncomeSource(**row_dict)
                    doc = obj.model_dump()
                    doc["createdAt"] = doc["createdAt"].isoformat()
                    await db.income_sources.insert_one(doc)
                elif sheet_name == "Expenses":
                    row_dict["userId"] = user_id
                    obj = Expense(**row_dict)
                    doc = obj.model_dump()
                    doc["createdAt"] = doc["createdAt"].isoformat()
                    await db.expenses.insert_one(doc)
                elif sheet_name == "Investments":
                    row_dict["userId"] = user_id
                    obj = Investment(**row_dict)
                    doc = obj.model_dump()
                    doc["createdAt"] = doc["createdAt"].isoformat()
                    await db.investments.insert_one(doc)
                elif sheet_name == "Assets":
                    row_dict["userId"] = user_id
                    obj = Asset(**row_dict)
                    doc = obj.model_dump()
                    doc["createdAt"] = doc["createdAt"].isoformat()
                    await db.assets.insert_one(doc)
                elif sheet_name == "Loans":
                    row_dict["userId"] = user_id
                    obj = Loan(**row_dict)
                    doc = obj.model_dump()
                    doc["createdAt"] = doc["createdAt"].isoformat()
                    await db.loans.insert_one(doc)
                elif sheet_name == "Credit Cards":
                    row_dict["userId"] = user_id
                    obj = CreditCard(**row_dict)
                    doc = obj.model_dump()
                    doc["createdAt"] = doc["createdAt"].isoformat()
                    await db.credit_cards.insert_one(doc)
                elif sheet_name == "Insurance":
                    row_dict["userId"] = user_id
                    obj = Insurance(**row_dict)
                    doc = obj.model_dump()
                    doc["createdAt"] = doc["createdAt"].isoformat()
                    await db.insurances.insert_one(doc)
                elif sheet_name == "Accounts":
                    row_dict["userId"] = user_id
                    obj = Account(**row_dict)
                    doc = obj.model_dump()
                    doc["createdAt"] = doc["createdAt"].isoformat()
                    await db.accounts.insert_one(doc)
                elif sheet_name == "Goals":
                    row_dict["userId"] = user_id
                    obj = Goal(**row_dict)
                    doc = obj.model_dump()
                    doc["createdAt"] = doc["createdAt"].isoformat()
                    await db.goals.insert_one(doc)
                else:
                    continue
                created += 1
            except Exception as e:
                errors.append(f"Row {row_idx + 1}: {str(e)[:100]}")

        results[sheet_name] = {"created": created, "errors": errors}

    return {"message": "Import complete", "results": results}
