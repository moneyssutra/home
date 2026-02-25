"""Reports routes - Bank-statement style PDF/Excel report generation."""
from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
from typing import Optional
from io import BytesIO
from collections import defaultdict

router = APIRouter(prefix="/reports", tags=["reports"])

import sys
sys.path.insert(0, '/app/backend')
from database import db
from routes.auth import get_current_user

RUPEE = "\u20b9"
import os
FONT_PATH = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
FONT_BOLD_PATH = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
if not os.path.exists(FONT_PATH):
    FONT_PATH = "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"
    FONT_BOLD_PATH = "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"

# Brand colors
NAVY = '#1B263B'
DARK_GRAY = '#1E293B'
MED_GRAY = '#475569'
LIGHT_GRAY = '#F1F5F9'
BORDER_GRAY = '#CBD5E1'
ACCENT_GREEN = '#059669'
ACCENT_RED = '#DC2626'
ACCENT_BLUE = '#00D1CC'
WHITE = '#FFFFFF'

REPORT_META = {
    "income": {"title": "Income Statement", "accent": ACCENT_GREEN, "icon": "INCOME"},
    "expense": {"title": "Expense Statement", "accent": ACCENT_RED, "icon": "EXPENSE"},
    "cashflow": {"title": "Cash Flow Statement", "accent": ACCENT_BLUE, "icon": "CASHFLOW"},
    "loan": {"title": "Loan Statement", "accent": '#D97706', "icon": "LOANS"},
    "investment": {"title": "Investment Statement", "accent": '#7C3AED', "icon": "INVESTMENTS"},
    "networth": {"title": "Net Worth Statement", "accent": ACCENT_BLUE, "icon": "NET WORTH"},
    "goal": {"title": "Goal Progress Report", "accent": '#DB2777', "icon": "GOALS"},
    "asset": {"title": "Asset Statement", "accent": '#0891B2', "icon": "ASSETS"},
    "insurance": {"title": "Insurance Statement", "accent": '#BE123C', "icon": "INSURANCE"},
}


def fmt_date(val):
    if not val:
        return "-"
    if isinstance(val, str):
        try:
            val = datetime.fromisoformat(val.replace("Z", "+00:00"))
        except Exception:
            return val[:10] if len(val) >= 10 else val
    if isinstance(val, datetime):
        return val.strftime("%d %b %Y")
    return str(val)


def fmt_amount(val, show_sign=False):
    if val is None:
        val = 0
    if show_sign and val > 0:
        return f"+{RUPEE}{val:,.0f}"
    if show_sign and val < 0:
        return f"-{RUPEE}{abs(val):,.0f}"
    return f"{RUPEE}{val:,.0f}"


def fmt_amount_plain(val):
    if val is None:
        val = 0
    return f"{val:,.0f}"


@router.get("/generate/{report_type}")
async def generate_report(
    request: Request,
    report_type: str,
    format: str = "pdf",
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    user_id = user.get('user_id')
    user_name = user.get('name', 'User')
    user_email = user.get('email', '')
    user_filter = {"userId": user_id}

    try:
        start_date = datetime.strptime(from_date, "%Y-%m-%d") if from_date else datetime.now(timezone.utc).replace(day=1)
        end_date = datetime.strptime(to_date, "%Y-%m-%d") if to_date else datetime.now(timezone.utc)
    except Exception:
        start_date = datetime.now(timezone.utc).replace(day=1)
        end_date = datetime.now(timezone.utc)

    data = {}
    if report_type in ["income", "cashflow", "networth"]:
        data["incomes"] = await db.income_sources.find(user_filter, {"_id": 0}).to_list(1000)
        data["other_incomes"] = await db.other_income.find(user_filter, {"_id": 0}).to_list(1000)
    if report_type in ["expense", "cashflow", "networth"]:
        data["expenses"] = await db.expenses.find(user_filter, {"_id": 0}).to_list(1000)
    if report_type in ["loan", "networth"]:
        data["loans"] = await db.loans.find(user_filter, {"_id": 0}).to_list(1000)
    if report_type in ["investment", "networth"]:
        data["investments"] = await db.investments.find(user_filter, {"_id": 0}).to_list(1000)
    if report_type in ["networth"]:
        data["assets"] = await db.assets.find(user_filter, {"_id": 0}).to_list(1000)
        data["accounts"] = await db.accounts.find(user_filter, {"_id": 0}).to_list(1000)
        data["credit_cards"] = await db.credit_cards.find(user_filter, {"_id": 0}).to_list(1000)
    if report_type == "goal":
        data["goals"] = await db.goals.find(user_filter, {"_id": 0}).to_list(1000)
    if report_type == "asset":
        data["assets"] = await db.assets.find(user_filter, {"_id": 0}).to_list(1000)
    if report_type == "insurance":
        data["insurances"] = await db.insurances.find(user_filter, {"_id": 0}).to_list(1000)

    ctx = {"user_name": user_name, "user_email": user_email, "start_date": start_date, "end_date": end_date}

    if format == "excel":
        return await generate_excel_report(report_type, data, ctx)
    return await generate_pdf_report(report_type, data, ctx)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PDF GENERATION — BANK STATEMENT STYLE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _register_fonts():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    try:
        pdfmetrics.getFont('DejaVu')
    except Exception:
        pdfmetrics.registerFont(TTFont('DejaVu', FONT_PATH))
        pdfmetrics.registerFont(TTFont('DejaVu-Bold', FONT_BOLD_PATH))


def _build_header_block(elements, report_type, ctx):
    """Build the dark navy header banner like a bank statement."""
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch

    meta = REPORT_META.get(report_type, {"title": "Financial Statement", "accent": ACCENT_BLUE, "icon": "REPORT"})
    accent = meta["accent"]

    # Brand header bar
    brand_style = ParagraphStyle('Brand', fontName='DejaVu-Bold', fontSize=18, textColor=colors.white)
    type_style = ParagraphStyle('Type', fontName='DejaVu', fontSize=9, textColor=colors.HexColor('#94A3B8'))
    header_data = [[
        Paragraph("MONEYSSUTRA", brand_style),
        Paragraph(meta["title"].upper(), ParagraphStyle('RT', fontName='DejaVu-Bold', fontSize=12, textColor=colors.HexColor(accent), alignment=2))
    ]]
    header_table = Table(header_data, colWidths=[3.5 * inch, 3.5 * inch])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(NAVY)),
        ('TOPPADDING', (0, 0), (-1, -1), 14),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 14),
        ('LEFTPADDING', (0, 0), (0, -1), 16),
        ('RIGHTPADDING', (-1, 0), (-1, -1), 16),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, 0), 3, colors.HexColor(accent)),
    ]))
    elements.append(header_table)

    # Info row
    info_style = ParagraphStyle('Info', fontName='DejaVu', fontSize=8, textColor=colors.HexColor(MED_GRAY))
    info_bold = ParagraphStyle('InfoBold', fontName='DejaVu-Bold', fontSize=8, textColor=colors.HexColor(DARK_GRAY))
    info_data = [[
        Paragraph(f"<b>Account Holder:</b>  {ctx['user_name']}", info_style),
        Paragraph(f"<b>Statement Period:</b>  {ctx['start_date'].strftime('%d %b %Y')}  to  {ctx['end_date'].strftime('%d %b %Y')}", info_style),
        Paragraph(f"<b>Generated:</b>  {datetime.now().strftime('%d %b %Y, %I:%M %p')}", info_style),
    ]]
    info_table = Table(info_data, colWidths=[2.5 * inch, 2.5 * inch, 2 * inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(LIGHT_GRAY)),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor(BORDER_GRAY)),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 12))


def _build_summary_boxes(elements, boxes):
    """Build the summary metric boxes (like bank statement quick glance)."""
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch

    label_style = ParagraphStyle('SumLabel', fontName='DejaVu', fontSize=7, textColor=colors.HexColor(MED_GRAY))
    value_style = ParagraphStyle('SumVal', fontName='DejaVu-Bold', fontSize=12, textColor=colors.HexColor(DARK_GRAY))

    n = len(boxes)
    col_w = 7.0 / n * inch
    row1 = []
    row2 = []
    for label, value, color in boxes:
        row1.append(Paragraph(label.upper(), label_style))
        row2.append(Paragraph(value, ParagraphStyle('V', fontName='DejaVu-Bold', fontSize=12, textColor=colors.HexColor(color or DARK_GRAY))))

    summary_table = Table([row1, row2], colWidths=[col_w] * n)
    styles = [
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.HexColor(BORDER_GRAY)),
    ]
    # Add vertical separators between columns
    for i in range(1, n):
        styles.append(('LINEBEFORE', (i, 0), (i, -1), 0.5, colors.HexColor(BORDER_GRAY)))
    summary_table.setStyle(TableStyle(styles))
    elements.append(summary_table)

    from reportlab.platypus import Spacer
    elements.append(Spacer(1, 14))


def _build_statement_table(elements, headers, rows, total_row=None, accent_color=NAVY):
    """Build a clean bank-statement-style data table."""
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch

    h_style = ParagraphStyle('TH', fontName='DejaVu-Bold', fontSize=7.5, textColor=colors.HexColor(MED_GRAY))
    cell_style = ParagraphStyle('TD', fontName='DejaVu', fontSize=8, textColor=colors.HexColor(DARK_GRAY))
    cell_bold = ParagraphStyle('TDB', fontName='DejaVu-Bold', fontSize=8, textColor=colors.HexColor(DARK_GRAY))
    amount_style = ParagraphStyle('AMT', fontName='DejaVu-Bold', fontSize=8, textColor=colors.HexColor(DARK_GRAY), alignment=2)
    total_style = ParagraphStyle('TOT', fontName='DejaVu-Bold', fontSize=9, textColor=colors.HexColor(accent_color), alignment=2)
    total_label = ParagraphStyle('TOTL', fontName='DejaVu-Bold', fontSize=9, textColor=colors.HexColor(accent_color))

    n_cols = len(headers)
    total_w = 7.0
    col_widths = [total_w / n_cols * inch] * n_cols
    # Make first column wider, amount columns narrower
    if n_cols >= 3:
        col_widths[0] = 2.2 * inch
        remaining = (total_w - 2.2) / (n_cols - 1)
        for i in range(1, n_cols):
            col_widths[i] = remaining * inch

    # Header row
    header_row = [Paragraph(h.upper(), h_style) for h in headers]
    table_data = [header_row]

    # Data rows
    for r_idx, row in enumerate(rows):
        styled_row = []
        for c_idx, cell in enumerate(row):
            if c_idx == 0:
                styled_row.append(Paragraph(str(cell), cell_bold))
            elif 'amount' in headers[c_idx].lower() or headers[c_idx].lower() in ('invested', 'current', 'outstanding', 'emi', 'premium', 'sum assured', 'gain/loss', 'purchase value', 'current value', 'target', 'progress'):
                styled_row.append(Paragraph(str(cell), amount_style))
            else:
                styled_row.append(Paragraph(str(cell), cell_style))
        table_data.append(styled_row)

    # Total row
    if total_row:
        styled_total = []
        for c_idx, cell in enumerate(total_row):
            if c_idx == 0:
                styled_total.append(Paragraph(str(cell), total_label))
            elif cell and cell != "":
                styled_total.append(Paragraph(str(cell), total_style))
            else:
                styled_total.append(Paragraph("", cell_style))
        table_data.append(styled_total)

    table = Table(table_data, colWidths=col_widths)
    styles = [
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(LIGHT_GRAY)),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor(BORDER_GRAY)),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        # Data rows
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]

    # Alternating row backgrounds
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            styles.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))

    # Thin line under each data row
    for i in range(1, len(table_data)):
        styles.append(('LINEBELOW', (0, i), (-1, i), 0.3, colors.HexColor('#E2E8F0')))

    # Total row styling
    if total_row:
        last = len(table_data) - 1
        styles.append(('LINEABOVE', (0, last), (-1, last), 1.5, colors.HexColor(accent_color)))
        styles.append(('BACKGROUND', (0, last), (-1, last), colors.HexColor('#F0F9FF')))
        styles.append(('TOPPADDING', (0, last), (-1, last), 10))
        styles.append(('BOTTOMPADDING', (0, last), (-1, last), 10))

    table.setStyle(TableStyle(styles))
    elements.append(table)
    elements.append(Spacer(1, 8))


def _build_section_header(elements, title, count=None):
    from reportlab.platypus import Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    label = title
    if count is not None:
        label = f"{title}  <font color='{MED_GRAY}' size=8>({count} entries)</font>"
    style = ParagraphStyle('SH', fontName='DejaVu-Bold', fontSize=10, textColor=colors.HexColor(DARK_GRAY), spaceBefore=12, spaceAfter=6)
    elements.append(Paragraph(label, style))


def _build_footer(elements):
    from reportlab.platypus import Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle('Footer', fontName='DejaVu', fontSize=7, textColor=colors.HexColor('#94A3B8'), alignment=1)
    elements.append(Paragraph("This is a computer-generated financial statement from Moneyssutra. No signature required.", footer_style))
    elements.append(Paragraph(f"Statement ID: MS-{datetime.now().strftime('%Y%m%d%H%M%S')}  |  Confidential", footer_style))


async def generate_pdf_report(report_type, data, ctx):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Spacer

    _register_fonts()
    meta = REPORT_META.get(report_type, {"title": "Financial Statement", "accent": ACCENT_BLUE})
    accent = meta["accent"]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.4 * inch, bottomMargin=0.4 * inch, leftMargin=0.5 * inch, rightMargin=0.5 * inch)
    elements = []

    _build_header_block(elements, report_type, ctx)

    # ── INCOME ──
    if report_type == "income":
        incomes = data.get("incomes", [])
        other_incomes = data.get("other_incomes", [])
        all_items = []
        for inc in incomes:
            amt = inc.get('expectedAmount', 0) or 0
            all_items.append({"name": inc.get('name', 'N/A'), "type": inc.get('type', 'N/A').title(), "amount": amt, "freq": inc.get('frequency', 'Monthly'), "date": fmt_date(inc.get('createdAt') or inc.get('lastEntryDate'))})
        for oi in other_incomes:
            amt = oi.get('amount', 0) or 0
            cat = oi.get('customCategory') or oi.get('category', 'Other')
            all_items.append({"name": oi.get('incomeName', 'N/A'), "type": cat.title(), "amount": amt, "freq": oi.get('frequency', 'One-time'), "date": fmt_date(oi.get('createdAt') or oi.get('dateReceived'))})

        total = sum(i["amount"] for i in all_items)
        by_type = defaultdict(float)
        for i in all_items:
            by_type[i["type"]] += i["amount"]
        top_source = max(all_items, key=lambda x: x["amount"])["name"] if all_items else "-"

        _build_summary_boxes(elements, [
            ("Total Income", fmt_amount(total), ACCENT_GREEN),
            ("Sources", str(len(all_items)), DARK_GRAY),
            ("Top Source", top_source[:18], DARK_GRAY),
            ("Categories", str(len(by_type)), DARK_GRAY),
        ])

        # Group by type
        grouped = defaultdict(list)
        for i in all_items:
            grouped[i["type"]].append(i)

        for cat_name, items in sorted(grouped.items()):
            cat_total = sum(i["amount"] for i in items)
            _build_section_header(elements, f"{cat_name}", count=len(items))
            rows = [[i["name"][:28], i["freq"], fmt_amount(i["amount"]), i["date"]] for i in items]
            _build_statement_table(elements,
                ["Description", "Frequency", "Amount", "Date"],
                rows,
                total_row=[f"Subtotal - {cat_name}", "", fmt_amount(cat_total), ""],
                accent_color=accent)

        # Grand total
        from reportlab.platypus import Paragraph
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors
        gt_style = ParagraphStyle('GT', fontName='DejaVu-Bold', fontSize=11, textColor=colors.HexColor(ACCENT_GREEN), alignment=2, spaceBefore=8)
        elements.append(Paragraph(f"GRAND TOTAL:  {fmt_amount(total)}", gt_style))

    # ── EXPENSE ──
    elif report_type == "expense":
        expenses = data.get("expenses", [])
        total = sum(e.get('expectedAmount', 0) or 0 for e in expenses)
        by_cat = defaultdict(float)
        for e in expenses:
            by_cat[e.get('category', 'Other')] += e.get('expectedAmount', 0) or 0
        top_cat = max(by_cat, key=by_cat.get) if by_cat else "-"

        _build_summary_boxes(elements, [
            ("Total Expenses", fmt_amount(total), ACCENT_RED),
            ("Items", str(len(expenses)), DARK_GRAY),
            ("Top Category", top_cat[:18], DARK_GRAY),
            ("Categories", str(len(by_cat)), DARK_GRAY),
        ])

        grouped = defaultdict(list)
        for e in expenses:
            grouped[e.get('category', 'Other')].append(e)

        for cat_name, items in sorted(grouped.items()):
            cat_total = sum(i.get('expectedAmount', 0) or 0 for i in items)
            _build_section_header(elements, cat_name, count=len(items))
            rows = [[i.get('expenseName', 'N/A')[:28], i.get('frequency', 'Monthly'), fmt_amount(i.get('expectedAmount', 0) or 0), fmt_date(i.get('createdAt'))] for i in items]
            _build_statement_table(elements,
                ["Description", "Frequency", "Amount", "Date"],
                rows,
                total_row=[f"Subtotal - {cat_name}", "", fmt_amount(cat_total), ""],
                accent_color=accent)

        from reportlab.platypus import Paragraph
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors
        gt_style = ParagraphStyle('GT2', fontName='DejaVu-Bold', fontSize=11, textColor=colors.HexColor(ACCENT_RED), alignment=2, spaceBefore=8)
        elements.append(Paragraph(f"GRAND TOTAL:  {fmt_amount(total)}", gt_style))

    # ── CASHFLOW ──
    elif report_type == "cashflow":
        incomes = data.get("incomes", [])
        other_incomes = data.get("other_incomes", [])
        expenses = data.get("expenses", [])
        inc_total = sum(i.get('expectedAmount', 0) or 0 for i in incomes) + sum(o.get('amount', 0) or 0 for o in other_incomes)
        exp_total = sum(e.get('expectedAmount', 0) or 0 for e in expenses)
        net = inc_total - exp_total
        net_color = ACCENT_GREEN if net >= 0 else ACCENT_RED

        _build_summary_boxes(elements, [
            ("Total Inflow", fmt_amount(inc_total), ACCENT_GREEN),
            ("Total Outflow", fmt_amount(exp_total), ACCENT_RED),
            ("Net Cash Flow", fmt_amount(net, show_sign=True), net_color),
            ("Savings Rate", f"{(net / inc_total * 100):.0f}%" if inc_total > 0 else "0%", net_color),
        ])

        if incomes or other_incomes:
            _build_section_header(elements, "Inflows (Income)", count=len(incomes) + len(other_incomes))
            rows = []
            for inc in incomes:
                rows.append([inc.get('name', 'N/A')[:28], inc.get('type', 'N/A').title(), inc.get('frequency', 'Monthly'), fmt_amount(inc.get('expectedAmount', 0) or 0)])
            for oi in other_incomes:
                cat = oi.get('customCategory') or oi.get('category', 'Other')
                rows.append([oi.get('incomeName', 'N/A')[:28], cat.title(), oi.get('frequency', 'One-time'), fmt_amount(oi.get('amount', 0) or 0)])
            _build_statement_table(elements,
                ["Description", "Type", "Frequency", "Amount"],
                rows,
                total_row=["Total Inflows", "", "", fmt_amount(inc_total)],
                accent_color=ACCENT_GREEN)

        if expenses:
            _build_section_header(elements, "Outflows (Expenses)", count=len(expenses))
            rows = [[e.get('expenseName', 'N/A')[:28], e.get('category', 'N/A'), e.get('frequency', 'Monthly'), fmt_amount(e.get('expectedAmount', 0) or 0)] for e in expenses]
            _build_statement_table(elements,
                ["Description", "Category", "Frequency", "Amount"],
                rows,
                total_row=["Total Outflows", "", "", fmt_amount(exp_total)],
                accent_color=ACCENT_RED)

        from reportlab.platypus import Paragraph
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors
        gt_style = ParagraphStyle('GTCF', fontName='DejaVu-Bold', fontSize=11, textColor=colors.HexColor(net_color), alignment=2, spaceBefore=8)
        elements.append(Paragraph(f"NET CASH FLOW:  {fmt_amount(net, show_sign=True)}", gt_style))

    # ── INVESTMENT ──
    elif report_type == "investment":
        investments = data.get("investments", [])
        ti = sum(i.get('principal', 0) or 0 for i in investments)
        tc = sum(i.get('currentValue', 0) or 0 for i in investments)
        gain = tc - ti
        gain_pct = (gain / ti * 100) if ti > 0 else 0
        gain_color = ACCENT_GREEN if gain >= 0 else ACCENT_RED

        _build_summary_boxes(elements, [
            ("Total Invested", fmt_amount(ti), DARK_GRAY),
            ("Current Value", fmt_amount(tc), DARK_GRAY),
            ("Total Returns", fmt_amount(gain, show_sign=True), gain_color),
            ("Return %", f"{gain_pct:+.1f}%", gain_color),
        ])

        _build_section_header(elements, "Portfolio Details", count=len(investments))
        rows = []
        for inv in investments:
            invested = inv.get('principal', 0) or 0
            current = inv.get('currentValue', 0) or 0
            g = current - invested
            rows.append([
                inv.get('name', 'N/A')[:24],
                inv.get('investmentCategory', 'N/A')[:14],
                fmt_amount(invested), fmt_amount(current), fmt_amount(g, show_sign=True),
            ])
        _build_statement_table(elements,
            ["Investment", "Category", "Invested", "Current", "Gain/Loss"],
            rows,
            total_row=["TOTAL", "", fmt_amount(ti), fmt_amount(tc), fmt_amount(gain, show_sign=True)],
            accent_color=accent)

    # ── LOAN ──
    elif report_type == "loan":
        loans = data.get("loans", [])
        total_outstanding = sum(l.get('outstandingAmount', 0) or 0 for l in loans)
        total_emi = sum(l.get('emiAmount', 0) or 0 for l in loans)
        avg_rate = (sum(l.get('interestRate', 0) or 0 for l in loans) / len(loans)) if loans else 0

        _build_summary_boxes(elements, [
            ("Outstanding", fmt_amount(total_outstanding), ACCENT_RED),
            ("Monthly EMI", fmt_amount(total_emi), '#D97706'),
            ("Active Loans", str(len(loans)), DARK_GRAY),
            ("Avg Rate", f"{avg_rate:.1f}%", DARK_GRAY),
        ])

        _build_section_header(elements, "Loan Details", count=len(loans))
        rows = []
        for loan in loans:
            rows.append([
                loan.get('loanName', 'N/A')[:22], loan.get('loanType', 'N/A')[:14],
                fmt_amount(loan.get('outstandingAmount', 0)),
                fmt_amount(loan.get('emiAmount', 0)),
                f"{loan.get('interestRate', 0)}%",
            ])
        _build_statement_table(elements,
            ["Loan", "Type", "Outstanding", "EMI", "Rate"],
            rows,
            total_row=["TOTAL", "", fmt_amount(total_outstanding), fmt_amount(total_emi), ""],
            accent_color=accent)

    # ── NET WORTH ──
    elif report_type == "networth":
        total_assets = sum(a.get('currentValue', 0) or 0 for a in data.get("assets", []))
        total_inv = sum(i.get('currentValue', 0) or 0 for i in data.get("investments", []))
        liquid = sum(a.get('currentBalance', 0) or a.get('balance', 0) or 0 for a in data.get("accounts", []))
        total_loans = sum(ln.get('outstandingAmount', 0) or 0 for ln in data.get("loans", []))
        total_cc = sum(c.get('currentOutstanding', 0) or c.get('outstandingAmount', 0) or 0 for c in data.get("credit_cards", []))
        total_liabilities = total_loans + total_cc
        gross = total_assets + total_inv + liquid
        nw = gross - total_liabilities
        nw_color = ACCENT_GREEN if nw >= 0 else ACCENT_RED

        _build_summary_boxes(elements, [
            ("Gross Assets", fmt_amount(gross), DARK_GRAY),
            ("Liabilities", fmt_amount(total_liabilities), ACCENT_RED),
            ("Net Worth", fmt_amount(nw), nw_color),
        ])

        _build_section_header(elements, "Balance Sheet")
        rows = [
            ["Physical Assets", "", fmt_amount(total_assets)],
            ["Investments", "", fmt_amount(total_inv)],
            ["Bank Balances", "", fmt_amount(liquid)],
            ["", "TOTAL ASSETS", fmt_amount(gross)],
            ["", "", ""],
            ["Loan Outstanding", "", fmt_amount(total_loans)],
            ["Credit Card Due", "", fmt_amount(total_cc)],
            ["", "TOTAL LIABILITIES", fmt_amount(total_liabilities)],
        ]
        _build_statement_table(elements,
            ["Category", "", "Amount"],
            rows,
            total_row=["NET WORTH", "", fmt_amount(nw)],
            accent_color=nw_color)

    # ── GOAL ──
    elif report_type == "goal":
        goals = data.get("goals", [])
        total_target = sum(g.get('targetAmount', 0) or 0 for g in goals)
        total_current = sum(g.get('currentAmount', 0) or 0 for g in goals)
        overall_pct = (total_current / total_target * 100) if total_target > 0 else 0

        _build_summary_boxes(elements, [
            ("Total Target", fmt_amount(total_target), DARK_GRAY),
            ("Achieved", fmt_amount(total_current), ACCENT_GREEN),
            ("Remaining", fmt_amount(total_target - total_current), '#D97706'),
            ("Progress", f"{overall_pct:.0f}%", ACCENT_GREEN if overall_pct >= 50 else '#D97706'),
        ])

        _build_section_header(elements, "Goal Details", count=len(goals))
        rows = []
        for goal in goals:
            target = goal.get('targetAmount', 0) or 0
            current = goal.get('currentAmount', 0) or 0
            pct = (current / target * 100) if target > 0 else 0
            td = goal.get('targetDate', '-')
            if isinstance(td, datetime):
                td = td.strftime('%d %b %Y')
            rows.append([goal.get('goalName', 'N/A')[:22], fmt_amount(target), fmt_amount(current), f"{pct:.0f}%", str(td)[:10] if td else '-'])
        _build_statement_table(elements,
            ["Goal", "Target", "Current", "Progress", "Target Date"],
            rows, accent_color=accent)

    # ── ASSET ──
    elif report_type == "asset":
        assets = data.get("assets", [])
        total_purchase = sum(a.get('purchaseValue', 0) or 0 for a in assets)
        total_current = sum(a.get('currentValue', 0) or 0 for a in assets)
        appreciation = total_current - total_purchase
        app_color = ACCENT_GREEN if appreciation >= 0 else ACCENT_RED

        _build_summary_boxes(elements, [
            ("Purchase Value", fmt_amount(total_purchase), DARK_GRAY),
            ("Current Value", fmt_amount(total_current), DARK_GRAY),
            ("Appreciation", fmt_amount(appreciation, show_sign=True), app_color),
            ("Assets", str(len(assets)), DARK_GRAY),
        ])

        _build_section_header(elements, "Asset Details", count=len(assets))
        rows = []
        for a in assets:
            pv = a.get('purchaseValue', 0) or 0
            cv = a.get('currentValue', 0) or 0
            rows.append([a.get('name', a.get('assetName', 'N/A'))[:22], a.get('assetType', 'N/A'), fmt_amount(pv), fmt_amount(cv)])
        _build_statement_table(elements,
            ["Asset", "Type", "Purchase Value", "Current Value"],
            rows,
            total_row=["TOTAL", "", fmt_amount(total_purchase), fmt_amount(total_current)],
            accent_color=accent)

    # ── INSURANCE ──
    elif report_type == "insurance":
        insurances = data.get("insurances", [])
        total_premium = sum(i.get('premiumAmount', 0) or 0 for i in insurances)
        total_cover = sum(i.get('sumAssured', 0) or 0 for i in insurances)

        _build_summary_boxes(elements, [
            ("Annual Premium", fmt_amount(total_premium), DARK_GRAY),
            ("Total Coverage", fmt_amount(total_cover), ACCENT_GREEN),
            ("Policies", str(len(insurances)), DARK_GRAY),
            ("Cover Ratio", f"{(total_cover / total_premium):.0f}x" if total_premium > 0 else "-", DARK_GRAY),
        ])

        _build_section_header(elements, "Policy Details", count=len(insurances))
        rows = []
        for ins in insurances:
            rows.append([ins.get('policyName', 'N/A')[:22], ins.get('insuranceType', 'N/A'), fmt_amount(ins.get('premiumAmount', 0)), fmt_amount(ins.get('sumAssured', 0))])
        _build_statement_table(elements,
            ["Policy", "Type", "Premium", "Sum Assured"],
            rows,
            total_row=["TOTAL", "", fmt_amount(total_premium), fmt_amount(total_cover)],
            accent_color=accent)

    else:
        from reportlab.platypus import Paragraph
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors
        elements.append(Paragraph("No data available for this report type.", ParagraphStyle('ND', fontName='DejaVu', fontSize=10, textColor=colors.HexColor(MED_GRAY))))

    _build_footer(elements)

    doc.build(elements)
    buffer.seek(0)
    filename = f"Moneyssutra_{report_type.title()}_Statement_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXCEL GENERATION — BANK STATEMENT STYLE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

async def generate_excel_report(report_type, data, ctx):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
    from openpyxl.utils import get_column_letter

    meta = REPORT_META.get(report_type, {"title": "Financial Statement", "accent": ACCENT_BLUE})
    accent_hex = meta["accent"].lstrip('#')

    wb = Workbook()
    ws = wb.active
    ws.title = meta["title"][:31]

    # Styles
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
    data_font = Font(size=10, name='Calibri', color='1E293B')
    bold_font = Font(bold=True, size=10, name='Calibri', color='1E293B')
    amount_font = Font(bold=True, size=10, name='Calibri', color='1E293B')
    summary_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    accent_fill = PatternFill(start_color=accent_hex, end_color=accent_hex, fill_type='solid')
    accent_font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    total_fill = PatternFill(start_color='F0F9FF', end_color='F0F9FF', fill_type='solid')
    total_font = Font(bold=True, size=11, name='Calibri', color=accent_hex)

    row = 1

    # ── Header Block ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    c = ws.cell(row=1, column=1, value="MONEYSSUTRA")
    c.font = Font(bold=True, size=16, name='Calibri', color='0F172A')
    row = 2
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    c = ws.cell(row=2, column=1, value=meta["title"])
    c.font = Font(bold=True, size=12, name='Calibri', color=accent_hex)
    row = 3
    ws.cell(row=3, column=1, value="Account Holder:").font = Font(bold=True, size=9, name='Calibri', color='475569')
    ws.cell(row=3, column=2, value=ctx['user_name']).font = data_font
    ws.cell(row=3, column=4, value="Generated:").font = Font(bold=True, size=9, name='Calibri', color='475569')
    ws.cell(row=3, column=5, value=datetime.now().strftime('%d %b %Y, %I:%M %p')).font = data_font
    row = 4
    ws.cell(row=4, column=1, value="Statement Period:").font = Font(bold=True, size=9, name='Calibri', color='475569')
    ws.cell(row=4, column=2, value=f"{ctx['start_date'].strftime('%d %b %Y')} to {ctx['end_date'].strftime('%d %b %Y')}").font = data_font
    row = 6

    def write_summary(labels_values):
        nonlocal row
        for col_idx, (label, value) in enumerate(labels_values, 1):
            c1 = ws.cell(row=row, column=col_idx, value=label)
            c1.font = Font(bold=True, size=8, name='Calibri', color='475569')
            c1.fill = summary_fill
            c1.alignment = Alignment(horizontal='center')
            c2 = ws.cell(row=row + 1, column=col_idx, value=value)
            c2.font = Font(bold=True, size=11, name='Calibri', color=accent_hex)
            c2.fill = summary_fill
            c2.alignment = Alignment(horizontal='center')
        row += 3

    def write_table_header(headers):
        nonlocal row
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')
        row += 1

    def write_data_row(values, is_alt=False):
        nonlocal row
        for col_idx, v in enumerate(values, 1):
            c = ws.cell(row=row, column=col_idx, value=v)
            c.font = data_font
            c.border = thin_border
            if is_alt:
                c.fill = alt_fill
            # Right-align numbers
            if isinstance(v, (int, float)):
                c.alignment = Alignment(horizontal='right')
                c.number_format = '#,##0'
        row += 1

    def write_total_row(values):
        nonlocal row
        for col_idx, v in enumerate(values, 1):
            c = ws.cell(row=row, column=col_idx, value=v)
            c.font = total_font
            c.fill = total_fill
            c.border = thin_border
            if isinstance(v, (int, float)):
                c.alignment = Alignment(horizontal='right')
                c.number_format = '#,##0'
        row += 1

    def write_section_label(text):
        nonlocal row
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=10, name='Calibri', color='1E293B')
        row += 1

    # ── INCOME ──
    if report_type == "income":
        all_items = []
        for inc in data.get("incomes", []):
            amt = inc.get('expectedAmount', 0) or 0
            all_items.append(("Regular", inc.get('name', 'N/A'), inc.get('type', 'N/A').title(), amt, inc.get('frequency', 'Monthly'), fmt_date(inc.get('createdAt'))))
        for oi in data.get("other_incomes", []):
            amt = oi.get('amount', 0) or 0
            cat = oi.get('customCategory') or oi.get('category', 'Other')
            all_items.append(("Other", oi.get('incomeName', 'N/A'), cat.title(), amt, oi.get('frequency', 'One-time'), fmt_date(oi.get('createdAt'))))
        total = sum(i[3] for i in all_items)
        write_summary([("TOTAL INCOME", fmt_amount(total)), ("SOURCES", str(len(all_items))), ("CATEGORIES", str(len(set(i[2] for i in all_items))))])
        write_table_header(["#", "Source", "Category", "Amount", "Frequency", "Date"])
        for idx, (src_type, name, cat, amt, freq, dt) in enumerate(all_items, 1):
            write_data_row([idx, name, cat, amt, freq, dt], idx % 2 == 0)
        write_total_row(["", "TOTAL", "", total, "", ""])

    # ── EXPENSE ──
    elif report_type == "expense":
        expenses = data.get("expenses", [])
        total = sum(e.get('expectedAmount', 0) or 0 for e in expenses)
        write_summary([("TOTAL EXPENSES", fmt_amount(total)), ("ITEMS", str(len(expenses)))])
        write_table_header(["#", "Expense", "Category", "Amount", "Frequency", "Date"])
        for idx, exp in enumerate(expenses, 1):
            amt = exp.get('expectedAmount', 0) or 0
            write_data_row([idx, exp.get('expenseName', 'N/A'), exp.get('category', 'N/A'), amt, exp.get('frequency', 'Monthly'), fmt_date(exp.get('createdAt'))], idx % 2 == 0)
        write_total_row(["", "TOTAL", "", total, "", ""])

    # ── CASHFLOW ──
    elif report_type == "cashflow":
        incomes = data.get("incomes", [])
        other_incomes = data.get("other_incomes", [])
        expenses = data.get("expenses", [])
        inc_total = sum(i.get('expectedAmount', 0) or 0 for i in incomes) + sum(o.get('amount', 0) or 0 for o in other_incomes)
        exp_total = sum(e.get('expectedAmount', 0) or 0 for e in expenses)
        net = inc_total - exp_total
        write_summary([("INFLOWS", fmt_amount(inc_total)), ("OUTFLOWS", fmt_amount(exp_total)), ("NET CASH FLOW", fmt_amount(net))])

        write_section_label("INFLOWS")
        write_table_header(["#", "Source", "Type", "Amount", "Frequency"])
        idx = 1
        for inc in incomes:
            write_data_row([idx, inc.get('name', 'N/A'), inc.get('type', 'N/A').title(), inc.get('expectedAmount', 0) or 0, inc.get('frequency', 'Monthly')], idx % 2 == 0)
            idx += 1
        for oi in other_incomes:
            cat = oi.get('customCategory') or oi.get('category', 'Other')
            write_data_row([idx, oi.get('incomeName', 'N/A'), cat.title(), oi.get('amount', 0) or 0, oi.get('frequency', 'One-time')], idx % 2 == 0)
            idx += 1
        write_total_row(["", "TOTAL INFLOWS", "", inc_total, ""])

        write_section_label("OUTFLOWS")
        write_table_header(["#", "Expense", "Category", "Amount", "Frequency"])
        for idx, exp in enumerate(expenses, 1):
            write_data_row([idx, exp.get('expenseName', 'N/A'), exp.get('category', 'N/A'), exp.get('expectedAmount', 0) or 0, exp.get('frequency', 'Monthly')], idx % 2 == 0)
        write_total_row(["", "TOTAL OUTFLOWS", "", exp_total, ""])
        row += 1
        write_total_row(["", "NET CASH FLOW", "", net, ""])

    # ── INVESTMENT ──
    elif report_type == "investment":
        investments = data.get("investments", [])
        ti = sum(i.get('principal', 0) or 0 for i in investments)
        tc = sum(i.get('currentValue', 0) or 0 for i in investments)
        write_summary([("INVESTED", fmt_amount(ti)), ("CURRENT VALUE", fmt_amount(tc)), ("RETURNS", fmt_amount(tc - ti))])
        write_table_header(["#", "Investment", "Category", "Invested", "Current Value", "Gain/Loss"])
        for idx, inv in enumerate(investments, 1):
            invested = inv.get('principal', 0) or 0
            current = inv.get('currentValue', 0) or 0
            write_data_row([idx, inv.get('name', 'N/A'), inv.get('investmentCategory', 'N/A'), invested, current, current - invested], idx % 2 == 0)
        write_total_row(["", "TOTAL", "", ti, tc, tc - ti])

    # ── LOAN ──
    elif report_type == "loan":
        loans = data.get("loans", [])
        total_out = sum(l.get('outstandingAmount', 0) or 0 for l in loans)
        total_emi = sum(l.get('emiAmount', 0) or 0 for l in loans)
        write_summary([("OUTSTANDING", fmt_amount(total_out)), ("MONTHLY EMI", fmt_amount(total_emi)), ("ACTIVE LOANS", str(len(loans)))])
        write_table_header(["#", "Loan", "Type", "Outstanding", "EMI", "Rate"])
        for idx, loan in enumerate(loans, 1):
            write_data_row([idx, loan.get('loanName', 'N/A'), loan.get('loanType', 'N/A'), loan.get('outstandingAmount', 0) or 0, loan.get('emiAmount', 0) or 0, f"{loan.get('interestRate', 0)}%"], idx % 2 == 0)
        write_total_row(["", "TOTAL", "", total_out, total_emi, ""])

    # ── NET WORTH ──
    elif report_type == "networth":
        total_assets = sum(a.get('currentValue', 0) or 0 for a in data.get("assets", []))
        total_inv = sum(i.get('currentValue', 0) or 0 for i in data.get("investments", []))
        liquid = sum(a.get('currentBalance', 0) or a.get('balance', 0) or 0 for a in data.get("accounts", []))
        total_loans = sum(ln.get('outstandingAmount', 0) or 0 for ln in data.get("loans", []))
        total_cc = sum(c.get('currentOutstanding', 0) or c.get('outstandingAmount', 0) or 0 for c in data.get("credit_cards", []))
        gross = total_assets + total_inv + liquid
        nw = gross - total_loans - total_cc
        write_summary([("GROSS ASSETS", fmt_amount(gross)), ("LIABILITIES", fmt_amount(total_loans + total_cc)), ("NET WORTH", fmt_amount(nw))])
        write_table_header(["Category", "Amount"])
        write_data_row(["Physical Assets", total_assets])
        write_data_row(["Investments", total_inv], True)
        write_data_row(["Bank Balances", liquid])
        write_total_row(["TOTAL ASSETS", gross])
        row += 1
        write_data_row(["Loan Outstanding", total_loans])
        write_data_row(["Credit Card Due", total_cc], True)
        write_total_row(["TOTAL LIABILITIES", total_loans + total_cc])
        row += 1
        write_total_row(["NET WORTH", nw])

    # ── GOAL ──
    elif report_type == "goal":
        goals = data.get("goals", [])
        total_target = sum(g.get('targetAmount', 0) or 0 for g in goals)
        total_current = sum(g.get('currentAmount', 0) or 0 for g in goals)
        write_summary([("TARGET", fmt_amount(total_target)), ("ACHIEVED", fmt_amount(total_current)), ("REMAINING", fmt_amount(total_target - total_current))])
        write_table_header(["#", "Goal", "Target", "Current", "Progress", "Target Date"])
        for idx, goal in enumerate(goals, 1):
            target = goal.get('targetAmount', 0) or 0
            current = goal.get('currentAmount', 0) or 0
            pct = (current / target * 100) if target > 0 else 0
            td = goal.get('targetDate', '-')
            if isinstance(td, datetime):
                td = td.strftime('%d %b %Y')
            write_data_row([idx, goal.get('goalName', 'N/A'), target, current, f"{pct:.0f}%", str(td)[:10] if td else '-'], idx % 2 == 0)

    # ── ASSET ──
    elif report_type == "asset":
        assets = data.get("assets", [])
        total_pv = sum(a.get('purchaseValue', 0) or 0 for a in assets)
        total_cv = sum(a.get('currentValue', 0) or 0 for a in assets)
        write_summary([("PURCHASE VALUE", fmt_amount(total_pv)), ("CURRENT VALUE", fmt_amount(total_cv)), ("APPRECIATION", fmt_amount(total_cv - total_pv))])
        write_table_header(["#", "Asset", "Type", "Purchase Value", "Current Value"])
        for idx, a in enumerate(assets, 1):
            write_data_row([idx, a.get('name', a.get('assetName', 'N/A')), a.get('assetType', 'N/A'), a.get('purchaseValue', 0) or 0, a.get('currentValue', 0) or 0], idx % 2 == 0)
        write_total_row(["", "TOTAL", "", total_pv, total_cv])

    # ── INSURANCE ──
    elif report_type == "insurance":
        insurances = data.get("insurances", [])
        total_premium = sum(i.get('premiumAmount', 0) or 0 for i in insurances)
        total_cover = sum(i.get('sumAssured', 0) or 0 for i in insurances)
        write_summary([("ANNUAL PREMIUM", fmt_amount(total_premium)), ("TOTAL COVERAGE", fmt_amount(total_cover)), ("POLICIES", str(len(insurances)))])
        write_table_header(["#", "Policy", "Type", "Premium", "Sum Assured"])
        for idx, ins in enumerate(insurances, 1):
            write_data_row([idx, ins.get('policyName', 'N/A'), ins.get('insuranceType', 'N/A'), ins.get('premiumAmount', 0) or 0, ins.get('sumAssured', 0) or 0], idx % 2 == 0)
        write_total_row(["", "TOTAL", "", total_premium, total_cover])

    # Footer disclaimer
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value="This is a computer-generated financial statement from Moneyssutra. No signature required.")
    c.font = Font(italic=True, size=8, name='Calibri', color='94A3B8')

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_len + 3, 35)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Moneyssutra_{report_type.title()}_Statement_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})
