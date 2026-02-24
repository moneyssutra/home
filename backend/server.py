from fastapi import FastAPI, APIRouter, HTTPException, Response, Request, Cookie
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import httpx
import hashlib
import asyncio
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta

# Configure logging
logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# Define Models
class StatusCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")  # Ignore MongoDB's _id field
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StatusCheckCreate(BaseModel):
    client_name: str

# ============ AUTH MODELS ============
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    user_id: str = Field(default_factory=lambda: f"user_{uuid.uuid4().hex[:12]}")
    email: str
    name: str
    picture: Optional[str] = None
    auth_type: str = "google"  # "google" or "jwt"
    password_hash: Optional[str] = None  # For JWT auth only
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserSession(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    session_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    session_token: str
    expires_at: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class JWTLoginRequest(BaseModel):
    username: str
    password: str
    remember_me: bool = False

class GoogleSessionRequest(BaseModel):
    session_id: str

# ============ WORKSPACE MODELS ============
class Workspace(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: f"ws_{uuid.uuid4().hex[:12]}")
    name: str
    type: str = "Personal"  # "Personal" or "Business"
    owner_id: str  # User ID of the owner
    invite_code: Optional[str] = None  # For sharing via code
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class WorkspaceCreate(BaseModel):
    name: str
    type: str = "Personal"

class WorkspaceMember(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: f"wm_{uuid.uuid4().hex[:12]}")
    workspace_id: str
    user_id: str
    user_email: str
    user_name: str
    role: str = "viewer"  # "owner", "admin", "editor", "viewer"
    invited_by: Optional[str] = None  # User ID of inviter
    invited_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    accepted_at: Optional[datetime] = None
    status: str = "pending"  # "pending", "active", "removed"

class WorkspaceInvite(BaseModel):
    email: str
    role: str = "viewer"

class WorkspaceInviteByCode(BaseModel):
    invite_code: str

# Permission levels for roles
ROLE_PERMISSIONS = {
    "owner": {"view": True, "add": True, "edit": True, "delete": True, "invite": True},
    "admin": {"view": True, "add": True, "edit": True, "delete": False, "invite": False},
    "editor": {"view": True, "add": True, "edit": True, "delete": False, "invite": False},
    "viewer": {"view": True, "add": False, "edit": False, "delete": False, "invite": False}
}

class IncomeSource(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: Optional[str] = None  # User isolation
    type: str
    name: str
    expectedAmount: float
    frequency: str
    selectedDay: Optional[str] = None
    selectedDate: Optional[str] = None
    selectedQuarter: Optional[str] = None
    selectedHalf: Optional[str] = None
    selectedMonth: Optional[str] = None
    customFrequency: Optional[str] = None
    customDate: Optional[str] = None
    # Interest-specific fields
    principal: Optional[float] = None
    rate: Optional[float] = None
    interestType: Optional[str] = None
    compoundingFrequency: Optional[str] = None
    manualOverride: Optional[bool] = None
    startDate: Optional[str] = None
    endDate: Optional[str] = None
    currentAmount: Optional[float] = None
    # Rental-specific fields
    tenantName: Optional[str] = None
    # Rental with Asset link
    assetId: Optional[str] = None
    securityDeposit: Optional[float] = None
    # Commission-specific fields
    isVariable: Optional[bool] = None
    # Dividend-specific fields
    sourceCategory: Optional[str] = None
    units: Optional[float] = None
    # Self-Employed specific fields
    profession: Optional[str] = None
    # Variable Income fields
    incomeType: Optional[str] = "fixed"  # "fixed" or "variable"
    lastRecordedAmount: Optional[float] = None
    reminderTime: Optional[str] = None  # HH:MM format (e.g., "19:00")
    lastEntryDate: Optional[str] = None  # Last date when entry was recorded
    nextDueDate: Optional[str] = None  # Next expected due date for variable income
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class IncomeSourceCreate(BaseModel):
    type: str
    name: str
    expectedAmount: float
    frequency: str
    selectedDay: Optional[str] = None
    selectedDate: Optional[str] = None
    selectedQuarter: Optional[str] = None
    selectedHalf: Optional[str] = None
    selectedMonth: Optional[str] = None
    customFrequency: Optional[str] = None
    customDate: Optional[str] = None
    # Interest-specific fields
    principal: Optional[float] = None
    rate: Optional[float] = None
    interestType: Optional[str] = None
    compoundingFrequency: Optional[str] = None
    manualOverride: Optional[bool] = None
    startDate: Optional[str] = None
    endDate: Optional[str] = None
    currentAmount: Optional[float] = None
    # Rental-specific fields
    tenantName: Optional[str] = None
    # Rental with Asset link
    assetId: Optional[str] = None
    securityDeposit: Optional[float] = None
    # Commission-specific fields
    isVariable: Optional[bool] = None
    # Dividend-specific fields
    sourceCategory: Optional[str] = None
    units: Optional[float] = None
    # Self-Employed specific fields
    profession: Optional[str] = None
    # Variable Income fields
    incomeType: Optional[str] = "fixed"  # "fixed" or "variable"
    lastRecordedAmount: Optional[float] = None
    reminderTime: Optional[str] = None  # HH:MM format
    lastEntryDate: Optional[str] = None
    nextDueDate: Optional[str] = None

# Account Model
class Account(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: Optional[str] = None  # User isolation
    accountName: str
    accountType: str
    currentBalance: float = 0
    openingBalance: float = 0
    accountNumber: Optional[str] = None
    isPrimary: bool = False
    notes: Optional[str] = None
    # Credit Card specific
    creditLimit: Optional[float] = None
    outstandingAmount: Optional[float] = None
    dueDate: Optional[str] = None
    minimumDue: Optional[float] = None
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AccountCreate(BaseModel):
    accountName: str
    accountType: str
    currentBalance: float = 0
    openingBalance: float = 0
    accountNumber: Optional[str] = None
    isPrimary: bool = False
    notes: Optional[str] = None
    creditLimit: Optional[float] = None
    outstandingAmount: Optional[float] = None
    dueDate: Optional[str] = None
    minimumDue: Optional[float] = None

# Expense Model
class Expense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: Optional[str] = None  # User isolation
    expenseName: str
    expenseType: str  # Fixed or Variable
    category: str
    expectedAmount: float
    frequency: str
    linkedAccountId: Optional[str] = None
    linkedLoanId: Optional[str] = None
    linkedInsuranceId: Optional[str] = None
    linkedInvestmentId: Optional[str] = None  # Link to SIP investment
    selectedDay: Optional[str] = None
    selectedDate: Optional[str] = None
    selectedQuarter: Optional[str] = None
    selectedHalf: Optional[str] = None
    selectedMonth: Optional[str] = None
    oneTimeDate: Optional[str] = None
    isPaid: bool = False
    lastPaidDate: Optional[str] = None
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ExpenseCreate(BaseModel):
    expenseName: str
    expenseType: str
    category: str
    expectedAmount: float
    frequency: str
    linkedAccountId: Optional[str] = None
    linkedLoanId: Optional[str] = None
    linkedInsuranceId: Optional[str] = None
    linkedInvestmentId: Optional[str] = None
    selectedDay: Optional[str] = None
    selectedDate: Optional[str] = None
    selectedQuarter: Optional[str] = None
    selectedHalf: Optional[str] = None
    selectedMonth: Optional[str] = None
    oneTimeDate: Optional[str] = None
    isPaid: bool = False
    lastPaidDate: Optional[str] = None

# Insurance Model
class Insurance(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: Optional[str] = None  # User isolation
    insuranceType: str
    policyName: str
    coverageAmount: float
    premiumAmount: float
    premiumFrequency: str
    startDate: str
    endDate: Optional[str] = None
    premiumPaymentDate: Optional[str] = None  # Next/first premium due date
    linkedAssetId: Optional[str] = None
    coveredPerson: Optional[str] = None
    linkedExpenseId: Optional[str] = None
    maturityType: Optional[str] = None
    expectedMaturityAmount: Optional[float] = None
    autoCreateExpense: bool = False
    premiumEndDate: Optional[str] = None
    notes: Optional[str] = None
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InsuranceCreate(BaseModel):
    insuranceType: str
    policyName: str
    coverageAmount: float
    premiumAmount: float
    premiumFrequency: str
    startDate: str
    endDate: Optional[str] = None
    premiumPaymentDate: Optional[str] = None  # Next/first premium due date
    linkedAssetId: Optional[str] = None
    coveredPerson: Optional[str] = None
    maturityType: Optional[str] = None
    expectedMaturityAmount: Optional[float] = None
    autoCreateExpense: bool = False
    premiumEndDate: Optional[str] = None
    notes: Optional[str] = None

# Loan Model
class Loan(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: Optional[str] = None  # User isolation
    loanType: Optional[str] = None
    loanName: str
    lenderName: Optional[str] = None
    principalAmount: float
    outstandingAmount: float
    interestRate: float
    emiAmount: float
    emiFrequency: str = "Monthly"
    tenureMonths: Optional[int] = None
    startDate: str
    endDate: Optional[str] = None
    linkedAssetId: Optional[str] = None
    linkedAccountId: Optional[str] = None
    autoCreateExpense: bool = True
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LoanCreate(BaseModel):
    loanType: Optional[str] = None
    loanName: str
    lenderName: Optional[str] = None
    principalAmount: float
    outstandingAmount: float
    interestRate: float
    emiAmount: float
    emiFrequency: str = "Monthly"
    tenureMonths: Optional[int] = None
    startDate: str
    endDate: Optional[str] = None
    linkedAssetId: Optional[str] = None
    linkedAccountId: Optional[str] = None
    autoCreateExpense: bool = True

# Asset Model
class Asset(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: Optional[str] = None  # User isolation
    assetType: str
    assetName: str
    purchaseValue: Optional[float] = None
    currentValue: float
    purchaseDate: Optional[str] = None
    depreciationType: Optional[str] = None  # Appreciating, Depreciating, Market Driven
    isFinanced: bool = False
    linkedLoanId: Optional[str] = None
    generatesIncome: bool = False
    linkedIncomeId: Optional[str] = None
    incomeAmount: Optional[float] = None
    incomeFrequency: Optional[str] = None
    renterName: Optional[str] = None
    securityDeposit: Optional[float] = None
    isInsured: bool = False
    linkedInsuranceId: Optional[str] = None
    location: Optional[str] = None
    notes: Optional[str] = None
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AssetCreate(BaseModel):
    assetType: str
    assetName: str
    purchaseValue: Optional[float] = None
    currentValue: float
    purchaseDate: Optional[str] = None
    depreciationType: Optional[str] = None
    isFinanced: bool = False
    linkedLoanId: Optional[str] = None
    generatesIncome: bool = False
    linkedIncomeId: Optional[str] = None
    incomeAmount: Optional[float] = None
    incomeFrequency: Optional[str] = None
    renterName: Optional[str] = None
    securityDeposit: Optional[float] = None
    isInsured: bool = False
    linkedInsuranceId: Optional[str] = None
    location: Optional[str] = None
    notes: Optional[str] = None

# Investment Model
class Investment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: Optional[str] = None  # User isolation
    investmentCategory: str
    investmentMode: str
    name: str
    principal: float
    currentValue: float
    startDate: str
    linkedAccountId: Optional[str] = None
    notes: Optional[str] = None
    quantity: Optional[float] = None
    unitPrice: Optional[float] = None
    currentPrice: Optional[float] = None
    returnRate: Optional[float] = None
    compoundingType: Optional[str] = None
    compoundingFrequency: Optional[str] = None
    payoutFrequency: Optional[str] = None
    maturityDate: Optional[str] = None
    expectedMaturityValue: Optional[float] = None
    lockInPeriod: Optional[int] = None
    investmentFrequency: Optional[str] = None  # For SIP: Weekly, Monthly, Quarterly, Yearly
    sipAmount: Optional[float] = None  # SIP amount per frequency
    sipSelectedDay: Optional[str] = None  # For weekly: Mon, Tue, etc.
    sipSelectedDate: Optional[str] = None  # Full date string for monthly/quarterly/yearly
    sipSelectedQuarter: Optional[str] = None  # For quarterly: Q1, Q2, etc.
    sipSelectedHalf: Optional[str] = None  # For half-yearly: H1, H2
    sipSelectedMonth: Optional[str] = None  # For quarterly/half-yearly/yearly: January, etc.
    autoCreateExpense: bool = False  # Auto create linked recurring expense
    isLiquidAsset: bool = False  # Consider for emergency fund
    linkedExpenseId: Optional[str] = None  # Link to auto-created expense
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InvestmentCreate(BaseModel):
    investmentCategory: str
    investmentMode: str
    name: str
    principal: float
    currentValue: float
    startDate: str
    linkedAccountId: Optional[str] = None
    notes: Optional[str] = None
    quantity: Optional[float] = None
    unitPrice: Optional[float] = None
    currentPrice: Optional[float] = None
    returnRate: Optional[float] = None
    compoundingType: Optional[str] = None
    compoundingFrequency: Optional[str] = None
    payoutFrequency: Optional[str] = None
    maturityDate: Optional[str] = None
    expectedMaturityValue: Optional[float] = None
    lockInPeriod: Optional[int] = None
    investmentFrequency: Optional[str] = None
    sipAmount: Optional[float] = None
    sipSelectedDay: Optional[str] = None
    sipSelectedDate: Optional[str] = None
    sipSelectedQuarter: Optional[str] = None
    sipSelectedHalf: Optional[str] = None
    sipSelectedMonth: Optional[str] = None
    autoCreateExpense: bool = False
    isLiquidAsset: bool = False

# Credit Card Model
class CreditCard(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: Optional[str] = None  # User isolation
    cardName: str
    bankName: str
    creditLimit: float
    outstandingAmount: float = 0
    billingDate: Optional[int] = None
    dueDate: Optional[int] = None
    minimumDue: Optional[float] = None
    interestRate: Optional[float] = None
    linkedAccountId: Optional[str] = None
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CreditCardCreate(BaseModel):
    cardName: str
    bankName: str
    creditLimit: float
    outstandingAmount: float = 0
    billingDate: Optional[int] = None
    dueDate: Optional[int] = None
    minimumDue: Optional[float] = None
    interestRate: Optional[float] = None
    linkedAccountId: Optional[str] = None

# Goal Model
class Goal(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: Optional[str] = None  # User isolation
    goalName: str
    goalType: str  # Wealth Creation, Debt Elimination, Investment Target, Emergency Fund, Other
    customTypeName: Optional[str] = None  # For "Other" type
    targetAmount: float
    currentAmount: float = 0
    targetDate: str  # ISO date string
    linkedInvestmentIds: List[str] = []  # Deprecated - kept for backward compatibility
    linkedInvestments: List[dict] = []  # New: [{id: str, allocatedAmount: float, name: str}]
    linkedLoanId: Optional[str] = None
    linkedCreditCardId: Optional[str] = None
    linkedAccountIds: List[str] = []  # Deprecated - kept for backward compatibility
    linkedAccounts: List[dict] = []  # New: [{id: str, allocatedAmount: float, name: str}]
    autoCalculate: bool = True  # Auto-calculate from linked sources
    manualOverride: bool = False  # User has manually set currentAmount
    priority: int = 1  # 1 = High, 2 = Medium, 3 = Low
    notes: Optional[str] = None
    isCompleted: bool = False
    completedDate: Optional[str] = None
    reachedMilestones: List[int] = []  # Track milestones reached: [25, 50, 75, 100]
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class GoalCreate(BaseModel):
    goalName: str
    goalType: str
    customTypeName: Optional[str] = None
    targetAmount: float
    currentAmount: float = 0
    targetDate: str
    linkedInvestmentIds: List[str] = []  # Deprecated
    linkedInvestments: List[dict] = []  # New: [{id, allocatedAmount, name}]
    linkedLoanId: Optional[str] = None
    linkedCreditCardId: Optional[str] = None
    linkedAccountIds: List[str] = []  # Deprecated
    linkedAccounts: List[dict] = []  # New: [{id, allocatedAmount, name}]
    autoCalculate: bool = True
    manualOverride: bool = False
    priority: int = 1
    notes: Optional[str] = None
    reachedMilestones: List[int] = []

# Other Income Model (Non-recurring income: gifts, bonuses, capital gains, etc.)
class OtherIncome(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: Optional[str] = None  # User isolation
    incomeName: str
    category: str  # Gift, Bonus, Incentive, Capital Gain, Asset Sale, Tax Refund, Cashback / Reward, Reimbursement, Freelance / Side Work, Windfall, Refund, Miscellaneous, Other
    customCategory: Optional[str] = None  # For "Other" category
    amount: float
    frequency: str  # One-Time, Monthly, Quarterly, Yearly, Irregular
    dateReceived: Optional[str] = None  # ISO date string for one-time or specific date
    selectedDay: Optional[str] = None  # For weekly
    selectedDate: Optional[str] = None  # Day of month (1-31)
    selectedMonth: Optional[str] = None  # For yearly
    selectedQuarter: Optional[str] = None  # For quarterly
    notes: Optional[str] = None
    isReceived: bool = False  # Track if already received
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class OtherIncomeCreate(BaseModel):
    incomeName: str
    category: str
    customCategory: Optional[str] = None
    amount: float
    frequency: str
    dateReceived: Optional[str] = None
    selectedDay: Optional[str] = None
    selectedDate: Optional[str] = None
    selectedMonth: Optional[str] = None
    selectedQuarter: Optional[str] = None
    notes: Optional[str] = None
    isReceived: bool = False

# ============ NOTIFICATION MODEL ============
class Notification(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: str
    title: str
    message: str
    type: str = "income_reminder"  # income_reminder, auto_entry, system
    relatedIncomeId: Optional[str] = None
    relatedIncomeName: Optional[str] = None
    isRead: bool = False
    actionUrl: Optional[str] = None  # URL to navigate when clicked
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NotificationCreate(BaseModel):
    userId: str
    title: str
    message: str
    type: str = "income_reminder"
    relatedIncomeId: Optional[str] = None
    relatedIncomeName: Optional[str] = None
    actionUrl: Optional[str] = None


# ============ INCOME TRANSACTION MODEL ============
class IncomeTransaction(BaseModel):
    """
    Immutable record of an income entry. Each time income is recorded,
    a new transaction is created rather than updating the income source.
    """
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: str
    entityId: str  # Reference to income_source.id (the template)
    entityType: str  # Business, Job, Rental, Interest, Dividend, etc.
    entityName: str  # Name of the income source for quick reference
    amount: float
    transactionDate: str  # YYYY-MM-DD - the date the income occurred
    notes: Optional[str] = None
    source: str = "manual"  # "manual", "auto_fallback", "scheduled"
    isLocked: bool = False  # Becomes True after 24 hours, cannot be deleted
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class IncomeTransactionCreate(BaseModel):
    entityId: str  # Reference to income source
    amount: float
    transactionDate: str  # YYYY-MM-DD
    notes: Optional[str] = None

class ExpenseTransaction(BaseModel):
    """
    Immutable record of an expense entry.
    """
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: str
    entityId: str  # Reference to expense.id (the template)
    entityName: str  # Name of the expense for quick reference
    category: str
    amount: float
    transactionDate: str  # YYYY-MM-DD
    notes: Optional[str] = None
    source: str = "manual"  # "manual", "auto_scheduled"
    isLocked: bool = False
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ExpenseTransactionCreate(BaseModel):
    entityId: str
    amount: float
    transactionDate: str
    notes: Optional[str] = None

# ============ PUSH SUBSCRIPTION MODEL ============
class PushSubscription(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: str
    endpoint: str
    keys: dict  # {p256dh, auth}
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Add your routes to the router instead of directly to app
@api_router.get("/")
async def root():
    return {"message": "Hello World"}

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.model_dump()
    status_obj = StatusCheck(**status_dict)
    
    # Convert to dict and serialize datetime to ISO string for MongoDB
    doc = status_obj.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    
    _ = await db.status_checks.insert_one(doc)
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    # Exclude MongoDB's _id field from the query results
    status_checks = await db.status_checks.find({}, {"_id": 0}).to_list(1000)
    
    # Convert ISO string timestamps back to datetime objects
    for check in status_checks:
        if isinstance(check['timestamp'], str):
            check['timestamp'] = datetime.fromisoformat(check['timestamp'])
    
    return status_checks

# ============ AUTH ENDPOINTS ============

class RegisterRequest(BaseModel):
    firstName: str
    middleName: Optional[str] = None
    lastName: str
    email: str
    mobile: Optional[str] = None
    sex: str  # "male" or "female"
    dateOfBirth: str  # YYYY-MM-DD format
    password: str

def hash_password(password: str) -> str:
    """Hash password using SHA-256"""
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password: str, hashed: str) -> bool:
    """Verify password against hash"""
    return hash_password(password) == hashed

def validate_password_strength(password: str) -> tuple[bool, str]:
    """
    Validate password strength:
    - Min 8 chars
    - At least 1 uppercase
    - At least 1 number
    - At least 1 special character
    Returns (is_valid, error_message)
    """
    if len(password) < 8:
        return False, "Password must be at least 8 characters"
    if not any(c.isupper() for c in password):
        return False, "Password must contain at least 1 uppercase letter"
    if not any(c.isdigit() for c in password):
        return False, "Password must contain at least 1 number"
    if not any(c in "!@#$%^&*()_+-=[]{}|;':\",./<>?" for c in password):
        return False, "Password must contain at least 1 special character"
    return True, ""

@api_router.post("/auth/register")
async def register_user(request: RegisterRequest, response: Response):
    """Register a new user with comprehensive profile data"""
    import re
    
    # Normalize inputs
    email = request.email.strip().lower()
    firstName = request.firstName.strip()
    middleName = request.middleName.strip() if request.middleName else ""
    lastName = request.lastName.strip()
    fullName = f"{firstName} {middleName} {lastName}".replace("  ", " ").strip()
    
    # Validate required fields
    if not firstName:
        raise HTTPException(status_code=400, detail="First name is required")
    if not lastName:
        raise HTTPException(status_code=400, detail="Last name is required")
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")
    if not request.sex or request.sex.lower() not in ["male", "female"]:
        raise HTTPException(status_code=400, detail="Please select your sex (Male/Female)")
    if not request.dateOfBirth:
        raise HTTPException(status_code=400, detail="Date of birth is required")
    
    # Validate name format (no numbers or special characters)
    name_pattern = re.compile(r'^[A-Za-z\s]+$')
    if not name_pattern.match(firstName):
        raise HTTPException(status_code=400, detail="First name should contain only letters")
    if not name_pattern.match(lastName):
        raise HTTPException(status_code=400, detail="Last name should contain only letters")
    if middleName and not name_pattern.match(middleName):
        raise HTTPException(status_code=400, detail="Middle name should contain only letters")
    
    # Validate mobile (if provided) - must be exactly 10 digits
    if request.mobile:
        mobile = request.mobile.strip()
        if not re.match(r'^\d{10}$', mobile):
            raise HTTPException(status_code=400, detail="Mobile number must be exactly 10 digits")
    
    # Validate date of birth (no future dates)
    try:
        dob = datetime.fromisoformat(request.dateOfBirth)
        # Make dob timezone-aware for comparison
        if dob.tzinfo is None:
            dob = dob.replace(tzinfo=timezone.utc)
        if dob > datetime.now(timezone.utc):
            raise HTTPException(status_code=400, detail="Date of birth cannot be in the future")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Validate password strength
    is_valid, error_msg = validate_password_strength(request.password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Check if email already exists (case-insensitive)
    existing_email = await db.users.find_one(
        {"email": {"$regex": f"^{email}$", "$options": "i"}}, 
        {"_id": 0}
    )
    if existing_email:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Check if full name already exists (case-insensitive)
    existing_name = await db.users.find_one(
        {"name": {"$regex": f"^{fullName}$", "$options": "i"}},
        {"_id": 0}
    )
    if existing_name:
        raise HTTPException(status_code=400, detail="This name is already registered")
    
    # Create new user
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    user = {
        "user_id": user_id,
        "email": email,
        "name": fullName,
        "firstName": firstName,
        "middleName": middleName,
        "lastName": lastName,
        "picture": None,
        "auth_type": "jwt",
        "password_hash": hash_password(request.password),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user)
    
    # Create BasicProfile entry
    basic_profile = {
        "user_id": user_id,
        "firstName": firstName,
        "middleName": middleName,
        "lastName": lastName,
        "fullName": fullName,
        "email": email,
        "mobile": request.mobile.strip() if request.mobile else None,
        "sex": request.sex.lower(),
        "dateOfBirth": request.dateOfBirth,
        "profilePicture": None,
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "updatedAt": datetime.now(timezone.utc).isoformat()
    }
    await db.basic_profiles.insert_one(basic_profile)
    
    # Create session
    session_token = str(uuid.uuid4())
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    session = {
        "session_id": str(uuid.uuid4()),
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.user_sessions.insert_one(session)
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7 * 24 * 60 * 60
    )
    
    return {
        "user_id": user_id,
        "email": email,
        "name": fullName,
        "firstName": firstName,
        "lastName": lastName,
        "picture": None,
        "session_token": session_token,
        "isNewUser": True  # Flag for welcome/onboarding flow
    }

async def get_current_user(request: Request):
    """Get current user from session token (cookie or header)"""
    # Get token from cookies
    token = request.cookies.get("session_token")
    
    # Fallback to Authorization header
    if not token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            token = auth_header.split(" ")[1]
    
    if not token:
        return None
    
    # Find session
    session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not session:
        return None
    
    # Check expiry
    expires_at = session.get("expires_at")
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        return None
    
    # Get user
    user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    return user

@api_router.post("/auth/login")
async def jwt_login(request: JWTLoginRequest, response: Response):
    """JWT-based login using Email ID or Mobile Number"""
    identifier = request.username.strip()
    
    # Determine if identifier is email or mobile
    is_mobile = identifier.isdigit() and len(identifier) == 10
    is_email = "@" in identifier
    
    # For demo: accept test/test credentials
    if identifier == "test" and request.password == "test":
        # Check if test user exists
        user = await db.users.find_one({"email": "test@moneyssutra.com"}, {"_id": 0})
        
        if not user:
            # Create test user
            user_id = f"user_{uuid.uuid4().hex[:12]}"
            user = {
                "user_id": user_id,
                "email": "test@moneyssutra.com",
                "name": "Test User",
                "firstName": "Test",
                "lastName": "User",
                "picture": None,
                "auth_type": "jwt",
                "password_hash": hash_password("test"),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.users.insert_one(user)
        else:
            user_id = user["user_id"]
        
        # Create session
        session_token = str(uuid.uuid4())
        session_days = 30 if request.remember_me else 7
        expires_at = datetime.now(timezone.utc) + timedelta(days=session_days)
        
        session = {
            "session_id": str(uuid.uuid4()),
            "user_id": user_id,
            "session_token": session_token,
            "expires_at": expires_at.isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.user_sessions.insert_one(session)
        
        # Set cookie
        response.set_cookie(
            key="session_token",
            value=session_token,
            httponly=True,
            secure=True,
            samesite="none",
            path="/",
            max_age=session_days * 24 * 60 * 60
        )
        
        return {
            "user_id": user_id,
            "email": user.get("email"),
            "name": user.get("name"),
            "firstName": user.get("firstName", user.get("name", "").split()[0] if user.get("name") else ""),
            "picture": user.get("picture"),
            "session_token": session_token
        }
    
    # Build query with $or for email OR mobile
    query_conditions = []
    if is_email:
        query_conditions.append({"email": {"$regex": f"^{identifier}$", "$options": "i"}})
    if is_mobile:
        query_conditions.append({"mobile": identifier})
    
    # If neither email nor mobile format, try both
    if not is_email and not is_mobile:
        query_conditions = [
            {"email": {"$regex": f"^{identifier}$", "$options": "i"}},
            {"mobile": identifier}
        ]
    
    # Check actual user credentials - allow both jwt and google users with passwords
    user = await db.users.find_one(
        {"$or": query_conditions},
        {"_id": 0}
    )
    
    # Also check basic_profiles for mobile if user not found
    if not user and is_mobile:
        profile = await db.basic_profiles.find_one({"mobile": identifier}, {"_id": 0, "user_id": 1})
        if profile:
            user = await db.users.find_one({"user_id": profile["user_id"]}, {"_id": 0})
    
    if not user:
        raise HTTPException(status_code=401, detail="Invalid email/mobile or password")
    
    # Check if user has a password set
    if not user.get("password_hash"):
        if user.get("auth_type") == "google":
            raise HTTPException(status_code=401, detail="No password set. Please login with Google or set a password in your profile.")
        raise HTTPException(status_code=401, detail="Invalid email/mobile or password")
    
    if not verify_password(request.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Invalid email/mobile or password")
    
    # Create session
    session_token = str(uuid.uuid4())
    session_days = 30 if request.remember_me else 7
    expires_at = datetime.now(timezone.utc) + timedelta(days=session_days)
    
    session = {
        "session_id": str(uuid.uuid4()),
        "user_id": user["user_id"],
        "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.user_sessions.insert_one(session)
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=session_days * 24 * 60 * 60
    )
    
    return {
        "user_id": user["user_id"],
        "email": user.get("email"),
        "name": user.get("name"),
        "firstName": user.get("firstName", user.get("name", "").split()[0] if user.get("name") else ""),
        "picture": user.get("picture"),
        "session_token": session_token
    }

@api_router.post("/auth/google/session")
async def google_session(request: GoogleSessionRequest, response: Response):
    """Process Google OAuth session_id from Emergent Auth"""
    # REMINDER: DO NOT HARDCODE THE URL, OR ADD ANY FALLBACKS OR REDIRECT URLS, THIS BREAKS THE AUTH
    
    try:
        # Call Emergent Auth to get user data
        async with httpx.AsyncClient() as client:
            auth_response = await client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": request.session_id}
            )
            
            if auth_response.status_code != 200:
                raise HTTPException(status_code=401, detail="Invalid session")
            
            session_data = auth_response.json()
    except Exception as e:
        logging.error(f"Google auth error: {e}")
        raise HTTPException(status_code=401, detail="Authentication failed")
    
    # Check if user exists
    email = session_data.get("email")
    user = await db.users.find_one({"email": email}, {"_id": 0})
    
    is_new_user = False
    has_password = False
    
    if not user:
        # Create new user
        is_new_user = True
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        user = {
            "user_id": user_id,
            "email": email,
            "name": session_data.get("name"),
            "picture": session_data.get("picture"),
            "auth_type": "google",
            "password_hash": None,
            "has_password": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(user)
    else:
        user_id = user["user_id"]
        has_password = user.get("has_password", False) or (user.get("password_hash") is not None)
        # Update user data
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {
                "name": session_data.get("name"),
                "picture": session_data.get("picture")
            }}
        )
    
    # Create session
    session_token = session_data.get("session_token") or str(uuid.uuid4())
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    session = {
        "session_id": str(uuid.uuid4()),
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.user_sessions.insert_one(session)
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7 * 24 * 60 * 60
    )
    
    return {
        "user_id": user_id,
        "email": email,
        "name": session_data.get("name"),
        "picture": session_data.get("picture"),
        "session_token": session_token,
        "auth_type": "google",
        "has_password": has_password,
        "is_new_user": is_new_user
    }

@api_router.get("/auth/me")
async def get_me(request: Request):
    """Get current authenticated user"""
    user = await get_current_user(request)
    
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    return {
        "user_id": user.get("user_id"),
        "email": user.get("email"),
        "name": user.get("name"),
        "picture": user.get("picture"),
        "auth_type": user.get("auth_type"),
        "has_password": user.get("has_password", user.get("auth_type") == "jwt" or user.get("password_hash") is not None)
    }


class SetPasswordRequest(BaseModel):
    password: str


@api_router.post("/auth/set-password")
async def set_password(request: SetPasswordRequest, req: Request):
    """Allow Google users to set a password for email/password login"""
    user = await get_current_user(req)
    
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    if len(request.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    
    # Update user with password hash
    await db.users.update_one(
        {"user_id": user["user_id"]},
        {"$set": {
            "password_hash": hash_password(request.password),
            "has_password": True
        }}
    )
    
    return {"message": "Password set successfully. You can now login with email and password."}


@api_router.post("/auth/logout")
async def logout(request: Request, response: Response, session_token: Optional[str] = Cookie(None)):
    """Logout and invalidate session"""
    token = session_token
    
    # Fallback to Authorization header
    if not token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            token = auth_header.split(" ")[1]
    
    if token:
        await db.user_sessions.delete_one({"session_token": token})
    
    # Clear cookie
    response.delete_cookie(key="session_token", path="/")
    
    return {"message": "Logged out successfully"}

# ============ AUTH RECOVERY ENDPOINTS ============

from email_service import (
    send_username_recovery_email,
    send_password_reset_email,
    send_password_changed_notification
)
import secrets

class CheckAvailabilityRequest(BaseModel):
    username: Optional[str] = None
    email: Optional[str] = None

class ForgotUsernameRequest(BaseModel):
    email: str

class ForgotPasswordRequest(BaseModel):
    username: str

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

@api_router.post("/auth/check-availability")
async def check_availability(request: CheckAvailabilityRequest):
    """
    Check if username or email is already taken.
    Used for real-time validation during registration.
    Case-insensitive matching for both username and email.
    """
    result = {
        "username_available": True,
        "email_available": True,
        "message": ""
    }
    
    if request.username:
        username = request.username.strip()
        # Check if username (name field) is taken - case insensitive, exact match
        existing_by_name = await db.users.find_one(
            {"name": {"$regex": f"^{username}$", "$options": "i"}},
            {"_id": 0, "name": 1}
        )
        
        # Also check if username matches an email's local part (before @)
        # This prevents user from registering with "john" if "john@example.com" exists
        existing_by_email_prefix = await db.users.find_one(
            {"email": {"$regex": f"^{username}@", "$options": "i"}},
            {"_id": 0, "email": 1}
        )
        
        if existing_by_name or existing_by_email_prefix:
            result["username_available"] = False
            result["message"] = "This username is already taken. Please choose another."
    
    if request.email:
        email = request.email.strip().lower()
        # Check if email is taken - case insensitive, exact match
        existing_user = await db.users.find_one(
            {"email": {"$regex": f"^{email}$", "$options": "i"}},
            {"_id": 0, "email": 1}
        )
        if existing_user:
            result["email_available"] = False
            result["message"] = "This email is already registered."
    
    return result

@api_router.post("/auth/forgot-username")
async def forgot_username(request: ForgotUsernameRequest):
    """
    Send username recovery email.
    Uses generic messaging to prevent email enumeration attacks.
    """
    # Find user by email
    user = await db.users.find_one(
        {"email": {"$regex": f"^{request.email}$", "$options": "i"}},
        {"_id": 0, "name": 1, "email": 1, "auth_type": 1}
    )
    
    # Always return success message to prevent email enumeration
    success_message = "If an account exists with this email, you will receive your username shortly."
    
    if user:
        # Only send for JWT auth users (Google users can just use Google login)
        if user.get("auth_type") == "jwt" or user.get("auth_type") is None:
            username = user.get("name", "User")
            email_result = await send_username_recovery_email(user["email"], username)
            
            if not email_result.get("success"):
                logger.error(f"Failed to send username recovery email: {email_result.get('error')}")
    
    return {"message": success_message}

@api_router.post("/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    """
    Send password reset email/OTP using registered Email ID or Mobile Number.
    """
    identifier = request.username.strip()
    
    # Determine if identifier is email or mobile
    is_mobile = identifier.isdigit() and len(identifier) == 10
    is_email = "@" in identifier
    
    # Build query conditions
    query_conditions = []
    if is_email:
        query_conditions.append({"email": {"$regex": f"^{identifier}$", "$options": "i"}})
    if is_mobile:
        query_conditions.append({"mobile": identifier})
    
    # If neither, check both
    if not is_email and not is_mobile:
        query_conditions = [
            {"email": {"$regex": f"^{identifier}$", "$options": "i"}},
            {"mobile": identifier}
        ]
    
    # Find user by email or mobile
    user = await db.users.find_one(
        {"$or": query_conditions},
        {"_id": 0, "user_id": 1, "name": 1, "firstName": 1, "email": 1, "mobile": 1, "auth_type": 1}
    )
    
    # Also check basic_profiles for mobile
    if not user and is_mobile:
        profile = await db.basic_profiles.find_one({"mobile": identifier}, {"_id": 0, "user_id": 1, "firstName": 1})
        if profile:
            user = await db.users.find_one({"user_id": profile["user_id"]}, {"_id": 0})
            if user:
                user["firstName"] = profile.get("firstName", user.get("name", "").split()[0] if user.get("name") else "User")
    
    # Always return success message to prevent user enumeration
    success_message = "If an account exists with this email or mobile number, you will receive a password reset link shortly."
    
    if user:
        # Only allow password reset for JWT auth users
        if user.get("auth_type") == "jwt" or user.get("auth_type") is None:
            # Generate secure reset token
            reset_token = secrets.token_urlsafe(32)
            expires_at = datetime.now(timezone.utc) + timedelta(minutes=30)
            
            # Store reset token in database
            reset_record = {
                "token_id": str(uuid.uuid4()),
                "user_id": user["user_id"],
                "reset_token": reset_token,
                "expires_at": expires_at.isoformat(),
                "used": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.password_reset_tokens.insert_one(reset_record)
            
            # Send reset email (use firstName for greeting)
            first_name = user.get("firstName", user.get("name", "User").split()[0] if user.get("name") else "User")
            email_result = await send_password_reset_email(user["email"], first_name, reset_token)
            
            if not email_result.get("success"):
                logger.error(f"Failed to send password reset email: {email_result.get('error')}")
    
    return {"message": success_message}

@api_router.post("/auth/reset-password")
async def reset_password(request: ResetPasswordRequest):
    """
    Reset password using valid token.
    """
    # Find valid reset token
    reset_record = await db.password_reset_tokens.find_one(
        {"reset_token": request.token, "used": False},
        {"_id": 0}
    )
    
    if not reset_record:
        raise HTTPException(status_code=400, detail="Invalid or expired reset link")
    
    # Check if token is expired
    expires_at = reset_record.get("expires_at")
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Reset link has expired. Please request a new one.")
    
    # Validate password
    if len(request.new_password) < 4:
        raise HTTPException(status_code=400, detail="Password must be at least 4 characters")
    
    # Get user
    user = await db.users.find_one(
        {"user_id": reset_record["user_id"]},
        {"_id": 0, "user_id": 1, "name": 1, "email": 1}
    )
    
    if not user:
        raise HTTPException(status_code=400, detail="User not found")
    
    # Update password
    new_password_hash = hash_password(request.new_password)
    await db.users.update_one(
        {"user_id": user["user_id"]},
        {"$set": {"password_hash": new_password_hash}}
    )
    
    # Mark token as used
    await db.password_reset_tokens.update_one(
        {"reset_token": request.token},
        {"$set": {"used": True}}
    )
    
    # Invalidate all existing sessions for security
    await db.user_sessions.delete_many({"user_id": user["user_id"]})
    
    # Send security notification email
    username = user.get("name", "User")
    email_result = await send_password_changed_notification(user["email"], username)
    
    if not email_result.get("success"):
        logger.error(f"Failed to send password changed notification: {email_result.get('error')}")
    
    return {"message": "Password reset successfully. Please log in with your new password."}

@api_router.get("/auth/verify-reset-token")
async def verify_reset_token(token: str):
    """
    Verify if a reset token is valid (not expired, not used).
    Used by frontend to validate token before showing reset form.
    """
    reset_record = await db.password_reset_tokens.find_one(
        {"reset_token": token, "used": False},
        {"_id": 0}
    )
    
    if not reset_record:
        return {"valid": False, "message": "Invalid or expired reset link"}
    
    # Check if token is expired
    expires_at = reset_record.get("expires_at")
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    
    if expires_at < datetime.now(timezone.utc):
        return {"valid": False, "message": "Reset link has expired. Please request a new one."}
    
    return {"valid": True}


# ============ ENTITY UNIQUENESS CHECK ENDPOINT ============

class EntityUniquenessRequest(BaseModel):
    collection: str  # Collection name: income_sources, expenses, assets, loans, credit_cards, insurances
    field: str       # Field name to check: name, expenseName, assetName, loanName, cardName, policyName
    value: str       # Value to check
    exclude_id: Optional[str] = None  # ID to exclude (for edit mode)
    type_filter: Optional[str] = None  # Optional type filter for income sources (e.g., "Business")

@api_router.post("/check-entity-uniqueness")
async def check_entity_uniqueness(request_data: EntityUniquenessRequest, request: Request):
    """
    Check if an entity name is unique for the current user.
    Returns whether the name is available or already exists.
    """
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    collection_name = request_data.collection
    field_name = request_data.field
    value = request_data.value.strip()
    
    # Validate collection name
    allowed_collections = [
        "income_sources", "expenses", "assets", "loans", 
        "credit_cards", "insurances", "accounts", "investments", "goals"
    ]
    if collection_name not in allowed_collections:
        raise HTTPException(status_code=400, detail="Invalid collection name")
    
    # Build the query - case-insensitive search
    query = {
        "userId": user_id,
        field_name: {"$regex": f"^{value}$", "$options": "i"}
    }
    
    # Add type filter if provided (for income sources)
    if request_data.type_filter:
        query["type"] = request_data.type_filter
    
    # Exclude current entity if editing
    if request_data.exclude_id:
        query["id"] = {"$ne": request_data.exclude_id}
    
    # Check in the collection
    collection = db[collection_name]
    existing = await collection.find_one(query, {"_id": 0, "id": 1, field_name: 1})
    
    if existing:
        return {
            "available": False,
            "message": "An entry with this name already exists. Please use a unique name.",
            "existing_id": existing.get("id")
        }
    
    return {
        "available": True,
        "message": "Name is available"
    }


# ============ WORKSPACE ENDPOINTS ============

@api_router.get("/workspaces")
async def get_user_workspaces(request: Request):
    """Get all workspaces the user has access to"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    
    # Get all workspace memberships for the user
    memberships = await db.workspace_members.find({
        "user_id": user_id,
        "status": "active"
    }, {"_id": 0}).to_list(100)
    
    workspaces = []
    for membership in memberships:
        workspace = await db.workspaces.find_one({"id": membership['workspace_id']}, {"_id": 0})
        if workspace:
            workspace['role'] = membership['role']
            workspace['member_id'] = membership['id']
            workspaces.append(workspace)
    
    return workspaces

@api_router.post("/workspaces")
async def create_workspace(input: WorkspaceCreate, request: Request):
    """Create a new workspace"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    user_email = user.get('email', '')
    user_name = user.get('name', 'User')
    
    # Create workspace
    workspace_id = f"ws_{uuid.uuid4().hex[:12]}"
    invite_code = uuid.uuid4().hex[:8].upper()
    
    workspace = {
        "id": workspace_id,
        "name": input.name,
        "type": input.type,
        "owner_id": user_id,
        "invite_code": invite_code,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.workspaces.insert_one(workspace)
    
    # Add creator as owner
    member = {
        "id": f"wm_{uuid.uuid4().hex[:12]}",
        "workspace_id": workspace_id,
        "user_id": user_id,
        "user_email": user_email,
        "user_name": user_name,
        "role": "owner",
        "invited_by": None,
        "invited_at": datetime.now(timezone.utc).isoformat(),
        "accepted_at": datetime.now(timezone.utc).isoformat(),
        "status": "active"
    }
    await db.workspace_members.insert_one(member)
    
    # Return clean response (exclude MongoDB's _id)
    response = {
        "id": workspace_id,
        "name": input.name,
        "type": input.type,
        "owner_id": user_id,
        "invite_code": invite_code,
        "created_at": workspace["created_at"],
        "role": "owner"
    }
    return response

@api_router.get("/workspaces/current")
async def get_current_workspace(request: Request, workspace_id: Optional[str] = None):
    """Get current workspace with user's role"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Ensure user has a workspace
    default_ws_id = await ensure_user_has_workspace(user)
    
    workspace, role = await get_user_workspace(user, workspace_id or default_ws_id)
    if not workspace:
        raise HTTPException(status_code=404, detail="No workspace found")
    
    workspace['role'] = role
    workspace['permissions'] = ROLE_PERMISSIONS.get(role, {})
    
    # Get member count
    member_count = await db.workspace_members.count_documents({
        "workspace_id": workspace['id'],
        "status": "active"
    })
    workspace['member_count'] = member_count
    
    return workspace

@api_router.get("/workspaces/{workspace_id}")
async def get_workspace(workspace_id: str, request: Request):
    """Get a specific workspace"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    workspace, role = await get_user_workspace(user, workspace_id)
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace not found or access denied")
    
    workspace['role'] = role
    workspace['permissions'] = ROLE_PERMISSIONS.get(role, {})
    
    return workspace

@api_router.get("/workspaces/{workspace_id}/members")
async def get_workspace_members(workspace_id: str, request: Request):
    """Get all members of a workspace"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    workspace, role = await get_user_workspace(user, workspace_id)
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace not found or access denied")
    
    members = await db.workspace_members.find({
        "workspace_id": workspace_id
    }, {"_id": 0}).to_list(100)
    
    return members

@api_router.post("/workspaces/{workspace_id}/invite")
async def invite_to_workspace(workspace_id: str, invite: WorkspaceInvite, request: Request):
    """Invite a user to workspace via email"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    workspace, role = await get_user_workspace(user, workspace_id)
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace not found or access denied")
    
    # Check if user has invite permission
    if not check_permission(role, 'invite'):
        raise HTTPException(status_code=403, detail="You don't have permission to invite members")
    
    # Check if email is already a member
    existing = await db.workspace_members.find_one({
        "workspace_id": workspace_id,
        "user_email": invite.email
    }, {"_id": 0})
    
    if existing:
        if existing['status'] == 'active':
            raise HTTPException(status_code=400, detail="User is already a member")
        elif existing['status'] == 'pending':
            raise HTTPException(status_code=400, detail="User already has a pending invitation")
    
    # Check if invited user exists
    invited_user = await db.users.find_one({"email": invite.email}, {"_id": 0})
    
    # Create invitation
    member = {
        "id": f"wm_{uuid.uuid4().hex[:12]}",
        "workspace_id": workspace_id,
        "user_id": invited_user['user_id'] if invited_user else None,
        "user_email": invite.email,
        "user_name": invited_user['name'] if invited_user else invite.email.split('@')[0],
        "role": invite.role,
        "invited_by": user.get('user_id'),
        "invited_at": datetime.now(timezone.utc).isoformat(),
        "accepted_at": None,
        "status": "pending"
    }
    await db.workspace_members.insert_one(member)
    
    return {
        "message": f"Invitation sent to {invite.email}",
        "member_id": member['id'],
        "status": "pending"
    }

@api_router.post("/workspaces/join")
async def join_workspace_by_code(invite: WorkspaceInviteByCode, request: Request):
    """Join a workspace using invite code"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    user_email = user.get('email', '')
    user_name = user.get('name', 'User')
    
    # Find workspace by invite code
    workspace = await db.workspaces.find_one({"invite_code": invite.invite_code.upper()}, {"_id": 0})
    if not workspace:
        raise HTTPException(status_code=404, detail="Invalid invite code")
    
    # Check if already a member
    existing = await db.workspace_members.find_one({
        "workspace_id": workspace['id'],
        "user_id": user_id,
        "status": "active"
    }, {"_id": 0})
    
    if existing:
        raise HTTPException(status_code=400, detail="You are already a member of this workspace")
    
    # Check for pending invitation
    pending = await db.workspace_members.find_one({
        "workspace_id": workspace['id'],
        "user_email": user_email,
        "status": "pending"
    }, {"_id": 0})
    
    if pending:
        # Accept pending invitation
        await db.workspace_members.update_one(
            {"id": pending['id']},
            {"$set": {
                "user_id": user_id,
                "user_name": user_name,
                "status": "active",
                "accepted_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        return {
            "message": f"Joined workspace: {workspace['name']}",
            "workspace_id": workspace['id'],
            "role": pending['role']
        }
    
    # Join as viewer by default (via code)
    member = {
        "id": f"wm_{uuid.uuid4().hex[:12]}",
        "workspace_id": workspace['id'],
        "user_id": user_id,
        "user_email": user_email,
        "user_name": user_name,
        "role": "viewer",
        "invited_by": None,
        "invited_at": datetime.now(timezone.utc).isoformat(),
        "accepted_at": datetime.now(timezone.utc).isoformat(),
        "status": "active"
    }
    await db.workspace_members.insert_one(member)
    
    return {
        "message": f"Joined workspace: {workspace['name']}",
        "workspace_id": workspace['id'],
        "role": "viewer"
    }

@api_router.post("/workspaces/accept/{member_id}")
async def accept_workspace_invitation(member_id: str, request: Request):
    """Accept a pending workspace invitation"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    user_email = user.get('email', '')
    user_name = user.get('name', 'User')
    
    # Find pending invitation
    invitation = await db.workspace_members.find_one({
        "id": member_id,
        "user_email": user_email,
        "status": "pending"
    }, {"_id": 0})
    
    if not invitation:
        raise HTTPException(status_code=404, detail="Invitation not found or already accepted")
    
    # Accept invitation
    await db.workspace_members.update_one(
        {"id": member_id},
        {"$set": {
            "user_id": user_id,
            "user_name": user_name,
            "status": "active",
            "accepted_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    workspace = await db.workspaces.find_one({"id": invitation['workspace_id']}, {"_id": 0})
    
    return {
        "message": f"Successfully joined {workspace['name'] if workspace else 'workspace'}",
        "workspace_id": invitation['workspace_id'],
        "role": invitation['role']
    }

@api_router.get("/workspaces/invitations/pending")
async def get_pending_invitations(request: Request):
    """Get all pending invitations for the current user"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_email = user.get('email', '')
    
    invitations = await db.workspace_members.find({
        "user_email": user_email,
        "status": "pending"
    }, {"_id": 0}).to_list(100)
    
    # Enrich with workspace details
    result = []
    for inv in invitations:
        workspace = await db.workspaces.find_one({"id": inv['workspace_id']}, {"_id": 0})
        if workspace:
            inv['workspace_name'] = workspace['name']
            inv['workspace_type'] = workspace['type']
            
            # Get inviter details
            if inv.get('invited_by'):
                inviter = await db.users.find_one({"user_id": inv['invited_by']}, {"_id": 0})
                if inviter:
                    inv['invited_by_name'] = inviter.get('name', 'Unknown')
            
            result.append(inv)
    
    return result

@api_router.put("/workspaces/{workspace_id}/members/{member_id}/role")
async def update_member_role(workspace_id: str, member_id: str, new_role: str, request: Request):
    """Update a member's role (owner only)"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    workspace, role = await get_user_workspace(user, workspace_id)
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace not found or access denied")
    
    if role != 'owner':
        raise HTTPException(status_code=403, detail="Only the owner can change member roles")
    
    if new_role not in ['admin', 'editor', 'viewer']:
        raise HTTPException(status_code=400, detail="Invalid role. Must be admin, editor, or viewer")
    
    # Cannot change owner's role
    member = await db.workspace_members.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    if member.get('role') == 'owner':
        raise HTTPException(status_code=400, detail="Cannot change owner's role")
    
    await db.workspace_members.update_one(
        {"id": member_id},
        {"$set": {"role": new_role}}
    )
    
    return {"message": f"Role updated to {new_role}", "member_id": member_id}

@api_router.delete("/workspaces/{workspace_id}/members/{member_id}")
async def remove_workspace_member(workspace_id: str, member_id: str, request: Request):
    """Remove a member from workspace (owner only) or leave workspace"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    workspace, role = await get_user_workspace(user, workspace_id)
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace not found or access denied")
    
    member = await db.workspace_members.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    # Owner cannot be removed
    if member.get('role') == 'owner':
        raise HTTPException(status_code=400, detail="Cannot remove the workspace owner")
    
    # User can remove themselves (leave) or owner can remove anyone
    if member.get('user_id') != user_id and role != 'owner':
        raise HTTPException(status_code=403, detail="You don't have permission to remove this member")
    
    await db.workspace_members.update_one(
        {"id": member_id},
        {"$set": {"status": "removed"}}
    )
    
    return {"message": "Member removed successfully", "member_id": member_id}

@api_router.put("/workspaces/{workspace_id}/regenerate-code")
async def regenerate_invite_code(workspace_id: str, request: Request):
    """Regenerate the workspace invite code (owner only)"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    workspace, role = await get_user_workspace(user, workspace_id)
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace not found or access denied")
    
    if role != 'owner':
        raise HTTPException(status_code=403, detail="Only the owner can regenerate the invite code")
    
    new_code = uuid.uuid4().hex[:8].upper()
    await db.workspaces.update_one(
        {"id": workspace_id},
        {"$set": {"invite_code": new_code}}
    )
    
    return {"invite_code": new_code}

# ============ INCOME ENDPOINTS ============

async def get_user_workspace(user, workspace_id: Optional[str] = None):
    """Get the current workspace for the user"""
    user_id = user.get('user_id')
    
    if workspace_id:
        # Check if user has access to the specified workspace
        member = await db.workspace_members.find_one({
            "workspace_id": workspace_id,
            "user_id": user_id,
            "status": "active"
        }, {"_id": 0})
        if member:
            workspace = await db.workspaces.find_one({"id": workspace_id}, {"_id": 0})
            return workspace, member.get('role', 'viewer')
    
    # Get user's default workspace (first one they own or are active in)
    member = await db.workspace_members.find_one({
        "user_id": user_id,
        "status": "active"
    }, {"_id": 0}, sort=[("role", 1)])  # Owner first
    
    if member:
        workspace = await db.workspaces.find_one({"id": member['workspace_id']}, {"_id": 0})
        return workspace, member.get('role', 'viewer')
    
    return None, None

async def ensure_user_has_workspace(user):
    """Ensure user has at least one workspace, create default if needed"""
    user_id = user.get('user_id')
    user_email = user.get('email', '')
    user_name = user.get('name', 'User')
    
    # Check if user already has a workspace
    existing_member = await db.workspace_members.find_one({
        "user_id": user_id,
        "status": "active"
    }, {"_id": 0})
    
    if existing_member:
        return existing_member['workspace_id']
    
    # Create default Personal workspace
    workspace_id = f"ws_{uuid.uuid4().hex[:12]}"
    invite_code = uuid.uuid4().hex[:8].upper()
    
    workspace = {
        "id": workspace_id,
        "name": f"{user_name}'s Finance",
        "type": "Personal",
        "owner_id": user_id,
        "invite_code": invite_code,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.workspaces.insert_one(workspace)
    
    # Add user as owner
    member = {
        "id": f"wm_{uuid.uuid4().hex[:12]}",
        "workspace_id": workspace_id,
        "user_id": user_id,
        "user_email": user_email,
        "user_name": user_name,
        "role": "owner",
        "invited_by": None,
        "invited_at": datetime.now(timezone.utc).isoformat(),
        "accepted_at": datetime.now(timezone.utc).isoformat(),
        "status": "active"
    }
    await db.workspace_members.insert_one(member)
    
    # Migrate user's existing data to this workspace (for legacy data)
    if user_email == 'test@moneyssutra.com':
        # Migrate legacy data (without workspaceId) to this workspace
        collections = ['income_sources', 'other_income', 'loans', 'assets', 'accounts', 
                      'expenses', 'investments', 'goals', 'credit_cards', 'insurances']
        for coll in collections:
            await db[coll].update_many(
                {"$or": [{"workspaceId": None}, {"workspaceId": {"$exists": False}}]},
                {"$set": {"workspaceId": workspace_id}}
            )
    
    return workspace_id

def get_workspace_filter(workspace_id: str, user_id: str = None, user_email: str = None):
    """Get the appropriate MongoDB filter for workspace data isolation"""
    if user_email == 'test@moneyssutra.com':
        # Test user: include legacy data AND workspace data
        return {"$or": [
            {"workspaceId": workspace_id},
            {"workspaceId": None},
            {"workspaceId": {"$exists": False}},
            {"userId": user_id},
            {"userId": None},
            {"userId": {"$exists": False}}
        ]}
    else:
        return {"workspaceId": workspace_id}

def check_permission(role: str, action: str) -> bool:
    """Check if a role has permission for an action"""
    return ROLE_PERMISSIONS.get(role, {}).get(action, False)

def get_user_filter(user):
    """Get the appropriate MongoDB filter for user data isolation"""
    user_id = user.get('user_id')
    user_email = user.get('email', '')
    
    # Test user can see legacy data (no userId), new users only see their own
    if user_email == 'test@moneyssutra.com' or user_id == 'test':
        return {"$or": [{"userId": user_id}, {"userId": None}, {"userId": {"$exists": False}}]}
    else:
        return {"userId": user_id}

@api_router.post("/income", response_model=IncomeSource)
async def create_income_source(input: IncomeSourceCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    income_dict = input.model_dump()
    income_dict['userId'] = user.get('user_id')  # Add user isolation
    income_obj = IncomeSource(**income_dict)
    
    # Convert to dict and serialize datetime to ISO string for MongoDB
    doc = income_obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    
    _ = await db.income_sources.insert_one(doc)
    return income_obj

@api_router.get("/income", response_model=List[IncomeSource])
async def get_income_sources(request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    income_sources = await db.income_sources.find(user_filter, {"_id": 0}).to_list(1000)
    
    # Convert ISO string timestamps back to datetime objects
    for source in income_sources:
        if isinstance(source.get('createdAt'), str):
            source['createdAt'] = datetime.fromisoformat(source['createdAt'])
    
    return income_sources

@api_router.get("/income/list/summary")
async def get_income_list_summary(request: Request, type: Optional[str] = None):
    """
    Optimized endpoint for list views - returns lightweight data with transaction summary.
    Only fetches essential fields + latest transaction + total recorded.
    """
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    if type:
        # Case-insensitive type filter using regex
        user_filter["type"] = {"$regex": f"^{type}$", "$options": "i"}
    
    # Fetch only essential fields for list view
    projection = {
        "_id": 0,
        "id": 1,
        "name": 1,
        "type": 1,
        "expectedAmount": 1,
        "frequency": 1,
        "selectedDay": 1,
        "selectedDate": 1,
        "selectedMonth": 1,
        "selectedQuarter": 1,
        "incomeType": 1
    }
    
    income_sources = await db.income_sources.find(user_filter, projection).to_list(1000)
    
    # Batch fetch transaction summaries for all entities
    entity_ids = [s["id"] for s in income_sources]
    
    # Get transaction counts and totals using aggregation
    pipeline = [
        {"$match": {"entityId": {"$in": entity_ids}}},
        {"$group": {
            "_id": "$entityId",
            "totalRecorded": {"$sum": "$amount"},
            "transactionCount": {"$sum": 1},
            "lastTransaction": {"$max": "$transactionDate"}
        }}
    ]
    
    transaction_stats = {}
    async for stat in db.income_transactions.aggregate(pipeline):
        transaction_stats[stat["_id"]] = {
            "totalRecorded": stat["totalRecorded"],
            "transactionCount": stat["transactionCount"],
            "lastTransaction": stat["lastTransaction"]
        }
    
    # Merge transaction stats into income sources
    for source in income_sources:
        stats = transaction_stats.get(source["id"], {})
        source["totalRecorded"] = stats.get("totalRecorded", 0)
        source["transactionCount"] = stats.get("transactionCount", 0)
        source["lastTransaction"] = stats.get("lastTransaction")
    
    return income_sources

@api_router.get("/income/{income_id}", response_model=IncomeSource)
async def get_income_source(income_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = income_id
    income_source = await db.income_sources.find_one(user_filter, {"_id": 0})
    
    if not income_source:
        raise HTTPException(status_code=404, detail="Income source not found")
    
    if isinstance(income_source.get('createdAt'), str):
        income_source['createdAt'] = datetime.fromisoformat(income_source['createdAt'])
    
    return income_source

@api_router.put("/income/{income_id}", response_model=IncomeSource)
async def update_income_source(income_id: str, input: IncomeSourceCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = income_id
    existing = await db.income_sources.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Income source not found")
    
    income_dict = input.model_dump()
    income_dict['id'] = income_id
    income_dict['userId'] = user.get('user_id')
    income_dict['createdAt'] = existing['createdAt']
    
    if isinstance(income_dict['createdAt'], str):
        pass
    else:
        income_dict['createdAt'] = income_dict['createdAt'].isoformat()
    
    await db.income_sources.replace_one({"id": income_id}, income_dict)
    
    income_obj = IncomeSource(**income_dict)
    if isinstance(income_obj.createdAt, str):
        income_obj.createdAt = datetime.fromisoformat(income_obj.createdAt)
    
    return income_obj

@api_router.delete("/income/{income_id}")
async def delete_income_source(income_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = income_id
    existing = await db.income_sources.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Income source not found")
    
    await db.income_sources.delete_one({"id": income_id})
    
    return {"message": "Income source deleted successfully", "id": income_id}

# ============ LOAN ENDPOINTS ============

@api_router.post("/loans", response_model=Loan)
async def create_loan(input: LoanCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    loan_dict = input.model_dump()
    loan_dict['userId'] = user.get('user_id')
    loan_obj = Loan(**loan_dict)
    
    doc = loan_obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    
    await db.loans.insert_one(doc)
    
    # Auto-create EMI expense if enabled
    if loan_obj.autoCreateExpense:
        existing_expense = await db.expenses.find_one({"linkedLoanId": loan_obj.id}, {"_id": 0})
        if not existing_expense:
            freq_map = {"Monthly": "Monthly", "Quarterly": "Quarterly", "Half-Yearly": "Half-Yearly"}
            expense_freq = freq_map.get(loan_obj.emiFrequency, "Monthly")
            
            start_date = datetime.fromisoformat(loan_obj.startDate) if loan_obj.startDate else datetime.now(timezone.utc)
            selected_date = str(start_date.day)
            
            expense_data = {
                "id": str(uuid.uuid4()),
                "userId": user.get('user_id'),
                "expenseName": f"{loan_obj.loanName} EMI",
                "expenseType": "Fixed",
                "category": "EMI",
                "expectedAmount": loan_obj.emiAmount,
                "frequency": expense_freq,
                "linkedAccountId": loan_obj.linkedAccountId,
                "linkedLoanId": loan_obj.id,
                "linkedInsuranceId": None,
                "selectedDay": None,
                "selectedDate": selected_date,
                "selectedQuarter": None,
                "selectedHalf": None,
                "selectedMonth": None,
                "oneTimeDate": None,
                "isPaid": False,
                "lastPaidDate": None,
                "createdAt": datetime.now(timezone.utc).isoformat()
            }
            await db.expenses.insert_one(expense_data)
    
    return loan_obj

@api_router.get("/loans", response_model=List[Loan])
async def get_loans(request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    loans = await db.loans.find(user_filter, {"_id": 0}).to_list(1000)
    
    for loan in loans:
        if isinstance(loan.get('createdAt'), str):
            loan['createdAt'] = datetime.fromisoformat(loan['createdAt'])
    
    return loans

@api_router.get("/loans/{loan_id}", response_model=Loan)
async def get_loan(loan_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = loan_id
    loan = await db.loans.find_one(user_filter, {"_id": 0})
    
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    if isinstance(loan.get('createdAt'), str):
        loan['createdAt'] = datetime.fromisoformat(loan['createdAt'])
    
    return loan

@api_router.put("/loans/{loan_id}", response_model=Loan)
async def update_loan(loan_id: str, input: LoanCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = loan_id
    existing = await db.loans.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    loan_dict = input.model_dump()
    loan_dict['id'] = loan_id
    loan_dict['userId'] = user.get('user_id')
    loan_dict['createdAt'] = existing['createdAt']
    
    await db.loans.replace_one({"id": loan_id}, loan_dict)
    
    loan_obj = Loan(**loan_dict)
    if isinstance(loan_obj.createdAt, str):
        loan_obj.createdAt = datetime.fromisoformat(loan_obj.createdAt)
    
    return loan_obj

@api_router.delete("/loans/{loan_id}")
async def delete_loan(loan_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = loan_id
    existing = await db.loans.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    await db.loans.delete_one({"id": loan_id})
    return {"message": "Loan deleted successfully", "id": loan_id}

@api_router.get("/loans/{loan_id}/linked-assets")
async def get_loan_linked_assets(loan_id: str, request: Request):
    """Get all assets that are linked to this loan (reverse lookup)"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["linkedLoanId"] = loan_id
    linked_assets = await db.assets.find(user_filter, {"_id": 0}).to_list(1000)
    
    result = []
    for asset in linked_assets:
        if isinstance(asset.get('createdAt'), str):
            asset['createdAt'] = datetime.fromisoformat(asset['createdAt'])
        result.append({
            "id": asset.get('id'),
            "assetName": asset.get('assetName'),
            "assetType": asset.get('assetType'),
            "currentValue": asset.get('currentValue', 0),
            "purchaseValue": asset.get('purchaseValue'),
            "location": asset.get('location')
        })
    
    return result

# ============ ASSET ENDPOINTS ============

@api_router.post("/assets", response_model=Asset)
async def create_asset(input: AssetCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    asset_dict = input.model_dump()
    asset_dict['userId'] = user.get('user_id')
    asset_obj = Asset(**asset_dict)
    
    # Auto-create Rental Income if asset generates income
    if asset_obj.generatesIncome and asset_obj.incomeAmount:
        existing_income = await db.income_sources.find_one({"assetId": asset_obj.id}, {"_id": 0})
        
        if not existing_income:
            rental_income = {
                "id": str(uuid.uuid4()),
                "userId": user.get('user_id'),
                "type": "Rental",
                "name": asset_obj.assetName,
                "expectedAmount": asset_obj.incomeAmount,
                "frequency": asset_obj.incomeFrequency or "Monthly",
                "tenantName": asset_dict.get("renterName") or None,
                "securityDeposit": asset_dict.get("securityDeposit") or None,
                "assetId": asset_obj.id,
                "assetValue": asset_obj.currentValue,
                "rentalYield": round((asset_obj.incomeAmount * 12 / asset_obj.currentValue) * 100, 2) if asset_obj.currentValue else None,
                "createdAt": datetime.now(timezone.utc).isoformat(),
            }
            
            await db.income_sources.insert_one(rental_income)
            asset_obj.linkedIncomeId = rental_income["id"]
    
    doc = asset_obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    
    await db.assets.insert_one(doc)
    return asset_obj

@api_router.get("/assets", response_model=List[Asset])
async def get_assets(request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    assets = await db.assets.find(user_filter, {"_id": 0}).to_list(1000)
    
    for asset in assets:
        if isinstance(asset.get('createdAt'), str):
            asset['createdAt'] = datetime.fromisoformat(asset['createdAt'])
    
    return assets

@api_router.get("/assets/{asset_id}", response_model=Asset)
async def get_asset(asset_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = asset_id
    asset = await db.assets.find_one(user_filter, {"_id": 0})
    
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    if isinstance(asset.get('createdAt'), str):
        asset['createdAt'] = datetime.fromisoformat(asset['createdAt'])
    
    return asset

@api_router.put("/assets/{asset_id}", response_model=Asset)
async def update_asset(asset_id: str, input: AssetCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = asset_id
    existing = await db.assets.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    asset_dict = input.model_dump()
    asset_dict['id'] = asset_id
    asset_dict['userId'] = user.get('user_id')
    asset_dict['createdAt'] = existing['createdAt']
    
    # Handle rental income linking
    if asset_dict.get('generatesIncome') and asset_dict.get('incomeAmount'):
        existing_income = await db.income_sources.find_one({"assetId": asset_id}, {"_id": 0})
        
        if existing_income:
            await db.income_sources.update_one(
                {"assetId": asset_id},
                {"$set": {
                    "name": asset_dict['assetName'],
                    "expectedAmount": asset_dict['incomeAmount'],
                    "frequency": asset_dict.get('incomeFrequency') or "Monthly",
                    "tenantName": asset_dict.get('renterName'),
                    "securityDeposit": asset_dict.get('securityDeposit'),
                    "assetValue": asset_dict['currentValue'],
                    "rentalYield": round((asset_dict['incomeAmount'] * 12 / asset_dict['currentValue']) * 100, 2) if asset_dict['currentValue'] else None,
                }}
            )
            asset_dict['linkedIncomeId'] = existing_income['id']
        else:
            rental_income = {
                "id": str(uuid.uuid4()),
                "userId": user.get('user_id'),
                "type": "Rental",
                "name": asset_dict['assetName'],
                "expectedAmount": asset_dict['incomeAmount'],
                "frequency": asset_dict.get('incomeFrequency') or "Monthly",
                "tenantName": asset_dict.get('renterName'),
                "securityDeposit": asset_dict.get('securityDeposit'),
                "assetId": asset_id,
                "assetValue": asset_dict['currentValue'],
                "rentalYield": round((asset_dict['incomeAmount'] * 12 / asset_dict['currentValue']) * 100, 2) if asset_dict['currentValue'] else None,
                "createdAt": datetime.now(timezone.utc).isoformat(),
            }
            await db.income_sources.insert_one(rental_income)
            asset_dict['linkedIncomeId'] = rental_income['id']
    elif not asset_dict.get('generatesIncome'):
        asset_dict['linkedIncomeId'] = None
        await db.income_sources.delete_many({"assetId": asset_id})
    
    await db.assets.replace_one({"id": asset_id}, asset_dict)
    
    asset_obj = Asset(**asset_dict)
    if isinstance(asset_obj.createdAt, str):
        asset_obj.createdAt = datetime.fromisoformat(asset_obj.createdAt)
    
    return asset_obj

@api_router.delete("/assets/{asset_id}")
async def delete_asset(asset_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = asset_id
    existing = await db.assets.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    await db.assets.delete_one({"id": asset_id})
    return {"message": "Asset deleted successfully", "id": asset_id}

# ============ ACCOUNT ENDPOINTS ============

@api_router.post("/accounts", response_model=Account)
async def create_account(input: AccountCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    account_dict = input.model_dump()
    account_dict['userId'] = user.get('user_id')
    account_obj = Account(**account_dict)
    
    doc = account_obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    
    await db.accounts.insert_one(doc)
    return account_obj

@api_router.get("/accounts", response_model=List[Account])
async def get_accounts(request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    accounts = await db.accounts.find(user_filter, {"_id": 0}).to_list(1000)
    
    for account in accounts:
        if isinstance(account.get('createdAt'), str):
            account['createdAt'] = datetime.fromisoformat(account['createdAt'])
    
    return accounts

@api_router.get("/accounts/{account_id}", response_model=Account)
async def get_account(account_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = account_id
    account = await db.accounts.find_one(user_filter, {"_id": 0})
    
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    if isinstance(account.get('createdAt'), str):
        account['createdAt'] = datetime.fromisoformat(account['createdAt'])
    
    return account

@api_router.put("/accounts/{account_id}", response_model=Account)
async def update_account(account_id: str, input: AccountCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = account_id
    existing = await db.accounts.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Account not found")
    
    account_dict = input.model_dump()
    account_dict['id'] = account_id
    account_dict['userId'] = user.get('user_id')
    account_dict['createdAt'] = existing['createdAt']
    
    await db.accounts.replace_one({"id": account_id}, account_dict)
    
    account_obj = Account(**account_dict)
    if isinstance(account_obj.createdAt, str):
        account_obj.createdAt = datetime.fromisoformat(account_obj.createdAt)
    
    return account_obj

@api_router.delete("/accounts/{account_id}")
async def delete_account(account_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = account_id
    existing = await db.accounts.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Account not found")
    
    await db.accounts.delete_one({"id": account_id})
    return {"message": "Account deleted successfully", "id": account_id}

# ============ EXPENSE ENDPOINTS ============

@api_router.post("/expenses", response_model=Expense)
async def create_expense(input: ExpenseCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    expense_dict = input.model_dump()
    expense_dict['userId'] = user.get('user_id')
    expense_obj = Expense(**expense_dict)
    
    doc = expense_obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    
    await db.expenses.insert_one(doc)
    return expense_obj

@api_router.get("/expenses", response_model=List[Expense])
async def get_expenses(request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    expenses = await db.expenses.find(user_filter, {"_id": 0}).to_list(1000)
    
    for expense in expenses:
        if isinstance(expense.get('createdAt'), str):
            expense['createdAt'] = datetime.fromisoformat(expense['createdAt'])
    
    return expenses

@api_router.get("/expenses/list/summary")
async def get_expense_list_summary(request: Request, category: Optional[str] = None, expense_type: Optional[str] = None):
    """
    Optimized endpoint for expense list views - returns lightweight data with transaction summary.
    """
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    if category:
        user_filter["category"] = category
    if expense_type:
        user_filter["expenseType"] = expense_type
    
    # Fetch only essential fields for list view
    projection = {
        "_id": 0,
        "id": 1,
        "expenseName": 1,
        "expenseType": 1,
        "category": 1,
        "expectedAmount": 1,
        "frequency": 1,
        "selectedDay": 1,
        "selectedDate": 1,
        "linkedLoanId": 1,
        "linkedInsuranceId": 1,
        "linkedInvestmentId": 1
    }
    
    expenses = await db.expenses.find(user_filter, projection).to_list(1000)
    
    # Batch fetch transaction summaries
    entity_ids = [e["id"] for e in expenses]
    
    pipeline = [
        {"$match": {"entityId": {"$in": entity_ids}}},
        {"$group": {
            "_id": "$entityId",
            "totalRecorded": {"$sum": "$amount"},
            "transactionCount": {"$sum": 1},
            "lastTransaction": {"$max": "$transactionDate"}
        }}
    ]
    
    transaction_stats = {}
    async for stat in db.expense_transactions.aggregate(pipeline):
        transaction_stats[stat["_id"]] = {
            "totalRecorded": stat["totalRecorded"],
            "transactionCount": stat["transactionCount"],
            "lastTransaction": stat["lastTransaction"]
        }
    
    for expense in expenses:
        stats = transaction_stats.get(expense["id"], {})
        expense["totalRecorded"] = stats.get("totalRecorded", 0)
        expense["transactionCount"] = stats.get("transactionCount", 0)
        expense["lastTransaction"] = stats.get("lastTransaction")
    
    return expenses

@api_router.get("/expenses/with-next-date")
async def get_expenses_with_next_date(request: Request):
    """Get all expenses with calculated next deduction dates"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    expenses = await db.expenses.find(user_filter, {"_id": 0}).to_list(1000)
    
    result = []
    for expense in expenses:
        if isinstance(expense.get('createdAt'), str):
            expense['createdAt'] = datetime.fromisoformat(expense['createdAt'])
        
        # Calculate next deduction date
        next_date = calculate_next_deduction_date(expense)
        expense['nextDeductionDate'] = next_date
        
        # Check if linked to loan and get loan details
        if expense.get('linkedLoanId'):
            loan = await db.loans.find_one({"id": expense['linkedLoanId']}, {"_id": 0})
            if loan:
                expense['linkedLoanName'] = loan.get('loanName')
        
        # Check if linked to insurance and get insurance details
        if expense.get('linkedInsuranceId'):
            insurance = await db.insurances.find_one({"id": expense['linkedInsuranceId']}, {"_id": 0})
            if insurance:
                expense['linkedInsuranceName'] = insurance.get('policyName')
        
        result.append(expense)
    
    return result

@api_router.get("/expenses/{expense_id}", response_model=Expense)
async def get_expense(expense_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = expense_id
    expense = await db.expenses.find_one(user_filter, {"_id": 0})
    
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    if isinstance(expense.get('createdAt'), str):
        expense['createdAt'] = datetime.fromisoformat(expense['createdAt'])
    
    return expense

@api_router.put("/expenses/{expense_id}", response_model=Expense)
async def update_expense(expense_id: str, input: ExpenseCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = expense_id
    existing = await db.expenses.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    expense_dict = input.model_dump()
    expense_dict['id'] = expense_id
    expense_dict['userId'] = user.get('user_id')
    expense_dict['createdAt'] = existing['createdAt']
    
    await db.expenses.replace_one({"id": expense_id}, expense_dict)
    
    expense_obj = Expense(**expense_dict)
    if isinstance(expense_obj.createdAt, str):
        expense_obj.createdAt = datetime.fromisoformat(expense_obj.createdAt)
    
    return expense_obj

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = expense_id
    existing = await db.expenses.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    await db.expenses.delete_one({"id": expense_id})
    return {"message": "Expense deleted successfully", "id": expense_id}

# ============ INSURANCE ENDPOINTS ============

@api_router.post("/insurances", response_model=Insurance)
async def create_insurance(input: InsuranceCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    insurance_dict = input.model_dump()
    insurance_dict['userId'] = user.get('user_id')  # Add user isolation
    insurance_obj = Insurance(**insurance_dict)
    
    doc = insurance_obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    
    await db.insurances.insert_one(doc)
    
    # Auto-create Asset entry for Market Linked / Returns on Maturity insurances
    maturity_types_needing_asset = ["Market Linked", "Returns on Maturity"]
    if insurance_obj.maturityType in maturity_types_needing_asset:
        # Check if asset already exists for this insurance
        existing_asset = await db.assets.find_one({"linkedInsuranceId": insurance_obj.id}, {"_id": 0})
        if not existing_asset:
            # Determine current value - use expectedMaturityAmount or premiumAmount as initial value
            current_value = insurance_obj.expectedMaturityAmount or insurance_obj.premiumAmount
            
            asset_data = {
                "id": str(uuid.uuid4()),
                "assetType": "Insurance Asset",
                "assetName": f"{insurance_obj.policyName} (Maturity Value)",
                "currentValue": current_value,
                "purchaseValue": insurance_obj.premiumAmount,  # Initial investment
                "purchaseDate": insurance_obj.startDate,
                "isFinanced": False,
                "linkedLoanId": None,
                "isInsured": True,  # It's itself an insurance
                "linkedInsuranceId": insurance_obj.id,
                "generatesIncome": False,
                "incomeAmount": None,
                "incomeFrequency": None,
                "notes": f"Auto-created from {insurance_obj.insuranceType} policy - {insurance_obj.maturityType}",
                "createdAt": datetime.now(timezone.utc).isoformat()
            }
            await db.assets.insert_one(asset_data)
            
            # Update insurance with linked asset ID
            await db.insurances.update_one(
                {"id": insurance_obj.id},
                {"$set": {"linkedAssetId": asset_data["id"]}}
            )
    
    # Auto-create premium expense if enabled
    if insurance_obj.autoCreateExpense:
        # Check if expense already exists for this insurance
        existing_expense = await db.expenses.find_one({"linkedInsuranceId": insurance_obj.id}, {"_id": 0})
        if not existing_expense:
            # Map premium frequency to expense frequency
            freq_map = {
                "One-Time": "One-Time",
                "Monthly": "Monthly", 
                "Quarterly": "Quarterly", 
                "Half-Yearly": "Half-Yearly", 
                "Yearly": "Yearly"
            }
            expense_freq = freq_map.get(insurance_obj.premiumFrequency, "Yearly")
            
            # Calculate selectedDate from startDate
            start_date = datetime.fromisoformat(insurance_obj.startDate) if insurance_obj.startDate else datetime.now(timezone.utc)
            selected_date = str(start_date.day)
            selected_month = start_date.strftime("%B") if expense_freq == "Yearly" else None
            
            expense_data = {
                "id": str(uuid.uuid4()),
                "expenseName": f"{insurance_obj.policyName} Premium",
                "expenseType": "Fixed",
                "category": "Insurance",
                "expectedAmount": insurance_obj.premiumAmount,
                "frequency": expense_freq,
                "linkedAccountId": None,
                "linkedLoanId": None,
                "linkedInsuranceId": insurance_obj.id,
                "selectedDay": None,
                "selectedDate": selected_date,
                "selectedQuarter": None,
                "selectedHalf": None,
                "selectedMonth": selected_month,
                "oneTimeDate": insurance_obj.startDate if expense_freq == "One-Time" else None,
                "isPaid": False,
                "lastPaidDate": None,
                "createdAt": datetime.now(timezone.utc).isoformat()
            }
            await db.expenses.insert_one(expense_data)
    
    return insurance_obj

@api_router.get("/insurances", response_model=List[Insurance])
async def get_insurances(request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    insurances = await db.insurances.find(user_filter, {"_id": 0}).to_list(1000)
    
    for insurance in insurances:
        if isinstance(insurance['createdAt'], str):
            insurance['createdAt'] = datetime.fromisoformat(insurance['createdAt'])
    
    return insurances

@api_router.get("/insurances/{insurance_id}", response_model=Insurance)
async def get_insurance(insurance_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = insurance_id
    insurance = await db.insurances.find_one(user_filter, {"_id": 0})
    
    if not insurance:
        raise HTTPException(status_code=404, detail="Insurance not found")
    
    if isinstance(insurance['createdAt'], str):
        insurance['createdAt'] = datetime.fromisoformat(insurance['createdAt'])
    
    return insurance

@api_router.put("/insurances/{insurance_id}", response_model=Insurance)
async def update_insurance(insurance_id: str, input: InsuranceCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = insurance_id
    existing = await db.insurances.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Insurance not found")
    
    insurance_dict = input.model_dump()
    insurance_dict['id'] = insurance_id
    insurance_dict['userId'] = user.get('user_id')
    insurance_dict['createdAt'] = existing['createdAt']
    
    await db.insurances.replace_one({"id": insurance_id}, insurance_dict)
    
    insurance_obj = Insurance(**insurance_dict)
    if isinstance(insurance_obj.createdAt, str):
        insurance_obj.createdAt = datetime.fromisoformat(insurance_obj.createdAt)
    
    return insurance_obj

@api_router.delete("/insurances/{insurance_id}")
async def delete_insurance(insurance_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = insurance_id
    existing = await db.insurances.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Insurance not found")
    
    await db.insurances.delete_one({"id": insurance_id})
    return {"message": "Insurance deleted successfully", "id": insurance_id}

# ============ EXPENSE SCHEDULER ENDPOINTS ============

def calculate_next_deduction_date(expense: dict) -> Optional[str]:
    """Calculate the next deduction date for an expense based on its frequency"""
    from calendar import monthrange
    
    today = datetime.now(timezone.utc).date()
    frequency = expense.get('frequency', '')
    
    if frequency == "Daily":
        return today.isoformat()
    
    elif frequency == "Weekly":
        selected_day = expense.get('selectedDay', '')
        if not selected_day:
            return None
        day_mapping = {"Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3, "Friday": 4, "Saturday": 5, "Sunday": 6}
        target_day = day_mapping.get(selected_day, 0)
        days_ahead = target_day - today.weekday()
        if days_ahead <= 0:
            days_ahead += 7
        next_date = today + timedelta(days=days_ahead)
        return next_date.isoformat()
    
    elif frequency == "Monthly":
        selected_date = expense.get('selectedDate', '')
        if not selected_date:
            return None
        day = int(selected_date)
        # Get max days in current month
        _, max_day = monthrange(today.year, today.month)
        day = min(day, max_day)
        
        if today.day < day:
            next_date = today.replace(day=day)
        else:
            # Move to next month
            if today.month == 12:
                next_date = today.replace(year=today.year + 1, month=1, day=min(day, 31))
            else:
                _, max_next_day = monthrange(today.year, today.month + 1)
                next_date = today.replace(month=today.month + 1, day=min(day, max_next_day))
        return next_date.isoformat()
    
    elif frequency == "Quarterly":
        selected_quarter = expense.get('selectedQuarter', '')
        selected_date = expense.get('selectedDate', '')
        if not selected_date:
            return None
        day = int(selected_date)
        # Q1: Jan, Q2: Apr, Q3: Jul, Q4: Oct
        quarter_starts = {"Q1 (Jan–Mar)": 1, "Q2 (Apr–Jun)": 4, "Q3 (Jul–Sep)": 7, "Q4 (Oct–Dec)": 10}
        for q_name, start_month in quarter_starts.items():
            if selected_quarter and q_name.startswith(selected_quarter[:2]):
                # Find next occurrence
                for m in [start_month, start_month + 3, start_month + 6, start_month + 9]:
                    m = ((m - 1) % 12) + 1
                    year = today.year if m >= today.month else today.year + 1
                    _, max_day = monthrange(year, m)
                    target_day = min(day, max_day)
                    target_date = datetime(year, m, target_day).date()
                    if target_date > today:
                        return target_date.isoformat()
        return None
    
    elif frequency == "Half-Yearly":
        selected_half = expense.get('selectedHalf', '')
        selected_date = expense.get('selectedDate', '')
        if not selected_date:
            return None
        day = int(selected_date)
        # H1: Jan-Jun, H2: Jul-Dec
        if "Jan" in selected_half:
            months = [1, 7]
        else:
            months = [7, 1]
        for m in months:
            year = today.year if m >= today.month else today.year + 1
            _, max_day = monthrange(year, m)
            target_day = min(day, max_day)
            target_date = datetime(year, m, target_day).date()
            if target_date > today:
                return target_date.isoformat()
        return None
    
    elif frequency == "Yearly":
        selected_month = expense.get('selectedMonth', '')
        selected_date = expense.get('selectedDate', '')
        if not selected_month or not selected_date:
            return None
        month_mapping = {"January": 1, "February": 2, "March": 3, "April": 4, "May": 5, "June": 6, 
                        "July": 7, "August": 8, "September": 9, "October": 10, "November": 11, "December": 12}
        month = month_mapping.get(selected_month, 1)
        day = int(selected_date)
        year = today.year
        _, max_day = monthrange(year, month)
        target_day = min(day, max_day)
        target_date = datetime(year, month, target_day).date()
        if target_date <= today:
            target_date = datetime(year + 1, month, target_day).date()
        return target_date.isoformat()
    
    elif frequency == "One-Time":
        one_time_date = expense.get('oneTimeDate', '')
        if one_time_date:
            return one_time_date
        return None
    
    return None

@api_router.post("/expenses/process-deductions")
async def process_fixed_expense_deductions():
    """Process fixed expense deductions for today - to be called by a scheduler"""
    today = datetime.now(timezone.utc).date().isoformat()
    
    # Get all fixed expenses
    fixed_expenses = await db.expenses.find({"expenseType": "Fixed"}, {"_id": 0}).to_list(1000)
    
    processed = []
    errors = []
    
    for expense in fixed_expenses:
        try:
            next_date = calculate_next_deduction_date(expense)
            
            # Check if due today
            if next_date == today:
                linked_account_id = expense.get('linkedAccountId')
                amount = expense.get('expectedAmount', 0)
                
                if linked_account_id and amount > 0:
                    # Deduct from linked account
                    account = await db.accounts.find_one({"id": linked_account_id}, {"_id": 0})
                    if account:
                        new_balance = account.get('currentBalance', 0) - amount
                        await db.accounts.update_one(
                            {"id": linked_account_id},
                            {"$set": {"currentBalance": new_balance}}
                        )
                        
                        # Update expense as paid
                        await db.expenses.update_one(
                            {"id": expense['id']},
                            {"$set": {"isPaid": True, "lastPaidDate": today}}
                        )
                        
                        processed.append({
                            "expenseId": expense['id'],
                            "expenseName": expense.get('expenseName'),
                            "amount": amount,
                            "accountId": linked_account_id,
                            "accountName": account.get('accountName'),
                            "newBalance": new_balance
                        })
        except Exception as e:
            errors.append({
                "expenseId": expense.get('id'),
                "error": str(e)
            })
    
    return {
        "processedCount": len(processed),
        "processed": processed,
        "errors": errors,
        "processedDate": today
    }

# ============ INVESTMENT ENDPOINTS ============

@api_router.post("/investments", response_model=Investment)
async def create_investment(input: InvestmentCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    investment_dict = input.model_dump()
    investment_dict['userId'] = user.get('user_id')
    investment_obj = Investment(**investment_dict)
    
    doc = investment_obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    
    # Auto-create SIP expense if enabled
    linked_expense_id = None
    if input.autoCreateExpense and input.investmentFrequency and input.sipAmount:
        expense_name = f"SIP - {input.name}"
        expense_dict = {
            'expenseName': expense_name,
            'expenseType': 'Fixed',
            'category': 'Investments',
            'expectedAmount': input.sipAmount,
            'frequency': input.investmentFrequency,
            'linkedAccountId': input.linkedAccountId,
            'linkedInvestmentId': investment_obj.id,  # Link back to investment
            'selectedDay': input.sipSelectedDay,
            'selectedDate': input.sipSelectedDate,
            'isPaid': False,
            'userId': user.get('user_id')
        }
        expense_obj = Expense(**expense_dict)
        expense_doc = expense_obj.model_dump()
        expense_doc['createdAt'] = expense_doc['createdAt'].isoformat()
        await db.expenses.insert_one(expense_doc)
        linked_expense_id = expense_obj.id
        doc['linkedExpenseId'] = linked_expense_id
        investment_obj = Investment(**doc)
        investment_obj.createdAt = datetime.fromisoformat(doc['createdAt'])
    
    await db.investments.insert_one(doc)
    return investment_obj

@api_router.get("/investments", response_model=List[Investment])
async def get_investments(request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    investments = await db.investments.find(user_filter, {"_id": 0}).to_list(1000)
    
    for investment in investments:
        if isinstance(investment.get('createdAt'), str):
            investment['createdAt'] = datetime.fromisoformat(investment['createdAt'])
    
    return investments

@api_router.get("/investments/{investment_id}", response_model=Investment)
async def get_investment(investment_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = investment_id
    investment = await db.investments.find_one(user_filter, {"_id": 0})
    
    if not investment:
        raise HTTPException(status_code=404, detail="Investment not found")
    
    if isinstance(investment.get('createdAt'), str):
        investment['createdAt'] = datetime.fromisoformat(investment['createdAt'])
    
    return investment

@api_router.put("/investments/{investment_id}", response_model=Investment)
async def update_investment(investment_id: str, input: InvestmentCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = investment_id
    existing = await db.investments.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Investment not found")
    
    investment_dict = input.model_dump()
    investment_dict['id'] = investment_id
    investment_dict['userId'] = user.get('user_id')
    investment_dict['createdAt'] = existing['createdAt']
    
    await db.investments.replace_one({"id": investment_id}, investment_dict)
    
    investment_obj = Investment(**investment_dict)
    if isinstance(investment_obj.createdAt, str):
        investment_obj.createdAt = datetime.fromisoformat(investment_obj.createdAt)
    
    return investment_obj

@api_router.delete("/investments/{investment_id}")
async def delete_investment(investment_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = investment_id
    existing = await db.investments.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Investment not found")
    
    await db.investments.delete_one({"id": investment_id})
    return {"message": "Investment deleted successfully", "id": investment_id}

# ============ NET WORTH DASHBOARD ENDPOINTS ============

@api_router.get("/dashboard/networth")
async def get_networth_summary(request: Request):
    """Aggregate all financial data for net worth calculation"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    user_email = user.get('email', '')
    
    # User filter - strict isolation for new users, legacy data only for test user
    if user_email == 'test@moneyssutra.com' or user_id == 'test':
        # Test user can see legacy data (no userId)
        user_filter = {"$or": [{"userId": user_id}, {"userId": None}, {"userId": {"$exists": False}}]}
    else:
        # New users only see their own data
        user_filter = {"userId": user_id}
    
    # Run all database queries in parallel for faster response
    assets_task = db.assets.find(user_filter, {"_id": 0}).to_list(1000)
    investments_task = db.investments.find(user_filter, {"_id": 0}).to_list(1000)
    accounts_task = db.accounts.find(user_filter, {"_id": 0}).to_list(1000)
    credit_cards_task = db.credit_cards.find(user_filter, {"_id": 0}).to_list(1000)
    loans_task = db.loans.find(user_filter, {"_id": 0}).to_list(1000)
    incomes_task = db.income_sources.find(user_filter, {"_id": 0}).to_list(1000)
    other_incomes_task = db.other_income.find(user_filter, {"_id": 0}).to_list(1000)
    expenses_task = db.expenses.find(user_filter, {"_id": 0}).to_list(1000)
    
    # Await all queries together
    assets, investments, accounts, credit_cards, loans, incomes, other_incomes, expenses = await asyncio.gather(
        assets_task, investments_task, accounts_task, credit_cards_task,
        loans_task, incomes_task, other_incomes_task, expenses_task
    )
    
    # Process results
    total_assets = sum(asset.get('currentValue', 0) for asset in assets)
    total_investments = sum(inv.get('currentValue', 0) for inv in investments)
    
    liquid_balance = sum(
        acc.get('currentBalance', 0) for acc in accounts 
        if acc.get('accountType') != 'Credit Card'
    )
    credit_outstanding = sum(
        acc.get('outstandingAmount', 0) or 0 for acc in accounts 
        if acc.get('accountType') == 'Credit Card'
    )
    
    credit_card_outstanding = sum(card.get('outstandingAmount', 0) for card in credit_cards)
    credit_card_limit = sum(card.get('creditLimit', 0) for card in credit_cards)
    
    total_liabilities = sum(loan.get('outstandingAmount', 0) for loan in loans)
    total_liabilities += credit_outstanding + credit_card_outstanding
    
    # Calculate monthly income for current month
    monthly_income = 0
    current_month = datetime.now(timezone.utc).month
    current_year = datetime.now(timezone.utc).year
    
    for income in incomes:
        amount = income.get('expectedAmount', 0)
        freq = income.get('frequency', 'Monthly')
        
        # For each frequency, determine if this income would be received this month
        if freq == 'Daily':
            # Daily income is received every day
            monthly_income += amount * 30
        elif freq == 'Weekly':
            # Weekly income is received ~4 times per month
            monthly_income += amount * 4
        elif freq == 'Monthly':
            # Monthly income is always received once per month
            monthly_income += amount
        elif freq == 'Quarterly':
            # Check if current month is a payment month (Jan, Apr, Jul, Oct typically)
            selected_quarter = income.get('selectedQuarter', '')
            # For quarterly, check if it matches this month's quarter
            quarter_months = {
                'Q1': [1, 2, 3],
                'Q2': [4, 5, 6],
                'Q3': [7, 8, 9],
                'Q4': [10, 11, 12]
            }
            for q_prefix, months in quarter_months.items():
                if selected_quarter and selected_quarter.startswith(q_prefix):
                    # First month of each quarter is the payment month
                    if current_month == months[0]:
                        monthly_income += amount
                    break
            else:
                # If no quarter specified, assume first month of each quarter
                if current_month in [1, 4, 7, 10]:
                    monthly_income += amount
        elif freq == 'Half-Yearly':
            # Check if current month is a payment month (Jan and Jul typically)
            selected_half = income.get('selectedHalf', '')
            if 'Jan' in selected_half:
                if current_month in [1, 7]:
                    monthly_income += amount
            else:
                if current_month in [7, 1]:
                    monthly_income += amount
        elif freq == 'Yearly':
            # Check if current month matches the selected month
            selected_month = income.get('selectedMonth', '')
            month_mapping = {
                "January": 1, "February": 2, "March": 3, "April": 4, 
                "May": 5, "June": 6, "July": 7, "August": 8, 
                "September": 9, "October": 10, "November": 11, "December": 12
            }
            if month_mapping.get(selected_month) == current_month:
                monthly_income += amount
        elif freq == 'Irregular' or freq == 'Others':
            # For irregular income, use custom date if it falls in current month
            custom_date = income.get('customDate', '')
            if custom_date:
                try:
                    date_obj = datetime.fromisoformat(custom_date).date()
                    if date_obj.month == current_month and date_obj.year == current_year:
                        monthly_income += amount
                except (ValueError, TypeError):
                    pass
        else:
            # Default case - assume monthly
            monthly_income += amount
    
    # Calculate other income for current month
    other_income_total = 0
    for other_inc in other_incomes:
        amount = other_inc.get('amount', 0)
        freq = other_inc.get('frequency', 'One-Time')
        
        if freq == 'One-Time':
            # Check if one-time income falls in current month
            date_received = other_inc.get('dateReceived', '')
            if date_received:
                try:
                    date_obj = datetime.fromisoformat(date_received).date()
                    if date_obj.month == current_month and date_obj.year == current_year:
                        other_income_total += amount
                except (ValueError, TypeError):
                    pass
        elif freq == 'Monthly':
            other_income_total += amount
        elif freq == 'Quarterly':
            selected_quarter = other_inc.get('selectedQuarter', '')
            quarter_months = {
                'Q1': [1, 2, 3], 'Q2': [4, 5, 6], 'Q3': [7, 8, 9], 'Q4': [10, 11, 12]
            }
            for q_prefix, months in quarter_months.items():
                if selected_quarter and selected_quarter.startswith(q_prefix):
                    if current_month == months[0]:
                        other_income_total += amount
                    break
            else:
                if current_month in [1, 4, 7, 10]:
                    other_income_total += amount
        elif freq == 'Yearly':
            selected_month = other_inc.get('selectedMonth', '')
            month_mapping = {
                "January": 1, "February": 2, "March": 3, "April": 4,
                "May": 5, "June": 6, "July": 7, "August": 8,
                "September": 9, "October": 10, "November": 11, "December": 12
            }
            if month_mapping.get(selected_month) == current_month:
                other_income_total += amount
        elif freq == 'Irregular':
            date_received = other_inc.get('dateReceived', '')
            if date_received:
                try:
                    date_obj = datetime.fromisoformat(date_received).date()
                    if date_obj.month == current_month and date_obj.year == current_year:
                        other_income_total += amount
                except (ValueError, TypeError):
                    pass
    
    # Add other income to monthly income
    monthly_income += other_income_total
    
    # Process expenses - Calculate actual monthly expenses for current month
    monthly_expenses = 0
    for expense in expenses:
        amount = expense.get('expectedAmount', 0)
        freq = expense.get('frequency', 'Monthly')
        
        # For each frequency, determine if this expense would occur this month
        if freq == 'Daily':
            monthly_expenses += amount * 30
        elif freq == 'Weekly':
            monthly_expenses += amount * 4
        elif freq == 'Monthly':
            monthly_expenses += amount
        elif freq == 'Quarterly':
            # Check if current month is a payment month
            selected_quarter = expense.get('selectedQuarter') or ''
            quarter_months = {
                'Q1': [1, 2, 3],
                'Q2': [4, 5, 6],
                'Q3': [7, 8, 9],
                'Q4': [10, 11, 12]
            }
            for q_prefix, months in quarter_months.items():
                if selected_quarter and selected_quarter.startswith(q_prefix):
                    if current_month == months[0]:
                        monthly_expenses += amount
                    break
            else:
                if current_month in [1, 4, 7, 10]:
                    monthly_expenses += amount
        elif freq == 'Half-Yearly':
            selected_half = expense.get('selectedHalf') or ''
            if 'Jan' in selected_half:
                if current_month in [1, 7]:
                    monthly_expenses += amount
            else:
                if current_month in [7, 1]:
                    monthly_expenses += amount
        elif freq == 'Yearly':
            selected_month = expense.get('selectedMonth') or ''
            month_mapping = {
                "January": 1, "February": 2, "March": 3, "April": 4, 
                "May": 5, "June": 6, "July": 7, "August": 8, 
                "September": 9, "October": 10, "November": 11, "December": 12
            }
            if month_mapping.get(selected_month) == current_month:
                monthly_expenses += amount
        elif freq == 'One-Time':
            # Check if one-time expense falls in current month
            one_time_date = expense.get('oneTimeDate', '')
            if one_time_date:
                try:
                    date_obj = datetime.fromisoformat(one_time_date).date()
                    if date_obj.month == current_month and date_obj.year == current_year:
                        monthly_expenses += amount
                except (ValueError, TypeError):
                    pass
        else:
            monthly_expenses += amount
    
    # Calculate net worth
    net_worth = total_assets + total_investments + liquid_balance - total_liabilities
    
    return {
        "netWorth": net_worth,
        "totalAssets": total_assets,
        "totalInvestments": total_investments,
        "liquidBalance": liquid_balance,
        "totalLiabilities": total_liabilities,
        "creditOutstanding": credit_outstanding,
        "creditCardOutstanding": credit_card_outstanding,
        "creditCardLimit": credit_card_limit,
        "creditCardUtilization": (credit_card_outstanding / credit_card_limit * 100) if credit_card_limit > 0 else 0,
        "monthlyIncome": monthly_income,
        "monthlyExpenses": monthly_expenses,
        "monthlySavings": monthly_income - monthly_expenses,
        "assetCount": len(assets),
        "investmentCount": len(investments),
        "accountCount": len(accounts),
        "loanCount": len(loans),
        "creditCardCount": len(credit_cards),
        "incomeCount": len(incomes),
        "expenseCount": len(expenses)
    }

@api_router.get("/dashboard/breakdown")
async def get_breakdown():
    """Get detailed breakdown by category"""
    
    # Asset breakdown by type
    assets = await db.assets.find({}, {"_id": 0}).to_list(1000)
    asset_breakdown = {}
    for asset in assets:
        asset_type = asset.get('assetType', 'Other')
        if asset_type not in asset_breakdown:
            asset_breakdown[asset_type] = 0
        asset_breakdown[asset_type] += asset.get('currentValue', 0)
    
    # Investment breakdown by category
    investments = await db.investments.find({}, {"_id": 0}).to_list(1000)
    investment_breakdown = {}
    for inv in investments:
        category = inv.get('investmentCategory', 'Other')
        if category not in investment_breakdown:
            investment_breakdown[category] = 0
        investment_breakdown[category] += inv.get('currentValue', 0)
    
    # Loan breakdown by type
    loans = await db.loans.find({}, {"_id": 0}).to_list(1000)
    loan_breakdown = {}
    for loan in loans:
        loan_type = loan.get('loanType', 'Other')
        if loan_type not in loan_breakdown:
            loan_breakdown[loan_type] = 0
        loan_breakdown[loan_type] += loan.get('outstandingAmount', 0)
    
    # Income breakdown by type
    incomes = await db.income_sources.find({}, {"_id": 0}).to_list(1000)
    income_breakdown = {}
    for income in incomes:
        income_type = income.get('type', 'Other')
        if income_type not in income_breakdown:
            income_breakdown[income_type] = 0
        income_breakdown[income_type] += income.get('expectedAmount', 0)
    
    # Expense breakdown by category
    expenses = await db.expenses.find({}, {"_id": 0}).to_list(1000)
    expense_breakdown = {}
    for expense in expenses:
        category = expense.get('category', 'Other')
        if category not in expense_breakdown:
            expense_breakdown[category] = 0
        expense_breakdown[category] += expense.get('expectedAmount', 0)
    
    return {
        "assetBreakdown": asset_breakdown,
        "investmentBreakdown": investment_breakdown,
        "loanBreakdown": loan_breakdown,
        "incomeBreakdown": income_breakdown,
        "expenseBreakdown": expense_breakdown
    }

# ============ USER PROFILE ENDPOINTS ============

# Profile Models
class BasicProfile(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    fullName: str
    monthlyIncome: float
    primaryGoals: List[str]
    riskAppetite: str
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BasicProfileCreate(BaseModel):
    fullName: str
    monthlyIncome: float
    primaryGoals: List[str]
    riskAppetite: str

class ExtendedProfile(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    userId: str
    dob: Optional[str] = None
    maritalStatus: Optional[str] = None
    dependents: Optional[int] = None
    retirementAge: Optional[int] = None
    emergencyFundTarget: Optional[str] = None
    debtComfortLevel: Optional[float] = None
    equityTarget: Optional[float] = None
    debtTarget: Optional[float] = None
    goldTarget: Optional[float] = None
    existingLifeCover: Optional[float] = None
    existingHealthCover: Optional[float] = None
    updatedAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ExtendedProfileCreate(BaseModel):
    dob: Optional[str] = None
    maritalStatus: Optional[str] = None
    dependents: Optional[int] = None
    retirementAge: Optional[int] = None
    emergencyFundTarget: Optional[str] = None
    debtComfortLevel: Optional[float] = None
    equityTarget: Optional[float] = None
    debtTarget: Optional[float] = None
    goldTarget: Optional[float] = None
    existingLifeCover: Optional[float] = None
    existingHealthCover: Optional[float] = None

@api_router.post("/profile/basic", response_model=BasicProfile)
async def create_basic_profile(input: BasicProfileCreate):
    # Check if profile exists
    existing = await db.profiles.find_one({}, {"_id": 0})
    if existing:
        # Update existing profile
        profile_dict = input.model_dump()
        profile_dict['id'] = existing['id']
        profile_dict['createdAt'] = existing['createdAt']
        await db.profiles.replace_one({"id": existing['id']}, profile_dict)
        return BasicProfile(**profile_dict)
    
    profile_dict = input.model_dump()
    profile_obj = BasicProfile(**profile_dict)
    
    doc = profile_obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    
    await db.profiles.insert_one(doc)
    return profile_obj

@api_router.get("/profile/basic")
async def get_basic_profile():
    profile = await db.profiles.find_one({}, {"_id": 0})
    
    if not profile:
        return None
    
    if isinstance(profile.get('createdAt'), str):
        profile['createdAt'] = datetime.fromisoformat(profile['createdAt'])
    
    return profile

@api_router.put("/profile/extended")
async def update_extended_profile(input: ExtendedProfileCreate):
    # Get basic profile first
    basic = await db.profiles.find_one({}, {"_id": 0})
    if not basic:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Basic profile not found. Complete basic setup first.")
    
    # Check if extended profile exists
    existing = await db.extended_profiles.find_one({"userId": basic['id']}, {"_id": 0})
    
    profile_dict = input.model_dump()
    profile_dict['userId'] = basic['id']
    profile_dict['updatedAt'] = datetime.now(timezone.utc).isoformat()
    
    if existing:
        await db.extended_profiles.replace_one({"userId": basic['id']}, profile_dict)
    else:
        await db.extended_profiles.insert_one(profile_dict)
    
    return profile_dict

@api_router.get("/profile/extended")
async def get_extended_profile():
    basic = await db.profiles.find_one({}, {"_id": 0})
    if not basic:
        return None
    
    extended = await db.extended_profiles.find_one({"userId": basic['id']}, {"_id": 0})
    return extended

@api_router.get("/profile/completion")
async def get_profile_completion():
    """Calculate profile completion percentage"""
    basic = await db.profiles.find_one({}, {"_id": 0})
    extended = await db.extended_profiles.find_one({}, {"_id": 0}) if basic else None
    
    completion = 0
    
    # Basic Info (25%)
    if basic:
        if basic.get('fullName'):
            completion += 10
        if basic.get('monthlyIncome'):
            completion += 10
        if basic.get('riskAppetite'):
            completion += 5
    
    # Extended profile fields
    if extended:
        if extended.get('dob'):
            completion += 5
        if extended.get('maritalStatus'):
            completion += 5
        if extended.get('dependents') is not None:
            completion += 5
        if extended.get('retirementAge'):
            completion += 10
        if extended.get('emergencyFundTarget'):
            completion += 5
        if extended.get('debtComfortLevel') is not None:
            completion += 5
        if extended.get('equityTarget') is not None:
            completion += 5
        if extended.get('debtTarget') is not None:
            completion += 5
        if extended.get('goldTarget') is not None:
            completion += 5
        if extended.get('existingLifeCover') is not None:
            completion += 10
        if extended.get('existingHealthCover') is not None:
            completion += 5
    
    return {
        "completion": min(completion, 100),
        "hasBasicProfile": basic is not None,
        "hasExtendedProfile": extended is not None
    }

# ============ CREDIT CARD ENDPOINTS ============

@api_router.post("/credit-cards", response_model=CreditCard)
async def create_credit_card(input: CreditCardCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    card_dict = input.model_dump()
    card_dict['userId'] = user.get('user_id')  # Add user isolation
    card_obj = CreditCard(**card_dict)
    
    doc = card_obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    
    await db.credit_cards.insert_one(doc)
    return card_obj

@api_router.get("/credit-cards", response_model=List[CreditCard])
async def get_credit_cards(request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    cards = await db.credit_cards.find(user_filter, {"_id": 0}).to_list(1000)
    
    for card in cards:
        if isinstance(card.get('createdAt'), str):
            card['createdAt'] = datetime.fromisoformat(card['createdAt'])
    
    return cards

@api_router.get("/credit-cards/{card_id}", response_model=CreditCard)
async def get_credit_card(card_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = card_id
    card = await db.credit_cards.find_one(user_filter, {"_id": 0})
    
    if not card:
        raise HTTPException(status_code=404, detail="Credit card not found")
    
    if isinstance(card.get('createdAt'), str):
        card['createdAt'] = datetime.fromisoformat(card['createdAt'])
    
    return card

@api_router.put("/credit-cards/{card_id}", response_model=CreditCard)
async def update_credit_card(card_id: str, input: CreditCardCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = card_id
    existing = await db.credit_cards.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Credit card not found")
    
    card_dict = input.model_dump()
    card_dict['id'] = card_id
    card_dict['userId'] = user.get('user_id')
    card_dict['createdAt'] = existing['createdAt']
    
    await db.credit_cards.replace_one({"id": card_id}, card_dict)
    
    card_obj = CreditCard(**card_dict)
    if isinstance(card_obj.createdAt, str):
        card_obj.createdAt = datetime.fromisoformat(card_obj.createdAt)
    
    return card_obj

@api_router.delete("/credit-cards/{card_id}")
async def delete_credit_card(card_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = card_id
    existing = await db.credit_cards.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Credit card not found")
    
    await db.credit_cards.delete_one({"id": card_id})
    return {"message": "Credit card deleted successfully", "id": card_id}

# ============ GOAL ENDPOINTS ============

async def calculate_goal_progress(goal: dict) -> dict:
    """Calculate the current progress of a goal based on linked sources"""
    calculated_amount = 0
    linked_details = []
    sip_projections = []
    
    # If manual override is set, use the stored currentAmount
    if goal.get('manualOverride') and not goal.get('autoCalculate'):
        return {
            "currentAmount": goal.get('currentAmount', 0),
            "linkedDetails": [],
            "sipProjections": [],
            "calculationMethod": "manual"
        }
    
    goal_type = goal.get('goalType', '')
    target_date = goal.get('targetDate')
    
    # Calculate months until target date for SIP projections
    months_to_target = 0
    if target_date:
        try:
            target_dt = datetime.fromisoformat(target_date).date()
            today = datetime.now(timezone.utc).date()
            days_diff = (target_dt - today).days
            months_to_target = max(0, days_diff / 30)
        except (ValueError, TypeError):
            months_to_target = 0
    
    # For Debt Elimination goals - track loan/credit card payoff
    if goal_type == "Debt Elimination":
        if goal.get('linkedLoanId'):
            loan = await db.loans.find_one({"id": goal['linkedLoanId']}, {"_id": 0})
            if loan:
                # Progress = Principal - Outstanding (amount paid off)
                principal = loan.get('principalAmount', 0)
                outstanding = loan.get('outstandingAmount', 0)
                paid_off = principal - outstanding
                calculated_amount += paid_off
                linked_details.append({
                    "type": "Loan",
                    "name": loan.get('loanName'),
                    "contribution": paid_off,
                    "principal": principal,
                    "outstanding": outstanding,
                    "emiAmount": loan.get('emiAmount', 0)
                })
        
        if goal.get('linkedCreditCardId'):
            card = await db.credit_cards.find_one({"id": goal['linkedCreditCardId']}, {"_id": 0})
            if card:
                # For credit card, progress is credit limit - outstanding
                limit = card.get('creditLimit', 0)
                outstanding = card.get('outstandingAmount', 0)
                available = limit - outstanding
                calculated_amount += available
                linked_details.append({
                    "type": "Credit Card",
                    "name": card.get('cardName'),
                    "contribution": available,
                    "creditLimit": limit,
                    "outstanding": outstanding
                })
    
    # For Investment Target / Wealth Creation goals
    elif goal_type in ["Investment Target", "Wealth Creation"]:
        # Sum up linked investments' current values and calculate SIP projections
        linked_investment_ids = goal.get('linkedInvestmentIds', [])
        for inv_id in linked_investment_ids:
            investment = await db.investments.find_one({"id": inv_id}, {"_id": 0})
            if investment:
                current_value = investment.get('currentValue', 0)
                calculated_amount += current_value
                
                # Calculate SIP projection if investment has recurring frequency
                sip_amount = investment.get('sipAmount', 0)
                frequency = investment.get('investmentFrequency', '')
                return_rate = investment.get('returnRate', 0)
                
                projected_value = current_value
                monthly_contribution = 0
                
                if sip_amount and frequency and months_to_target > 0:
                    # Convert frequency to monthly
                    if frequency == 'Daily':
                        monthly_contribution = sip_amount * 30
                    elif frequency == 'Weekly':
                        monthly_contribution = sip_amount * 4
                    elif frequency == 'Monthly':
                        monthly_contribution = sip_amount
                    elif frequency == 'Quarterly':
                        monthly_contribution = sip_amount / 3
                    elif frequency == 'Yearly':
                        monthly_contribution = sip_amount / 12
                    
                    # Calculate future value with compound growth
                    # FV = PV * (1 + r)^n + PMT * [((1 + r)^n - 1) / r]
                    monthly_rate = (return_rate / 100) / 12 if return_rate else 0
                    n = months_to_target
                    
                    if monthly_rate > 0:
                        # Compound growth of current value
                        pv_growth = current_value * ((1 + monthly_rate) ** n)
                        # Future value of monthly contributions
                        pmt_growth = monthly_contribution * (((1 + monthly_rate) ** n - 1) / monthly_rate)
                        projected_value = pv_growth + pmt_growth
                    else:
                        # No growth rate, simple addition
                        projected_value = current_value + (monthly_contribution * n)
                    
                    sip_projections.append({
                        "investmentId": inv_id,
                        "investmentName": investment.get('name'),
                        "currentValue": current_value,
                        "sipAmount": sip_amount,
                        "frequency": frequency,
                        "monthlyContribution": monthly_contribution,
                        "returnRate": return_rate,
                        "projectedValue": round(projected_value, 2),
                        "projectedGain": round(projected_value - current_value, 2),
                        "monthsToTarget": round(months_to_target, 1)
                    })
                
                linked_details.append({
                    "type": "Investment",
                    "name": investment.get('name'),
                    "category": investment.get('investmentCategory'),
                    "contribution": current_value,
                    "principal": investment.get('principal', 0),
                    "hasSIP": bool(sip_amount and frequency),
                    "sipAmount": sip_amount,
                    "frequency": frequency,
                    "projectedValue": round(projected_value, 2) if sip_amount else None
                })
        
        # Also add linked account balances - support both new and legacy format
        # New format: linkedAccounts with allocatedAmount
        for linked_acc in goal.get('linkedAccounts', []):
            acc_id = linked_acc.get('id')
            allocated = linked_acc.get('allocatedAmount', 0)
            account = await db.accounts.find_one({"id": acc_id}, {"_id": 0})
            if account and account.get('accountType') != 'Credit Card' and allocated > 0:
                calculated_amount += allocated
                linked_details.append({
                    "type": "Account",
                    "name": account.get('accountName'),
                    "accountType": account.get('accountType'),
                    "contribution": allocated,
                    "totalBalance": account.get('currentBalance', 0),
                    "isPartialAllocation": allocated < account.get('currentBalance', 0)
                })
        
        # Legacy format: linkedAccountIds (full amount)
        processed_acc_ids = [a.get('id') for a in goal.get('linkedAccounts', [])]
        linked_account_ids = goal.get('linkedAccountIds', [])
        for acc_id in linked_account_ids:
            if acc_id not in processed_acc_ids:
                account = await db.accounts.find_one({"id": acc_id}, {"_id": 0})
                if account and account.get('accountType') != 'Credit Card':
                    balance = account.get('currentBalance', 0)
                    calculated_amount += balance
                    linked_details.append({
                        "type": "Account",
                        "name": account.get('accountName'),
                        "accountType": account.get('accountType'),
                        "contribution": balance,
                        "isLegacy": True
                    })
    
    # For Emergency Fund goals
    elif goal_type == "Emergency Fund":
        # New format: linkedAccounts with allocatedAmount
        for linked_acc in goal.get('linkedAccounts', []):
            acc_id = linked_acc.get('id')
            allocated = linked_acc.get('allocatedAmount', 0)
            account = await db.accounts.find_one({"id": acc_id}, {"_id": 0})
            if account and account.get('accountType') != 'Credit Card' and allocated > 0:
                calculated_amount += allocated
                linked_details.append({
                    "type": "Account",
                    "name": account.get('accountName'),
                    "accountType": account.get('accountType'),
                    "contribution": allocated,
                    "totalBalance": account.get('currentBalance', 0),
                    "isPartialAllocation": allocated < account.get('currentBalance', 0)
                })
        
        # Legacy format: linkedAccountIds (full amount)
        processed_acc_ids = [a.get('id') for a in goal.get('linkedAccounts', [])]
        linked_account_ids = goal.get('linkedAccountIds', [])
        for acc_id in linked_account_ids:
            if acc_id not in processed_acc_ids:
                account = await db.accounts.find_one({"id": acc_id}, {"_id": 0})
                if account and account.get('accountType') != 'Credit Card':
                    balance = account.get('currentBalance', 0)
                    calculated_amount += balance
                    linked_details.append({
                        "type": "Account",
                        "name": account.get('accountName'),
                        "accountType": account.get('accountType'),
                        "contribution": balance,
                        "isLegacy": True
                    })
        
        # New format: linkedInvestments with allocatedAmount
        for linked_inv in goal.get('linkedInvestments', []):
            inv_id = linked_inv.get('id')
            allocated = linked_inv.get('allocatedAmount', 0)
            investment = await db.investments.find_one({"id": inv_id}, {"_id": 0})
            if investment and allocated > 0:
                calculated_amount += allocated
                
                # Calculate SIP projection proportionally
                current_value = investment.get('currentValue', 0)
                allocation_ratio = allocated / current_value if current_value > 0 else 0
                sip_amount = investment.get('sipAmount', 0)
                frequency = investment.get('investmentFrequency', '')
                return_rate = investment.get('returnRate', 0)
                
                projected_value = allocated
                monthly_contribution = 0
                
                if sip_amount and frequency and months_to_target > 0:
                    if frequency == 'Daily':
                        monthly_contribution = sip_amount * 30 * allocation_ratio
                    elif frequency == 'Weekly':
                        monthly_contribution = sip_amount * 4 * allocation_ratio
                    elif frequency == 'Monthly':
                        monthly_contribution = sip_amount * allocation_ratio
                    elif frequency == 'Quarterly':
                        monthly_contribution = (sip_amount / 3) * allocation_ratio
                    elif frequency == 'Yearly':
                        monthly_contribution = (sip_amount / 12) * allocation_ratio
                    
                    monthly_rate = (return_rate / 100) / 12 if return_rate else 0
                    n = months_to_target
                    
                    if monthly_rate > 0:
                        pv_growth = allocated * ((1 + monthly_rate) ** n)
                        pmt_growth = monthly_contribution * (((1 + monthly_rate) ** n - 1) / monthly_rate)
                        projected_value = pv_growth + pmt_growth
                    else:
                        projected_value = allocated + (monthly_contribution * n)
                    
                    sip_projections.append({
                        "investmentId": inv_id,
                        "investmentName": investment.get('name'),
                        "allocatedAmount": allocated,
                        "currentValue": current_value,
                        "sipAmount": sip_amount,
                        "frequency": frequency,
                        "monthlyContribution": monthly_contribution,
                        "returnRate": return_rate,
                        "projectedValue": round(projected_value, 2),
                        "projectedGain": round(projected_value - allocated, 2),
                        "monthsToTarget": round(months_to_target, 1)
                    })
                
                linked_details.append({
                    "type": "Investment",
                    "name": investment.get('name'),
                    "category": investment.get('investmentCategory'),
                    "contribution": allocated,
                    "totalValue": current_value,
                    "isPartialAllocation": allocated < current_value,
                    "hasSIP": bool(sip_amount and frequency),
                    "sipAmount": sip_amount,
                    "frequency": frequency,
                    "projectedValue": round(projected_value, 2) if sip_amount else None
                })
        
        # Legacy format: linkedInvestmentIds (full amount)
        processed_inv_ids = [i.get('id') for i in goal.get('linkedInvestments', [])]
        linked_investment_ids = goal.get('linkedInvestmentIds', [])
        for inv_id in linked_investment_ids:
            if inv_id not in processed_inv_ids:
                investment = await db.investments.find_one({"id": inv_id}, {"_id": 0})
                if investment:
                    current_value = investment.get('currentValue', 0)
                    calculated_amount += current_value
                    
                    sip_amount = investment.get('sipAmount', 0)
                    frequency = investment.get('investmentFrequency', '')
                    return_rate = investment.get('returnRate', 0)
                    
                    projected_value = current_value
                    monthly_contribution = 0
                    
                    if sip_amount and frequency and months_to_target > 0:
                        if frequency == 'Daily':
                            monthly_contribution = sip_amount * 30
                        elif frequency == 'Weekly':
                            monthly_contribution = sip_amount * 4
                        elif frequency == 'Monthly':
                            monthly_contribution = sip_amount
                        elif frequency == 'Quarterly':
                            monthly_contribution = sip_amount / 3
                        elif frequency == 'Yearly':
                            monthly_contribution = sip_amount / 12
                        
                        monthly_rate = (return_rate / 100) / 12 if return_rate else 0
                        n = months_to_target
                        
                        if monthly_rate > 0:
                            pv_growth = current_value * ((1 + monthly_rate) ** n)
                            pmt_growth = monthly_contribution * (((1 + monthly_rate) ** n - 1) / monthly_rate)
                            projected_value = pv_growth + pmt_growth
                        else:
                            projected_value = current_value + (monthly_contribution * n)
                        
                        sip_projections.append({
                            "investmentId": inv_id,
                            "investmentName": investment.get('name'),
                            "currentValue": current_value,
                            "sipAmount": sip_amount,
                            "frequency": frequency,
                            "monthlyContribution": monthly_contribution,
                            "returnRate": return_rate,
                            "projectedValue": round(projected_value, 2),
                            "projectedGain": round(projected_value - current_value, 2),
                            "monthsToTarget": round(months_to_target, 1),
                            "isLegacy": True
                        })
                    
                    linked_details.append({
                        "type": "Investment",
                        "name": investment.get('name'),
                        "category": investment.get('investmentCategory'),
                        "contribution": current_value,
                        "principal": investment.get('principal', 0),
                        "hasSIP": bool(sip_amount and frequency),
                        "sipAmount": sip_amount,
                        "frequency": frequency,
                        "projectedValue": round(projected_value, 2) if sip_amount else None,
                        "isLegacy": True
                    })
    
    # For Other/Custom goals - use linked sources or manual amount
    else:
        # Sum up any linked investments with SIP projections
        for inv_id in goal.get('linkedInvestmentIds', []):
            investment = await db.investments.find_one({"id": inv_id}, {"_id": 0})
            if investment:
                current_value = investment.get('currentValue', 0)
                calculated_amount += current_value
                
                # Calculate SIP projection
                sip_amount = investment.get('sipAmount', 0)
                frequency = investment.get('investmentFrequency', '')
                return_rate = investment.get('returnRate', 0)
                
                projected_value = current_value
                monthly_contribution = 0
                
                if sip_amount and frequency and months_to_target > 0:
                    if frequency == 'Daily':
                        monthly_contribution = sip_amount * 30
                    elif frequency == 'Weekly':
                        monthly_contribution = sip_amount * 4
                    elif frequency == 'Monthly':
                        monthly_contribution = sip_amount
                    elif frequency == 'Quarterly':
                        monthly_contribution = sip_amount / 3
                    elif frequency == 'Yearly':
                        monthly_contribution = sip_amount / 12
                    
                    monthly_rate = (return_rate / 100) / 12 if return_rate else 0
                    n = months_to_target
                    
                    if monthly_rate > 0:
                        pv_growth = current_value * ((1 + monthly_rate) ** n)
                        pmt_growth = monthly_contribution * (((1 + monthly_rate) ** n - 1) / monthly_rate)
                        projected_value = pv_growth + pmt_growth
                    else:
                        projected_value = current_value + (monthly_contribution * n)
                    
                    sip_projections.append({
                        "investmentId": inv_id,
                        "investmentName": investment.get('name'),
                        "currentValue": current_value,
                        "sipAmount": sip_amount,
                        "frequency": frequency,
                        "monthlyContribution": monthly_contribution,
                        "returnRate": return_rate,
                        "projectedValue": round(projected_value, 2),
                        "projectedGain": round(projected_value - current_value, 2),
                        "monthsToTarget": round(months_to_target, 1)
                    })
                
                linked_details.append({
                    "type": "Investment",
                    "name": investment.get('name'),
                    "contribution": current_value,
                    "hasSIP": bool(sip_amount and frequency),
                    "projectedValue": round(projected_value, 2) if sip_amount else None
                })
        
        # Sum up linked accounts
        for acc_id in goal.get('linkedAccountIds', []):
            account = await db.accounts.find_one({"id": acc_id}, {"_id": 0})
            if account and account.get('accountType') != 'Credit Card':
                calculated_amount += account.get('currentBalance', 0)
                linked_details.append({
                    "type": "Account",
                    "name": account.get('accountName'),
                    "contribution": account.get('currentBalance', 0)
                })
    
    # Calculate total projected value from all SIPs
    total_projected_from_sips = sum(sp.get('projectedValue', 0) for sp in sip_projections)
    total_monthly_sip_contribution = sum(sp.get('monthlyContribution', 0) for sp in sip_projections)
    
    # If manual override is set but auto-calculate is also on, use max of both
    if goal.get('manualOverride'):
        manual_amount = goal.get('currentAmount', 0)
        if manual_amount > calculated_amount:
            return {
                "currentAmount": manual_amount,
                "linkedDetails": linked_details,
                "sipProjections": sip_projections,
                "totalProjectedFromSIPs": total_projected_from_sips,
                "totalMonthlySIPContribution": total_monthly_sip_contribution,
                "monthsToTarget": round(months_to_target, 1),
                "calculationMethod": "manual_override"
            }
    
    return {
        "currentAmount": calculated_amount,
        "linkedDetails": linked_details,
        "sipProjections": sip_projections,
        "totalProjectedFromSIPs": total_projected_from_sips,
        "totalMonthlySIPContribution": total_monthly_sip_contribution,
        "monthsToTarget": round(months_to_target, 1),
        "calculationMethod": "auto"
    }

@api_router.post("/goals", response_model=Goal)
async def create_goal(input: GoalCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    goal_dict = input.model_dump()
    goal_dict['userId'] = user.get('user_id')
    goal_obj = Goal(**goal_dict)
    
    # For Debt Elimination, auto-set target to outstanding amount
    if goal_obj.goalType == "Debt Elimination":
        if goal_obj.linkedLoanId:
            loan = await db.loans.find_one({"id": goal_obj.linkedLoanId}, {"_id": 0})
            if loan and goal_obj.targetAmount == 0:
                goal_obj.targetAmount = loan.get('outstandingAmount', 0)
        elif goal_obj.linkedCreditCardId:
            card = await db.credit_cards.find_one({"id": goal_obj.linkedCreditCardId}, {"_id": 0})
            if card and goal_obj.targetAmount == 0:
                goal_obj.targetAmount = card.get('outstandingAmount', 0)
    
    doc = goal_obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    
    await db.goals.insert_one(doc)
    return goal_obj

@api_router.get("/goals")
async def get_goals(request: Request):
    """Get all goals with calculated progress"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    goals = await db.goals.find(user_filter, {"_id": 0}).to_list(1000)
    
    result = []
    for goal in goals:
        if isinstance(goal.get('createdAt'), str):
            goal['createdAt'] = datetime.fromisoformat(goal['createdAt'])
        
        # Calculate progress
        progress_data = await calculate_goal_progress(goal)
        goal['calculatedAmount'] = progress_data['currentAmount']
        goal['linkedDetails'] = progress_data['linkedDetails']
        goal['calculationMethod'] = progress_data['calculationMethod']
        goal['sipProjections'] = progress_data.get('sipProjections', [])
        goal['totalProjectedFromSIPs'] = progress_data.get('totalProjectedFromSIPs', 0)
        goal['totalMonthlySIPContribution'] = progress_data.get('totalMonthlySIPContribution', 0)
        
        # Calculate progress percentage
        target = goal.get('targetAmount', 0)
        current = progress_data['currentAmount']
        goal['progressPercent'] = round((current / target) * 100, 1) if target > 0 else 0
        
        # Calculate projected progress percentage (using SIP projections)
        projected_total = progress_data.get('totalProjectedFromSIPs', 0)
        if projected_total > 0:
            goal['projectedProgressPercent'] = round((projected_total / target) * 100, 1) if target > 0 else 0
        else:
            goal['projectedProgressPercent'] = goal['progressPercent']
        
        # Calculate days remaining
        target_date = goal.get('targetDate')
        if target_date:
            try:
                target_dt = datetime.fromisoformat(target_date).date()
                today = datetime.now(timezone.utc).date()
                days_remaining = (target_dt - today).days
                goal['daysRemaining'] = days_remaining
                goal['isOverdue'] = days_remaining < 0
            except (ValueError, TypeError):
                goal['daysRemaining'] = None
                goal['isOverdue'] = False
        
        result.append(goal)
    
    # Sort by priority (1=high first) then by days remaining
    result.sort(key=lambda x: (x.get('priority', 1), x.get('daysRemaining') or 9999))
    
    return result

@api_router.get("/goals/allocation-status")
async def get_allocation_status(request: Request):
    """Get allocation status of all investments and accounts across goals"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    
    # Get all goals with their allocations
    goals = await db.goals.find(user_filter, {"_id": 0}).to_list(1000)
    
    # Get all investments and accounts
    investments = await db.investments.find(user_filter, {"_id": 0}).to_list(1000)
    accounts = await db.accounts.find(user_filter, {"_id": 0}).to_list(1000)
    
    # Build allocation map for investments
    investment_allocations = {}
    for inv in investments:
        inv_id = inv.get('id')
        total_value = inv.get('currentValue', inv.get('principal', 0))
        investment_allocations[inv_id] = {
            "id": inv_id,
            "name": inv.get('name', 'Unknown'),
            "category": inv.get('investmentCategory', ''),
            "totalValue": total_value,
            "allocatedAmount": 0,
            "remainingAmount": total_value,
            "allocations": []
        }
    
    # Build allocation map for accounts
    account_allocations = {}
    for acc in accounts:
        acc_id = acc.get('id')
        total_balance = acc.get('currentBalance', 0)
        account_allocations[acc_id] = {
            "id": acc_id,
            "name": acc.get('accountName', 'Unknown'),
            "accountType": acc.get('accountType', ''),
            "totalBalance": total_balance,
            "allocatedAmount": 0,
            "remainingAmount": total_balance,
            "allocations": []
        }
    
    # Process goal allocations - check both new and legacy formats
    for goal in goals:
        goal_id = goal.get('id')
        goal_name = goal.get('goalName', 'Unknown Goal')
        
        # Process new format: linkedInvestments with allocatedAmount
        for linked_inv in goal.get('linkedInvestments', []):
            inv_id = linked_inv.get('id')
            allocated = linked_inv.get('allocatedAmount', 0)
            if inv_id in investment_allocations and allocated > 0:
                investment_allocations[inv_id]['allocatedAmount'] += allocated
                investment_allocations[inv_id]['remainingAmount'] -= allocated
                investment_allocations[inv_id]['allocations'].append({
                    "goalId": goal_id,
                    "goalName": goal_name,
                    "allocatedAmount": allocated
                })
        
        # Process legacy format: linkedInvestmentIds (full amount)
        for inv_id in goal.get('linkedInvestmentIds', []):
            # Skip if already processed via linkedInvestments
            already_processed = any(a.get('goalId') == goal_id for a in investment_allocations.get(inv_id, {}).get('allocations', []))
            if inv_id in investment_allocations and not already_processed:
                total_value = investment_allocations[inv_id]['totalValue']
                investment_allocations[inv_id]['allocatedAmount'] += total_value
                investment_allocations[inv_id]['remainingAmount'] = 0
                investment_allocations[inv_id]['allocations'].append({
                    "goalId": goal_id,
                    "goalName": goal_name,
                    "allocatedAmount": total_value,
                    "isLegacy": True
                })
        
        # Process new format: linkedAccounts with allocatedAmount
        for linked_acc in goal.get('linkedAccounts', []):
            acc_id = linked_acc.get('id')
            allocated = linked_acc.get('allocatedAmount', 0)
            if acc_id in account_allocations and allocated > 0:
                account_allocations[acc_id]['allocatedAmount'] += allocated
                account_allocations[acc_id]['remainingAmount'] -= allocated
                account_allocations[acc_id]['allocations'].append({
                    "goalId": goal_id,
                    "goalName": goal_name,
                    "allocatedAmount": allocated
                })
        
        # Process legacy format: linkedAccountIds (full amount)
        for acc_id in goal.get('linkedAccountIds', []):
            already_processed = any(a.get('goalId') == goal_id for a in account_allocations.get(acc_id, {}).get('allocations', []))
            if acc_id in account_allocations and not already_processed:
                total_balance = account_allocations[acc_id]['totalBalance']
                account_allocations[acc_id]['allocatedAmount'] += total_balance
                account_allocations[acc_id]['remainingAmount'] = 0
                account_allocations[acc_id]['allocations'].append({
                    "goalId": goal_id,
                    "goalName": goal_name,
                    "allocatedAmount": total_balance,
                    "isLegacy": True
                })
    
    return {
        "investments": list(investment_allocations.values()),
        "accounts": list(account_allocations.values())
    }

@api_router.get("/goals/achievements")
async def get_goal_achievements(request: Request):
    """Get all completed goals with their milestone history for the achievements page"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["isCompleted"] = True
    completed_goals = await db.goals.find(user_filter, {"_id": 0}).to_list(1000)
    
    achievements = []
    for goal in completed_goals:
        if isinstance(goal.get('createdAt'), str):
            goal['createdAt'] = datetime.fromisoformat(goal['createdAt'])
        
        # Calculate final progress data
        progress_data = await calculate_goal_progress(goal)
        
        # Build milestone history with dates
        reached_milestones = goal.get('reachedMilestones', [])
        milestone_history = []
        
        # For each standard milestone, check if it was reached
        for milestone in [25, 50, 75, 100]:
            is_reached = milestone in reached_milestones
            milestone_history.append({
                "milestone": milestone,
                "reached": is_reached,
                "label": f"{milestone}% Complete"
            })
        
        # Calculate duration from creation to completion
        created_at = goal.get('createdAt')
        completed_date = goal.get('completedDate')
        duration_days = None
        
        if created_at and completed_date:
            try:
                if isinstance(created_at, str):
                    created_dt = datetime.fromisoformat(created_at)
                else:
                    created_dt = created_at
                completed_dt = datetime.fromisoformat(completed_date)
                duration_days = (completed_dt - created_dt).days
            except (ValueError, TypeError):
                pass
        
        achievement = {
            "id": goal.get('id'),
            "goalName": goal.get('goalName'),
            "goalType": goal.get('goalType'),
            "customTypeName": goal.get('customTypeName'),
            "targetAmount": goal.get('targetAmount', 0),
            "finalAmount": progress_data['currentAmount'],
            "targetDate": goal.get('targetDate'),
            "completedDate": goal.get('completedDate'),
            "createdAt": goal.get('createdAt').isoformat() if isinstance(goal.get('createdAt'), datetime) else goal.get('createdAt'),
            "milestoneHistory": milestone_history,
            "reachedMilestones": reached_milestones,
            "durationDays": duration_days,
            "priority": goal.get('priority', 1),
            "notes": goal.get('notes'),
            "linkedDetails": progress_data.get('linkedDetails', [])
        }
        
        achievements.append(achievement)
    
    # Sort by completion date (most recent first)
    achievements.sort(key=lambda x: x.get('completedDate') or '', reverse=True)
    
    # Calculate summary stats
    total_achieved = sum(a.get('finalAmount', 0) for a in achievements)
    avg_duration = sum(a.get('durationDays', 0) or 0 for a in achievements) / len(achievements) if achievements else 0
    
    return {
        "totalCompleted": len(achievements),
        "totalAmountAchieved": total_achieved,
        "averageDurationDays": round(avg_duration),
        "achievements": achievements
    }

@api_router.get("/goals/{goal_id}")
async def get_goal(goal_id: str, request: Request):
    """Get a single goal with full details"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = goal_id
    goal = await db.goals.find_one(user_filter, {"_id": 0})
    
    if not goal:
        raise HTTPException(status_code=404, detail="Goal not found")
    
    if isinstance(goal.get('createdAt'), str):
        goal['createdAt'] = datetime.fromisoformat(goal['createdAt'])
    
    # Calculate progress
    progress_data = await calculate_goal_progress(goal)
    goal['calculatedAmount'] = progress_data['currentAmount']
    goal['linkedDetails'] = progress_data['linkedDetails']
    goal['calculationMethod'] = progress_data['calculationMethod']
    goal['sipProjections'] = progress_data.get('sipProjections', [])
    goal['totalProjectedFromSIPs'] = progress_data.get('totalProjectedFromSIPs', 0)
    goal['totalMonthlySIPContribution'] = progress_data.get('totalMonthlySIPContribution', 0)
    goal['monthsToTarget'] = progress_data.get('monthsToTarget', 0)
    
    # Calculate progress percentage
    target = goal.get('targetAmount', 0)
    current = progress_data['currentAmount']
    goal['progressPercent'] = round((current / target) * 100, 1) if target > 0 else 0
    
    # Calculate projected progress percentage
    projected_total = progress_data.get('totalProjectedFromSIPs', 0)
    if projected_total > 0:
        goal['projectedProgressPercent'] = round((projected_total / target) * 100, 1) if target > 0 else 0
    else:
        goal['projectedProgressPercent'] = goal['progressPercent']
    
    # Calculate additional monthly savings needed to reach goal
    remaining = target - current
    months_to_target = progress_data.get('monthsToTarget', 0)
    monthly_sip = progress_data.get('totalMonthlySIPContribution', 0)
    
    if months_to_target > 0:
        # Monthly needed without SIP growth
        monthly_needed_total = remaining / months_to_target
        # Additional monthly needed beyond current SIP
        additional_monthly_needed = max(0, monthly_needed_total - monthly_sip)
        goal['additionalMonthlySavingsNeeded'] = round(additional_monthly_needed, 2)
        goal['totalMonthlyNeeded'] = round(monthly_needed_total, 2)
    else:
        goal['additionalMonthlySavingsNeeded'] = 0
        goal['totalMonthlyNeeded'] = 0
    
    # Calculate days remaining
    target_date = goal.get('targetDate')
    if target_date:
        try:
            target_dt = datetime.fromisoformat(target_date).date()
            today = datetime.now(timezone.utc).date()
            days_remaining = (target_dt - today).days
            goal['daysRemaining'] = days_remaining
            goal['isOverdue'] = days_remaining < 0
        except (ValueError, TypeError):
            goal['daysRemaining'] = None
            goal['isOverdue'] = False
    
    # Get full details of linked sources
    linked_investments = []
    for inv_id in goal.get('linkedInvestmentIds', []):
        inv = await db.investments.find_one({"id": inv_id}, {"_id": 0})
        if inv:
            linked_investments.append(inv)
    goal['linkedInvestments'] = linked_investments
    
    linked_accounts = []
    for acc_id in goal.get('linkedAccountIds', []):
        acc = await db.accounts.find_one({"id": acc_id}, {"_id": 0})
        if acc:
            linked_accounts.append(acc)
    goal['linkedAccounts'] = linked_accounts
    
    if goal.get('linkedLoanId'):
        loan = await db.loans.find_one({"id": goal['linkedLoanId']}, {"_id": 0})
        goal['linkedLoan'] = loan
    
    if goal.get('linkedCreditCardId'):
        card = await db.credit_cards.find_one({"id": goal['linkedCreditCardId']}, {"_id": 0})
        goal['linkedCreditCard'] = card
    
    return goal

@api_router.get("/goals/{goal_id}/milestones")
async def check_goal_milestones(goal_id: str):
    """Check and update milestones for a goal, return newly reached milestones"""
    goal = await db.goals.find_one({"id": goal_id}, {"_id": 0})
    
    if not goal:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Goal not found")
    
    # Calculate current progress
    progress_data = await calculate_goal_progress(goal)
    current_amount = progress_data['currentAmount']
    target_amount = goal.get('targetAmount', 0)
    
    progress_percent = round((current_amount / target_amount) * 100, 1) if target_amount > 0 else 0
    
    # Define milestones
    milestones = [25, 50, 75, 100]
    reached_milestones = goal.get('reachedMilestones', [])
    newly_reached = []
    
    for milestone in milestones:
        if progress_percent >= milestone and milestone not in reached_milestones:
            newly_reached.append(milestone)
            reached_milestones.append(milestone)
    
    # Update milestones in database if any new ones reached
    if newly_reached:
        await db.goals.update_one(
            {"id": goal_id},
            {"$set": {"reachedMilestones": reached_milestones}}
        )
    
    return {
        "goalId": goal_id,
        "goalName": goal.get('goalName', ''),
        "progressPercent": progress_percent,
        "currentAmount": current_amount,
        "targetAmount": target_amount,
        "reachedMilestones": reached_milestones,
        "newlyReached": newly_reached,
        "isCompleted": goal.get('isCompleted', False)
    }

@api_router.put("/goals/{goal_id}")
async def update_goal(goal_id: str, input: GoalCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = goal_id
    existing = await db.goals.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Goal not found")
    
    goal_dict = input.model_dump()
    goal_dict['id'] = goal_id
    goal_dict['userId'] = user.get('user_id')
    goal_dict['createdAt'] = existing['createdAt']
    goal_dict['isCompleted'] = existing.get('isCompleted', False)
    goal_dict['completedDate'] = existing.get('completedDate')
    
    await db.goals.replace_one({"id": goal_id}, goal_dict)
    
    # Return updated goal with progress
    updated = await db.goals.find_one({"id": goal_id}, {"_id": 0})
    progress_data = await calculate_goal_progress(updated)
    updated['calculatedAmount'] = progress_data['currentAmount']
    updated['progressPercent'] = round((progress_data['currentAmount'] / updated.get('targetAmount', 1)) * 100, 1) if updated.get('targetAmount', 0) > 0 else 0
    
    return updated

@api_router.patch("/goals/{goal_id}/complete")
async def mark_goal_complete(goal_id: str, request: Request):
    """Mark a goal as completed"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = goal_id
    existing = await db.goals.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Goal not found")
    
    await db.goals.update_one(
        {"id": goal_id},
        {"$set": {
            "isCompleted": True,
            "completedDate": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Goal marked as completed", "id": goal_id}

@api_router.patch("/goals/{goal_id}/progress")
async def update_goal_progress(goal_id: str, current_amount: float, request: Request):
    """Manually update the current amount for a goal"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = goal_id
    existing = await db.goals.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Goal not found")
    
    await db.goals.update_one(
        {"id": goal_id},
        {"$set": {
            "currentAmount": current_amount,
            "manualOverride": True
        }}
    )
    
    # Check if goal is now complete
    target = existing.get('targetAmount', 0)
    if current_amount >= target and target > 0:
        await db.goals.update_one(
            {"id": goal_id},
            {"$set": {
                "isCompleted": True,
                "completedDate": datetime.now(timezone.utc).isoformat()
            }}
        )
        return {"message": "Goal progress updated and marked as completed!", "id": goal_id, "currentAmount": current_amount}
    
    return {"message": "Goal progress updated", "id": goal_id, "currentAmount": current_amount}

@api_router.delete("/goals/{goal_id}")
async def delete_goal(goal_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = goal_id
    existing = await db.goals.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Goal not found")
    
    await db.goals.delete_one({"id": goal_id})
    return {"message": "Goal deleted successfully", "id": goal_id}

# Pydantic model for batch priority update
class GoalPriorityUpdate(BaseModel):
    id: str
    priority: int

@api_router.patch("/goals/reorder")
async def reorder_goals(updates: List[GoalPriorityUpdate]):
    """Update priorities for multiple goals at once (for drag-and-drop reordering)"""
    updated_count = 0
    
    for update in updates:
        result = await db.goals.update_one(
            {"id": update.id},
            {"$set": {"priority": update.priority}}
        )
        if result.modified_count > 0:
            updated_count += 1
    
    return {"message": f"Updated priorities for {updated_count} goals", "updatedCount": updated_count}

@api_router.get("/goals/summary/dashboard")
async def get_goals_dashboard_summary(request: Request):
    """Get summary of goals for dashboard widget"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["isCompleted"] = False
    goals = await db.goals.find(user_filter, {"_id": 0}).to_list(1000)
    
    total_goals = len(goals)
    completed_filter = get_user_filter(user)
    completed_filter["isCompleted"] = True
    completed_count = await db.goals.count_documents(completed_filter)
    
    summary = {
        "totalActiveGoals": total_goals,
        "completedGoals": completed_count,
        "goals": []
    }
    
    for goal in goals[:5]:  # Top 5 goals for dashboard
        progress_data = await calculate_goal_progress(goal)
        target = goal.get('targetAmount', 0)
        current = progress_data['currentAmount']
        
        target_date = goal.get('targetDate')
        days_remaining = None
        if target_date:
            try:
                target_dt = datetime.fromisoformat(target_date).date()
                today = datetime.now(timezone.utc).date()
                days_remaining = (target_dt - today).days
            except (ValueError, TypeError):
                pass
        
        summary["goals"].append({
            "id": goal.get('id'),
            "goalName": goal.get('goalName'),
            "goalType": goal.get('goalType'),
            "targetAmount": target,
            "currentAmount": current,
            "progressPercent": round((current / target) * 100, 1) if target > 0 else 0,
            "daysRemaining": days_remaining,
            "priority": goal.get('priority', 1)
        })
    
    # Sort by priority then by progress (lowest progress first - needs attention)
    summary["goals"].sort(key=lambda x: (x.get('priority', 1), x.get('progressPercent', 0)))
    
    return summary

# ============ OTHER INCOME ENDPOINTS ============

@api_router.post("/other-income", response_model=OtherIncome)
async def create_other_income(input: OtherIncomeCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    income_dict = input.model_dump()
    income_dict['userId'] = user.get('user_id')
    income_obj = OtherIncome(**income_dict)
    
    doc = income_obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    
    await db.other_income.insert_one(doc)
    return income_obj

@api_router.get("/other-income", response_model=List[OtherIncome])
async def get_other_incomes(request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    incomes = await db.other_income.find(user_filter, {"_id": 0}).to_list(1000)
    
    for income in incomes:
        if isinstance(income.get('createdAt'), str):
            income['createdAt'] = datetime.fromisoformat(income['createdAt'])
    
    return incomes

@api_router.get("/other-income/{income_id}", response_model=OtherIncome)
async def get_other_income(income_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = income_id
    income = await db.other_income.find_one(user_filter, {"_id": 0})
    
    if not income:
        raise HTTPException(status_code=404, detail="Other income not found")
    
    if isinstance(income.get('createdAt'), str):
        income['createdAt'] = datetime.fromisoformat(income['createdAt'])
    
    return income

@api_router.put("/other-income/{income_id}", response_model=OtherIncome)
async def update_other_income(income_id: str, input: OtherIncomeCreate, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = income_id
    existing = await db.other_income.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Other income not found")
    
    income_dict = input.model_dump()
    income_dict['id'] = income_id
    income_dict['userId'] = user.get('user_id')
    income_dict['createdAt'] = existing['createdAt']
    
    await db.other_income.replace_one({"id": income_id}, income_dict)
    
    income_obj = OtherIncome(**income_dict)
    if isinstance(income_obj.createdAt, str):
        income_obj.createdAt = datetime.fromisoformat(income_obj.createdAt)
    
    return income_obj

@api_router.delete("/other-income/{income_id}")
async def delete_other_income(income_id: str, request: Request):
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    user_filter["id"] = income_id
    existing = await db.other_income.find_one(user_filter, {"_id": 0})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Other income not found")
    
    await db.other_income.delete_one({"id": income_id})
    return {"message": "Other income deleted successfully", "id": income_id}

# ============ AI INSIGHTS ============
from pydantic import BaseModel as PydanticBaseModel
from typing import List as ListType

class InsightItem(PydanticBaseModel):
    type: str
    icon: str
    title: str
    description: str
    priority: str
    actionable: bool
    action_text: Optional[str] = None
    action_link: Optional[str] = None

async def generate_ai_insights_internal(financial_data: dict) -> list:
    """Generate AI insights using OpenAI GPT-5.2"""
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    
    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        return get_fallback_insights_internal(financial_data)
    
    summary = f"""
    Financial Summary:
    - Total Net Worth: ₹{financial_data.get('net_worth', 0):,.0f}
    - Monthly Income: ₹{financial_data.get('monthly_income', 0):,.0f}
    - Monthly Expenses: ₹{financial_data.get('monthly_expenses', 0):,.0f}
    - Monthly Balance: ₹{financial_data.get('monthly_savings', 0):,.0f}
    - Total Assets: ₹{financial_data.get('total_assets', 0):,.0f}
    - Total Investments: ₹{financial_data.get('total_investments', 0):,.0f}
    - Total Liabilities: ₹{financial_data.get('total_liabilities', 0):,.0f}
    - Liquid Balance (Bank Accounts): ₹{financial_data.get('liquid_balance', 0):,.0f}
    - FD/RD Balance: ₹{financial_data.get('fd_rd_balance', 0):,.0f}
    - Liquid Investments (marked): ₹{financial_data.get('liquid_investments', 0):,.0f}
    - {financial_data.get('emergency_fund_goal_info', 'No Emergency Fund goal set')}
    - Total Emergency Fund Available: ₹{financial_data.get('emergency_fund', 0):,.0f}
    - Active Goals: {financial_data.get('active_goals', 0)}
    - Savings Rate: {financial_data.get('savings_rate', 0):.1f}%
    - Top Expense Categories: {financial_data.get('top_expenses', 'N/A')}
    
    Insurance Coverage:
    - {financial_data.get('insurance_summary', 'No insurance data')}
    - Insurance Gaps: {financial_data.get('insurance_gaps', 'Unknown')}
    """
    
    system_prompt = """You are a smart financial advisor AI. Analyze the user's financial data and provide 4-5 personalized, actionable insights.
    
    IMPORTANT: Always include at least one insurance-related insight if there are coverage gaps or if health insurance is missing.
    
    Return ONLY a valid JSON array with objects containing:
    - type: "spending", "savings", "goal", "alert", "trend", or "insurance"
    - icon: single emoji
    - title: max 6 words
    - description: max 25 words, practical advice
    - priority: "high", "medium", or "low"
    - actionable: boolean
    - action_text: button text if actionable
    - action_link: "/my-expenses", "/my-income", "/my-goals", "/my-investments", "/my-loans", "/my-insurance", or "/portfolio"
    
    No markdown, no explanation - ONLY the JSON array. 
    IMPORTANT: Use Indian number format - ₹ symbol with Lakhs (L) and Crores (Cr). Example: ₹70L, ₹2.3Cr. Never use M (millions)."""
    
    try:
        chat = LlmChat(
            api_key=api_key,
            session_id=f"insights-{datetime.now(timezone.utc).timestamp()}",
            system_message=system_prompt
        ).with_model("openai", "gpt-5.2")
        
        user_message = UserMessage(text=f"Analyze this financial data:\n{summary}")
        response = await chat.send_message(user_message)
        
        import json
        clean_response = response.strip()
        if clean_response.startswith("```"):
            clean_response = clean_response.split("```")[1]
            if clean_response.startswith("json"):
                clean_response = clean_response[4:]
        clean_response = clean_response.strip()
        
        return json.loads(clean_response)
    except Exception as e:
        logger.error(f"AI Insights Error: {str(e)}")
        return get_fallback_insights_internal(financial_data)

def get_fallback_insights_internal(data: dict) -> list:
    """Generate basic insights without AI"""
    insights = []
    savings_rate = data.get('savings_rate', 0)
    
    if savings_rate > 30:
        insights.append({
            "type": "trend", "icon": "🎉", "title": "Excellent Savings!",
            "description": f"You're saving {savings_rate:.0f}% of income. Great financial discipline!",
            "priority": "low", "actionable": False
        })
    elif savings_rate < 10 and savings_rate >= 0:
        insights.append({
            "type": "alert", "icon": "⚠️", "title": "Low Savings Alert",
            "description": f"Only {savings_rate:.0f}% savings rate. Review your expenses to save more.",
            "priority": "high", "actionable": True, "action_text": "View Expenses", "action_link": "/my-expenses"
        })
    
    if data.get('total_liabilities', 0) > data.get('emergency_fund', data.get('liquid_balance', 0)) * 2:
        insights.append({
            "type": "alert", "icon": "💳", "title": "High Debt Ratio",
            "description": "Liabilities exceed 2x your liquid funds. Focus on debt reduction.",
            "priority": "high", "actionable": True, "action_text": "View Loans", "action_link": "/my-loans"
        })
    
    if data.get('active_goals', 0) > 0:
        insights.append({
            "type": "goal", "icon": "🎯", "title": "Goals In Progress",
            "description": f"You have {data.get('active_goals')} active goals. Keep contributing!",
            "priority": "medium", "actionable": True, "action_text": "View Goals", "action_link": "/my-goals"
        })
    else:
        insights.append({
            "type": "savings", "icon": "💡", "title": "Set Financial Goals",
            "description": "Create goals for better financial planning and motivation.",
            "priority": "medium", "actionable": True, "action_text": "Add Goal", "action_link": "/my-goals"
        })
    
    # Insurance insights
    if not data.get('has_health_insurance', True):
        insights.append({
            "type": "insurance", "icon": "🏥", "title": "Get Health Insurance",
            "description": "No health insurance found. Medical emergencies can drain savings quickly.",
            "priority": "high", "actionable": True, "action_text": "Add Insurance", "action_link": "/my-insurance"
        })
    elif data.get('life_coverage', 0) < data.get('monthly_income', 0) * 120:
        insights.append({
            "type": "insurance", "icon": "🛡️", "title": "Increase Life Cover",
            "description": "Life cover should be 10x annual income. Consider term insurance.",
            "priority": "medium", "actionable": True, "action_text": "View Insurance", "action_link": "/my-insurance"
        })
    
    return insights

@api_router.get("/ai/insights")
async def get_ai_insights(request: Request):
    """Get AI-powered financial insights"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_filter = get_user_filter(user)
    current_month = datetime.now(timezone.utc).month
    current_year = datetime.now(timezone.utc).year
    
    try:
        # Calculate monthly income using the same logic as dashboard
        incomes = await db.income_sources.find(user_filter, {"_id": 0}).to_list(1000)
        monthly_income = 0
        for income in incomes:
            amount = income.get('expectedAmount', 0)
            freq = income.get('frequency', 'Monthly')
            if freq == 'Daily':
                monthly_income += amount * 30
            elif freq == 'Weekly':
                monthly_income += amount * 4
            elif freq == 'Monthly':
                monthly_income += amount
            elif freq == 'Quarterly':
                if current_month in [1, 4, 7, 10]:
                    monthly_income += amount
            elif freq == 'Half-Yearly':
                if current_month in [1, 7]:
                    monthly_income += amount
            elif freq == 'Yearly':
                selected_month = income.get('selectedMonth', '')
                month_mapping = {"January": 1, "February": 2, "March": 3, "April": 4, 
                    "May": 5, "June": 6, "July": 7, "August": 8, 
                    "September": 9, "October": 10, "November": 11, "December": 12}
                if month_mapping.get(selected_month) == current_month:
                    monthly_income += amount
            else:
                monthly_income += amount
        
        # Add other income
        other_incomes = await db.other_income.find(user_filter, {"_id": 0}).to_list(1000)
        for other_inc in other_incomes:
            amount = other_inc.get('amount', 0)
            freq = other_inc.get('frequency', 'One-Time')
            if freq == 'Monthly':
                monthly_income += amount
            elif freq == 'One-Time':
                date_received = other_inc.get('dateReceived', '')
                if date_received:
                    try:
                        date_obj = datetime.fromisoformat(date_received).date()
                        if date_obj.month == current_month and date_obj.year == current_year:
                            monthly_income += amount
                    except (ValueError, TypeError):
                        pass
        
        # Calculate monthly expenses using the same logic as dashboard
        expenses = await db.expenses.find(user_filter, {"_id": 0}).to_list(1000)
        monthly_expenses = 0
        expense_by_category = {}
        for expense in expenses:
            amount = expense.get('expectedAmount', 0)
            freq = expense.get('frequency', 'Monthly')
            cat = expense.get('category', 'Other')
            
            expense_amount = 0
            if freq == 'Daily':
                expense_amount = amount * 30
            elif freq == 'Weekly':
                expense_amount = amount * 4
            elif freq == 'Monthly':
                expense_amount = amount
            elif freq == 'Quarterly':
                if current_month in [1, 4, 7, 10]:
                    expense_amount = amount
            elif freq == 'Half-Yearly':
                if current_month in [1, 7]:
                    expense_amount = amount
            elif freq == 'Yearly':
                selected_month = expense.get('selectedMonth', '')
                month_mapping = {"January": 1, "February": 2, "March": 3, "April": 4, 
                    "May": 5, "June": 6, "July": 7, "August": 8, 
                    "September": 9, "October": 10, "November": 11, "December": 12}
                if month_mapping.get(selected_month) == current_month:
                    expense_amount = amount
            else:
                expense_amount = amount
            
            monthly_expenses += expense_amount
            expense_by_category[cat] = expense_by_category.get(cat, 0) + expense_amount
        
        top_expenses = sorted(expense_by_category.items(), key=lambda x: x[1], reverse=True)[:3]
        top_expenses_str = ", ".join([f"{cat}: ₹{amt:,.0f}" for cat, amt in top_expenses]) or "No expenses"
        
        # Get assets, investments, loans, accounts, goals
        assets = await db.assets.find(user_filter, {"_id": 0}).to_list(1000)
        total_assets = sum(a.get('currentValue', 0) for a in assets)
        
        investments = await db.investments.find(user_filter, {"_id": 0}).to_list(1000)
        total_investments = sum(inv.get('currentValue', inv.get('principal', 0)) for inv in investments)
        
        loans = await db.loans.find(user_filter, {"_id": 0}).to_list(1000)
        total_liabilities = sum(loan.get('outstandingAmount', 0) for loan in loans)
        
        credit_cards = await db.credit_cards.find(user_filter, {"_id": 0}).to_list(1000)
        total_liabilities += sum(cc.get('currentOutstanding', 0) for cc in credit_cards)
        
        accounts = await db.accounts.find(user_filter, {"_id": 0}).to_list(1000)
        liquid_balance = sum(a.get('currentBalance', 0) for a in accounts)
        
        # Get insurance data
        insurances = await db.insurances.find(user_filter, {"_id": 0}).to_list(1000)
        
        # Calculate insurance coverage by type
        life_coverage = sum(i.get('coverageAmount', 0) for i in insurances if i.get('insuranceType') in ['Life Insurance', 'Term Insurance'])
        health_coverage = sum(i.get('coverageAmount', 0) for i in insurances if i.get('insuranceType') == 'Health Insurance')
        vehicle_coverage = sum(i.get('coverageAmount', 0) for i in insurances if i.get('insuranceType') == 'Vehicle Insurance')
        
        # Calculate annual premium
        def get_annual_premium(ins):
            premium = ins.get('premiumAmount', 0)
            freq = ins.get('premiumFrequency', 'Yearly')
            multipliers = {'Monthly': 12, 'Quarterly': 4, 'Half-Yearly': 2, 'Yearly': 1, 'One-Time': 0}
            return premium * multipliers.get(freq, 1)
        
        total_annual_premium = sum(get_annual_premium(i) for i in insurances)
        
        # Insurance insights
        insurance_types = list(set(i.get('insuranceType') for i in insurances))
        has_health_insurance = 'Health Insurance' in insurance_types
        
        # Include FDs and RDs as liquid/emergency funds (they are accessible in emergencies)
        fd_rd_balance = sum(
            inv.get('currentValue', inv.get('principal', 0)) 
            for inv in investments 
            if inv.get('investmentCategory') in ['Fixed Deposit (FD)', 'Recurring Deposit (RD)']
        )
        
        # Include investments explicitly marked as liquid (exclude FD/RD to avoid double counting)
        liquid_investments = sum(
            inv.get('currentValue', inv.get('principal', 0))
            for inv in investments
            if inv.get('isLiquidAsset', False) and inv.get('investmentCategory') not in ['Fixed Deposit (FD)', 'Recurring Deposit (RD)']
        )
        
        # Get Emergency Fund goals and calculate their actual funded amounts from linked sources
        goals = await db.goals.find(user_filter, {"_id": 0}).to_list(1000)
        emergency_fund_goals_amount = 0
        emergency_fund_goal_info = []
        
        for g in goals:
            if g.get('goalType') == 'Emergency Fund' and not g.get('isCompleted', False):
                # Calculate actual amount from linked investments and accounts
                goal_amount = 0
                
                # Add linked investments
                for inv_id in g.get('linkedInvestmentIds', []):
                    inv = await db.investments.find_one({"id": inv_id}, {"_id": 0})
                    if inv:
                        goal_amount += inv.get('currentValue', inv.get('principal', 0))
                
                # Add linked accounts
                for acc_id in g.get('linkedAccountIds', []):
                    acc = await db.accounts.find_one({"id": acc_id}, {"_id": 0})
                    if acc and acc.get('accountType') != 'Credit Card':
                        goal_amount += acc.get('currentBalance', 0)
                
                # If no linked sources, use stored currentAmount
                if goal_amount == 0:
                    goal_amount = g.get('currentAmount', 0)
                
                emergency_fund_goals_amount += goal_amount
                target = g.get('targetAmount', 0)
                progress = round((goal_amount / target * 100), 1) if target > 0 else 0
                emergency_fund_goal_info.append({
                    "name": g.get('goalName'),
                    "current": goal_amount,
                    "target": target,
                    "progress": progress
                })
        
        active_goals = len([g for g in goals if not g.get('isCompleted', False)])
        
        # Total emergency fund = liquid cash + FD/RD + liquid investments (avoid double counting)
        # Don't add emergency_fund_goals_amount again since those sources are already counted
        emergency_fund = liquid_balance + fd_rd_balance + liquid_investments
        
        monthly_savings = monthly_income - monthly_expenses
        savings_rate = (monthly_savings / monthly_income * 100) if monthly_income > 0 else 0
        net_worth = total_assets + total_investments + liquid_balance - total_liabilities
        
        # Format emergency fund goal info for AI
        ef_goal_str = ""
        if emergency_fund_goal_info:
            ef = emergency_fund_goal_info[0]
            ef_goal_str = f"Emergency Fund Goal: {ef['progress']}% funded (₹{ef['current']:,.0f} of ₹{ef['target']:,.0f})"
        
        # Build insurance summary
        insurance_summary = f"Life/Term Coverage: ₹{life_coverage:,.0f}, Health Coverage: ₹{health_coverage:,.0f}, Vehicle Coverage: ₹{vehicle_coverage:,.0f}, Annual Premium: ₹{total_annual_premium:,.0f}"
        insurance_gaps = []
        if not has_health_insurance:
            insurance_gaps.append("No Health Insurance")
        if life_coverage < monthly_income * 120:  # Rule: Life cover should be 10x annual income
            insurance_gaps.append(f"Life cover low (have ₹{life_coverage:,.0f}, need ₹{monthly_income * 120:,.0f})")
        
        financial_data = {
            "net_worth": net_worth, "monthly_income": monthly_income,
            "monthly_expenses": monthly_expenses, "monthly_savings": monthly_savings,
            "total_assets": total_assets, "total_investments": total_investments,
            "total_liabilities": total_liabilities, "liquid_balance": liquid_balance,
            "emergency_fund": emergency_fund, "fd_rd_balance": fd_rd_balance,
            "liquid_investments": liquid_investments, "emergency_fund_goals": emergency_fund_goals_amount,
            "emergency_fund_goal_info": ef_goal_str,
            "active_goals": active_goals, "savings_rate": savings_rate,
            "top_expenses": top_expenses_str,
            "insurance_summary": insurance_summary,
            "insurance_gaps": ", ".join(insurance_gaps) if insurance_gaps else "Adequate coverage",
            "has_health_insurance": has_health_insurance,
            "life_coverage": life_coverage,
            "total_annual_premium": total_annual_premium
        }
        
        insights = await generate_ai_insights_internal(financial_data)
        
        return {
            "insights": insights,
            "generated_at": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        logger.error(f"Error generating insights: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ NOTIFICATION HELPER ============
MAX_NOTIFICATIONS_PER_USER = 10

async def create_notification_and_cleanup(notification: dict):
    """
    Create a new notification and remove old ones if user has more than MAX_NOTIFICATIONS_PER_USER.
    This ensures only the last 10 notifications are kept per user.
    """
    user_id = notification.get("userId")
    
    # Insert the new notification
    await db.notifications.insert_one(notification)
    
    # Count total notifications for this user
    total_count = await db.notifications.count_documents({"userId": user_id})
    
    # If more than limit, delete the oldest ones
    if total_count > MAX_NOTIFICATIONS_PER_USER:
        # Find the oldest notifications to delete
        excess_count = total_count - MAX_NOTIFICATIONS_PER_USER
        oldest_notifications = await db.notifications.find(
            {"userId": user_id},
            {"_id": 1}
        ).sort("createdAt", 1).limit(excess_count).to_list(excess_count)
        
        # Delete the oldest notifications
        if oldest_notifications:
            ids_to_delete = [n["_id"] for n in oldest_notifications]
            await db.notifications.delete_many({"_id": {"$in": ids_to_delete}})


@api_router.post("/notifications/test-reminder/{income_id}")
async def send_test_reminder(income_id: str, request: Request):
    """Manually send a test reminder notification for an income source (for testing purposes)"""
    session_token = request.cookies.get("session_token")
    if not session_token:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    user_id = session.get("user_id")
    
    # Find the income source
    source = await db.income_sources.find_one({"id": income_id, "userId": user_id}, {"_id": 0})
    if not source:
        raise HTTPException(status_code=404, detail="Income source not found")
    
    source_name = source.get("name", "Income")
    source_type = source.get("type", "job").lower().replace(' ', '-')
    expected_amount = source.get("expectedAmount", 0)
    
    # Create notification
    action_url = f"/{source_type}-income/{income_id}"
    notification = {
        "id": str(uuid.uuid4()),
        "userId": user_id,
        "title": f"Time to record {source_name}",
        "message": f"Hi! It's time to record your {source_name} income. Expected: ₹{expected_amount:,.0f}" if expected_amount else f"Hi! It's time to record your {source_name} income.",
        "type": "income_reminder",
        "relatedIncomeId": income_id,
        "relatedIncomeName": source_name,
        "expectedAmount": expected_amount,
        "actionUrl": action_url,
        "isRead": False,
        "createdAt": datetime.now(timezone.utc).isoformat()
    }
    
    await create_notification_and_cleanup(notification)
    
    return {"success": True, "message": f"Test reminder sent for {source_name}"}

# ============ NOTIFICATION ENDPOINTS ============
@api_router.get("/notifications")
async def get_notifications(user_id: str = None, request: Request = None):
    """Get the last 10 notifications for a user (sorted by newest first)"""
    if not user_id:
        session_token = request.cookies.get("session_token")
        if session_token:
            session = await db.user_sessions.find_one({"session_token": session_token})
            if session:
                user_id = session.get("user_id")
    
    if not user_id:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    # Get only the last 10 notifications, sorted by newest first
    notifications = await db.notifications.find(
        {"userId": user_id},
        {"_id": 0}
    ).sort("createdAt", -1).limit(10).to_list(10)
    
    return notifications

@api_router.get("/notifications/unread-count")
async def get_unread_notification_count(user_id: str = None, request: Request = None):
    """Get count of unread notifications"""
    if not user_id:
        session_token = request.cookies.get("session_token")
        if session_token:
            session = await db.user_sessions.find_one({"session_token": session_token})
            if session:
                user_id = session.get("user_id")
    
    if not user_id:
        return {"count": 0}
    
    count = await db.notifications.count_documents({"userId": user_id, "isRead": False})
    return {"count": count}

@api_router.patch("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, user_id: str = None, request: Request = None):
    """Mark a notification as read"""
    if not user_id:
        session_token = request.cookies.get("session_token")
        if session_token:
            session = await db.user_sessions.find_one({"session_token": session_token})
            if session:
                user_id = session.get("user_id")
    
    if not user_id:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    result = await db.notifications.update_one(
        {"id": notification_id, "userId": user_id},
        {"$set": {"isRead": True}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    return {"success": True}

@api_router.patch("/notifications/mark-all-read")
async def mark_all_notifications_read(user_id: str = None, request: Request = None):
    """Mark all notifications as read for a user"""
    if not user_id:
        session_token = request.cookies.get("session_token")
        if session_token:
            session = await db.user_sessions.find_one({"session_token": session_token})
            if session:
                user_id = session.get("user_id")
    
    if not user_id:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    await db.notifications.update_many(
        {"userId": user_id, "isRead": False},
        {"$set": {"isRead": True}}
    )
    
    return {"success": True}

@api_router.delete("/notifications/{notification_id}")
async def delete_notification(notification_id: str, user_id: str = None, request: Request = None):
    """Delete a notification"""
    if not user_id:
        session_token = request.cookies.get("session_token")
        if session_token:
            session = await db.user_sessions.find_one({"session_token": session_token})
            if session:
                user_id = session.get("user_id")
    
    if not user_id:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    result = await db.notifications.delete_one({"id": notification_id, "userId": user_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    return {"success": True}

@api_router.delete("/notifications")
async def clear_all_notifications(request: Request):
    """Clear all notifications for the current user"""
    session_token = request.cookies.get("session_token")
    if not session_token:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    user_id = session.get("user_id")
    result = await db.notifications.delete_many({"userId": user_id})
    
    return {"success": True, "deleted_count": result.deleted_count}

@api_router.delete("/notifications/by-entity/{entity_id}")
async def delete_notifications_by_entity(entity_id: str, request: Request):
    """Delete all notifications related to a specific income/expense entity (e.g., when user records income)"""
    session_token = request.cookies.get("session_token")
    if not session_token:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    user_id = session.get("user_id")
    
    # Delete notifications that match either relatedIncomeId or relatedExpenseId
    result = await db.notifications.delete_many({
        "userId": user_id,
        "$or": [
            {"relatedIncomeId": entity_id},
            {"relatedExpenseId": entity_id}
        ]
    })
    
    return {"success": True, "deleted_count": result.deleted_count}

# ============ PUSH NOTIFICATION SUBSCRIPTION ============
from push_service import get_vapid_public_key, send_push_notification, send_income_reminder, send_auto_entry_notification

@api_router.get("/push/vapid-key")
async def get_vapid_key():
    """Get the VAPID public key for push subscription"""
    public_key = get_vapid_public_key()
    if not public_key:
        raise HTTPException(status_code=500, detail="VAPID keys not configured")
    return {"public_key": public_key}

@api_router.post("/push/subscribe")
async def subscribe_push_notifications(subscription: dict, user_id: str = None, request: Request = None):
    """Subscribe to push notifications"""
    if not user_id:
        session_token = request.cookies.get("session_token")
        if session_token:
            session = await db.user_sessions.find_one({"session_token": session_token})
            if session:
                user_id = session.get("user_id")
    
    if not user_id:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    # Upsert subscription
    await db.push_subscriptions.update_one(
        {"userId": user_id, "endpoint": subscription.get("endpoint")},
        {"$set": {
            "userId": user_id,
            "endpoint": subscription.get("endpoint"),
            "keys": subscription.get("keys", {}),
            "createdAt": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    return {"success": True}

@api_router.delete("/push/unsubscribe")
async def unsubscribe_push_notifications(endpoint: str, user_id: str = None, request: Request = None):
    """Unsubscribe from push notifications"""
    if not user_id:
        session_token = request.cookies.get("session_token")
        if session_token:
            session = await db.user_sessions.find_one({"session_token": session_token})
            if session:
                user_id = session.get("user_id")
    
    if not user_id:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    await db.push_subscriptions.delete_one({"userId": user_id, "endpoint": endpoint})
    
    return {"success": True}

# ============ CRON JOB ENDPOINT FOR VARIABLE INCOME FALLBACK ============
@api_router.post("/cron/process-variable-income")
async def process_variable_income_fallback(api_key: str = None):
    """
    Daily cron job to process variable income entries.
    Should be called once per day at midnight by MongoDB Atlas Triggers or external scheduler.
    
    Logic:
    1. Find all variable income sources that were due yesterday but have no manual entry
    2. Auto-create transaction entries using lastRecordedAmount (or expectedAmount as fallback)
    3. Create notifications for users about auto-entries
    4. Update nextDueDate for processed sources
    """
    # Simple API key validation for cron security
    expected_key = os.environ.get("CRON_API_KEY", "moneyssutra_cron_secret_2026")
    if api_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid API key")
    
    today = datetime.now(timezone.utc).date()
    yesterday = today - timedelta(days=1)
    yesterday_str = yesterday.isoformat()
    
    processed_count = 0
    notifications_created = 0
    
    try:
        # Find all variable income sources that were due yesterday
        variable_sources = await db.income_sources.find({
            "incomeType": "variable",
            "nextDueDate": yesterday_str
        }, {"_id": 0}).to_list(10000)
        
        for source in variable_sources:
            user_id = source.get("userId")
            source_id = source.get("id")
            income_name = source.get("name", "Unknown Income")
            
            # Check if user already recorded an entry for this source on the due date
            existing_entry = await db.income_transactions.find_one({
                "userId": user_id,
                "incomeSourceId": source_id,
                "recordedDate": yesterday_str
            })
            
            if not existing_entry:
                # No manual entry - auto-create using fallback amount
                fallback_amount = source.get("lastRecordedAmount") or source.get("expectedAmount", 0)
                
                # Create auto-entry transaction
                auto_entry = {
                    "id": str(uuid.uuid4()),
                    "userId": user_id,
                    "incomeSourceId": source_id,
                    "incomeName": income_name,
                    "incomeType": source.get("type"),
                    "amount": fallback_amount,
                    "recordedDate": yesterday_str,
                    "isAutoEntry": True,
                    "notes": "Auto-recorded (24hr fallback)",
                    "createdAt": datetime.now(timezone.utc).isoformat()
                }
                await db.income_transactions.insert_one(auto_entry)
                processed_count += 1
                
                action_url = f"/{source.get('type', 'job').lower().replace(' ', '-')}-income/{source_id}"
                
                # Create in-app notification for user
                notification = {
                    "id": str(uuid.uuid4()),
                    "userId": user_id,
                    "title": "Auto-recorded Income",
                    "message": f"₹{fallback_amount:,.0f} was auto-recorded for {income_name} as you didn't log it within 24 hours.",
                    "type": "auto_entry",
                    "relatedIncomeId": source_id,
                    "relatedIncomeName": income_name,
                    "actionUrl": action_url,
                    "isRead": False,
                    "createdAt": datetime.now(timezone.utc).isoformat()
                }
                await create_notification_and_cleanup(notification)
                notifications_created += 1
                
                # Send browser push notifications
                subscriptions = await db.push_subscriptions.find(
                    {"userId": user_id}, {"_id": 0}
                ).to_list(100)
                
                for sub in subscriptions:
                    subscription_info = {
                        "endpoint": sub.get("endpoint"),
                        "keys": sub.get("keys", {})
                    }
                    result = await send_auto_entry_notification(
                        subscription_info, income_name, fallback_amount, source_id
                    )
                    if result.get("should_remove"):
                        await db.push_subscriptions.delete_one({"endpoint": sub.get("endpoint")})
            
            # Calculate next due date based on frequency
            next_due = calculate_next_due_date(source, yesterday)
            if next_due:
                await db.income_sources.update_one(
                    {"id": source_id},
                    {"$set": {"nextDueDate": next_due.isoformat()}}
                )
        
        return {
            "success": True,
            "processed_entries": processed_count,
            "notifications_created": notifications_created,
            "processed_date": yesterday_str
        }
    
    except Exception as e:
        logger.error(f"Error in cron job: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

def calculate_next_due_date(source: dict, current_due: datetime.date) -> Optional[datetime.date]:
    """Calculate the next due date based on frequency"""
    frequency = source.get("frequency", "Monthly")
    
    if frequency == "Daily":
        return current_due + timedelta(days=1)
    elif frequency == "Weekly":
        return current_due + timedelta(weeks=1)
    elif frequency == "Monthly":
        # Add one month
        next_month = current_due.month + 1
        next_year = current_due.year
        if next_month > 12:
            next_month = 1
            next_year += 1
        day = min(current_due.day, 28)  # Safe day for all months
        return datetime.date(next_year, next_month, day)
    elif frequency == "Quarterly":
        # Add 3 months
        next_month = current_due.month + 3
        next_year = current_due.year
        while next_month > 12:
            next_month -= 12
            next_year += 1
        day = min(current_due.day, 28)
        return datetime.date(next_year, next_month, day)
    elif frequency == "Half-Yearly":
        # Add 6 months
        next_month = current_due.month + 6
        next_year = current_due.year
        while next_month > 12:
            next_month -= 12
            next_year += 1
        day = min(current_due.day, 28)
        return datetime.date(next_year, next_month, day)
    elif frequency == "Yearly":
        return datetime.date(current_due.year + 1, current_due.month, current_due.day)
    
    return None

@api_router.post("/cron/send-reminder-notifications")
async def send_reminder_notifications(api_key: str = None):
    """
    Hourly cron job to send reminder notifications for variable income.
    Checks for income sources whose reminder time matches the current hour.
    Sends both in-app and browser push notifications.
    """
    expected_key = os.environ.get("CRON_API_KEY", "moneyssutra_cron_secret_2026")
    if api_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid API key")
    
    now = datetime.now(timezone.utc)
    current_hour = now.strftime("%H:00")
    today_str = now.date().isoformat()
    
    notifications_sent = 0
    push_sent = 0
    push_failed = 0
    
    try:
        # Find variable income sources due today with matching reminder time
        sources = await db.income_sources.find({
            "incomeType": "variable",
            "nextDueDate": today_str,
            "reminderTime": {"$regex": f"^{current_hour[:2]}"}  # Match hour
        }, {"_id": 0}).to_list(10000)
        
        for source in sources:
            user_id = source.get("userId")
            source_id = source.get("id")
            income_name = source.get("name", "Unknown Income")
            income_expected_amount = source.get("expectedAmount", 0)
            
            # Get user name
            user = await db.users.find_one({"user_id": user_id}, {"_id": 0, "name": 1})
            user_name = user.get("name", "there") if user else "there"
            
            action_url = f"/{source.get('type', 'job').lower().replace(' ', '-')}-income/{source_id}"
            
            # Create in-app notification
            notification = {
                "id": str(uuid.uuid4()),
                "userId": user_id,
                "title": f"Time to record {income_name}",
                "message": f"Hi {user_name}, it's time to record your {income_name}. Tap to enter today's actual amount.",
                "type": "income_reminder",
                "relatedIncomeId": source_id,
                "relatedIncomeName": income_name,
                "expectedAmount": income_expected_amount,
                "actionUrl": action_url,
                "isRead": False,
                "createdAt": datetime.now(timezone.utc).isoformat()
            }
            await create_notification_and_cleanup(notification)
            notifications_sent += 1
            
            # Send browser push notifications to all user's subscriptions
            subscriptions = await db.push_subscriptions.find(
                {"userId": user_id}, {"_id": 0}
            ).to_list(100)
            
            for sub in subscriptions:
                subscription_info = {
                    "endpoint": sub.get("endpoint"),
                    "keys": sub.get("keys", {})
                }
                result = await send_income_reminder(subscription_info, income_name, source_id)
                
                if result.get("success"):
                    push_sent += 1
                else:
                    push_failed += 1
                    # Remove expired subscriptions
                    if result.get("should_remove"):
                        await db.push_subscriptions.delete_one({"endpoint": sub.get("endpoint")})
        
        return {
            "success": True,
            "in_app_notifications": notifications_sent,
            "push_sent": push_sent,
            "push_failed": push_failed,
            "checked_hour": current_hour
        }
    
    except Exception as e:
        logger.error(f"Error sending reminders: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ============ INCOME TRANSACTION ENDPOINTS ============

@api_router.post("/income-transactions")
async def record_income_transaction(transaction: dict, user_id: str = None, request: Request = None):
    """
    Record a new income transaction (append, don't overwrite).
    Creates an immutable record linked to the income source template.
    """
    if not user_id:
        session_token = request.cookies.get("session_token")
        if session_token:
            session = await db.user_sessions.find_one({"session_token": session_token})
            if session:
                user_id = session.get("user_id")
    
    if not user_id:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    entity_id = transaction.get("entityId") or transaction.get("incomeSourceId")
    amount = float(transaction.get("amount", 0))
    transaction_date = transaction.get("transactionDate") or transaction.get("recordedDate") or datetime.now(timezone.utc).date().isoformat()
    
    # Get the income source template (including legacy data without userId)
    income_source = await db.income_sources.find_one({"id": entity_id, "userId": user_id}, {"_id": 0})
    if not income_source:
        # Also check for legacy data (userId is null/missing)
        income_source = await db.income_sources.find_one({
            "id": entity_id, 
            "$or": [{"userId": None}, {"userId": {"$exists": False}}]
        }, {"_id": 0})
    if not income_source:
        raise HTTPException(status_code=404, detail="Income source not found")
    
    # Create immutable transaction record
    entry = {
        "id": str(uuid.uuid4()),
        "userId": user_id,
        "entityId": entity_id,
        "entityType": income_source.get("type", "Unknown"),
        "entityName": income_source.get("name", "Unknown"),
        "amount": amount,
        "transactionDate": transaction_date,
        "notes": transaction.get("notes", ""),
        "source": transaction.get("source", "manual"),
        "isLocked": False,
        "createdAt": datetime.now(timezone.utc).isoformat()
    }
    await db.income_transactions.insert_one(entry)
    
    # Update the income source template with last recorded info (for reference only)
    await db.income_sources.update_one(
        {"id": entity_id, "userId": user_id},
        {"$set": {
            "lastRecordedAmount": amount,
            "lastEntryDate": transaction_date
        }}
    )
    
    return {"success": True, "transaction": {k: v for k, v in entry.items() if k != "_id"}}

@api_router.get("/income-transactions")
async def get_income_transactions(
    entity_id: str = None,
    income_source_id: str = None,  # Alias for backwards compatibility
    start_date: str = None,
    end_date: str = None,
    user_id: str = None,
    request: Request = None
):
    """Get income transactions with optional filters, sorted by date descending"""
    if not user_id:
        session_token = request.cookies.get("session_token")
        if session_token:
            session = await db.user_sessions.find_one({"session_token": session_token})
            if session:
                user_id = session.get("user_id")
    
    if not user_id:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    query = {"userId": user_id}
    
    # Support both entityId and incomeSourceId for backwards compatibility
    source_id = entity_id or income_source_id
    if source_id:
        query["$or"] = [{"entityId": source_id}, {"incomeSourceId": source_id}]
    
    if start_date and end_date:
        query["transactionDate"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["transactionDate"] = {"$gte": start_date}
    elif end_date:
        query["transactionDate"] = {"$lte": end_date}
    
    transactions = await db.income_transactions.find(query, {"_id": 0}).sort("transactionDate", -1).to_list(1000)
    
    return transactions

@api_router.get("/income-transactions/history/{entity_id}")
async def get_income_history(entity_id: str, request: Request):
    """
    Get complete transaction history for an income source, grouped by date.
    Used for the History Page view.
    """
    session_token = request.cookies.get("session_token")
    if not session_token:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    user_id = session.get("user_id")
    
    # Get income source info (including legacy data without userId)
    income_source = await db.income_sources.find_one({"id": entity_id, "userId": user_id}, {"_id": 0})
    if not income_source:
        # Also check for legacy data (userId is null/missing)
        income_source = await db.income_sources.find_one({
            "id": entity_id, 
            "$or": [{"userId": None}, {"userId": {"$exists": False}}]
        }, {"_id": 0})
    if not income_source:
        raise HTTPException(status_code=404, detail="Income source not found")
    
    # Get all transactions for this source
    transactions = await db.income_transactions.find(
        {"$or": [{"entityId": entity_id}, {"incomeSourceId": entity_id}], "userId": user_id},
        {"_id": 0}
    ).sort("transactionDate", -1).to_list(1000)
    
    # Calculate totals
    total_amount = sum(t.get("amount", 0) for t in transactions)
    transaction_count = len(transactions)
    
    return {
        "incomeSource": income_source,
        "transactions": transactions,
        "summary": {
            "totalAmount": total_amount,
            "transactionCount": transaction_count,
            "averageAmount": total_amount / transaction_count if transaction_count > 0 else 0
        }
    }

@api_router.get("/income-transactions/monthly-summary")
async def get_monthly_income_summary(month: str = None, request: Request = None):
    """
    Get aggregated income totals for a month.
    Used for Monthly Cash Flow calculation on Home Page.
    Returns sum of all transactions for the specified month.
    """
    session_token = request.cookies.get("session_token")
    if not session_token:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    user_id = session.get("user_id")
    
    # Default to current month if not specified
    if not month:
        month = datetime.now().strftime("%Y-%m")
    
    # Get start and end dates for the month
    year, mon = month.split("-")
    start_date = f"{month}-01"
    # Get last day of month
    if int(mon) == 12:
        end_date = f"{int(year)+1}-01-01"
    else:
        end_date = f"{year}-{int(mon)+1:02d}-01"
    
    # Aggregate transactions for the month
    pipeline = [
        {
            "$match": {
                "userId": user_id,
                "transactionDate": {"$gte": start_date, "$lt": end_date}
            }
        },
        {
            "$group": {
                "_id": "$entityType",
                "totalAmount": {"$sum": "$amount"},
                "count": {"$sum": 1}
            }
        }
    ]
    
    results = await db.income_transactions.aggregate(pipeline).to_list(100)
    
    # Calculate grand total
    grand_total = sum(r.get("totalAmount", 0) for r in results)
    
    # Also get fixed income that might not have transactions
    fixed_income = await db.income_sources.find(
        {"userId": user_id, "incomeType": {"$ne": "variable"}},
        {"_id": 0, "type": 1, "expectedAmount": 1, "frequency": 1}
    ).to_list(100)
    
    # Calculate expected fixed income for the month
    fixed_total = 0
    for income in fixed_income:
        freq = income.get("frequency", "Monthly")
        amount = income.get("expectedAmount", 0)
        if freq == "Daily":
            fixed_total += amount * 30
        elif freq == "Weekly":
            fixed_total += amount * 4
        elif freq == "Bi-Weekly":
            fixed_total += amount * 2
        elif freq == "Monthly":
            fixed_total += amount
        elif freq == "Quarterly":
            fixed_total += amount / 3
        elif freq == "Half-Yearly":
            fixed_total += amount / 6
        elif freq == "Yearly":
            fixed_total += amount / 12
    
    return {
        "month": month,
        "variableIncomeTotal": grand_total,
        "fixedIncomeTotal": fixed_total,
        "grandTotal": grand_total + fixed_total,
        "byType": results
    }

class IncomeTransactionUpdate(BaseModel):
    amount: float
    transactionDate: str
    notes: Optional[str] = None

@api_router.put("/income-transactions/{transaction_id}")
async def update_income_transaction(transaction_id: str, update: IncomeTransactionUpdate, request: Request):
    """
    Update an income transaction (only if not locked).
    Transactions become locked 24 hours after creation.
    """
    session_token = request.cookies.get("session_token")
    if not session_token:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    user_id = session.get("user_id")
    
    # Find the transaction
    transaction = await db.income_transactions.find_one(
        {"id": transaction_id, "userId": user_id},
        {"_id": 0}
    )
    
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Check if locked
    if transaction.get("isLocked"):
        raise HTTPException(
            status_code=403, 
            detail="This transaction is locked and cannot be updated. Create an adjustment entry instead."
        )
    
    # Check if older than 24 hours
    created_at = transaction.get("createdAt")
    if created_at:
        created_time = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
        if datetime.now(timezone.utc) - created_time > timedelta(hours=24):
            # Lock the transaction
            await db.income_transactions.update_one(
                {"id": transaction_id},
                {"$set": {"isLocked": True}}
            )
            raise HTTPException(
                status_code=403,
                detail="This transaction is now locked (older than 24 hours). Create an adjustment entry instead."
            )
    
    # Update the transaction
    update_data = {
        "amount": update.amount,
        "transactionDate": update.transactionDate,
        "notes": update.notes or "",
        "updatedAt": datetime.now(timezone.utc).isoformat()
    }
    
    await db.income_transactions.update_one(
        {"id": transaction_id, "userId": user_id},
        {"$set": update_data}
    )
    
    # Fetch updated transaction
    updated = await db.income_transactions.find_one(
        {"id": transaction_id},
        {"_id": 0}
    )
    
    return updated

@api_router.delete("/income-transactions/{transaction_id}")
async def delete_income_transaction(transaction_id: str, request: Request):
    """
    Delete an income transaction (only if not locked).
    Transactions become locked 24 hours after creation.
    """
    session_token = request.cookies.get("session_token")
    if not session_token:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    user_id = session.get("user_id")
    
    # Find the transaction
    transaction = await db.income_transactions.find_one(
        {"id": transaction_id, "userId": user_id},
        {"_id": 0}
    )
    
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Check if locked
    if transaction.get("isLocked"):
        raise HTTPException(
            status_code=403, 
            detail="This transaction is locked and cannot be deleted. Create an adjustment entry instead."
        )
    
    # Check if older than 24 hours
    created_at = transaction.get("createdAt")
    if created_at:
        created_time = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
        if datetime.now(timezone.utc) - created_time > timedelta(hours=24):
            # Lock the transaction
            await db.income_transactions.update_one(
                {"id": transaction_id},
                {"$set": {"isLocked": True}}
            )
            raise HTTPException(
                status_code=403,
                detail="This transaction is now locked (older than 24 hours). Create an adjustment entry instead."
            )
    
    # Delete the transaction
    await db.income_transactions.delete_one({"id": transaction_id, "userId": user_id})
    
    return {"success": True, "message": "Transaction deleted"}

@api_router.post("/income-transactions/{transaction_id}/adjust")
async def adjust_income_transaction(transaction_id: str, adjustment: dict, request: Request):
    """
    Create an adjustment entry for a locked transaction.
    This preserves the original record while allowing corrections.
    """
    session_token = request.cookies.get("session_token")
    if not session_token:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    user_id = session.get("user_id")
    
    # Find the original transaction
    original = await db.income_transactions.find_one(
        {"id": transaction_id, "userId": user_id},
        {"_id": 0}
    )
    
    if not original:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Create adjustment entry
    adjustment_amount = float(adjustment.get("amount", 0))
    adjustment_entry = {
        "id": str(uuid.uuid4()),
        "userId": user_id,
        "entityId": original.get("entityId"),
        "entityType": original.get("entityType"),
        "entityName": original.get("entityName"),
        "amount": adjustment_amount,
        "transactionDate": original.get("transactionDate"),
        "notes": f"Adjustment for transaction {transaction_id}: {adjustment.get('reason', 'Correction')}",
        "source": "adjustment",
        "originalTransactionId": transaction_id,
        "isLocked": False,
        "createdAt": datetime.now(timezone.utc).isoformat()
    }
    
    await db.income_transactions.insert_one(adjustment_entry)
    
    return {"success": True, "adjustment": {k: v for k, v in adjustment_entry.items() if k != "_id"}}


# ============ EXPENSE TRANSACTION ENDPOINTS ============

@api_router.post("/expense-transactions")
async def record_expense_transaction(transaction: dict, request: Request):
    """
    Record a new expense transaction (append, don't overwrite).
    Creates an immutable record linked to the expense template.
    """
    session_token = request.cookies.get("session_token")
    if not session_token:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    user_id = session.get("user_id")
    
    entity_id = transaction.get("entityId") or transaction.get("expenseId")
    amount = float(transaction.get("amount", 0))
    transaction_date = transaction.get("transactionDate") or datetime.now(timezone.utc).date().isoformat()
    
    # Get the expense template (including legacy data without userId)
    expense_template = await db.expenses.find_one({"id": entity_id, "userId": user_id}, {"_id": 0})
    if not expense_template:
        # Also check for legacy data (userId is null/missing)
        expense_template = await db.expenses.find_one({
            "id": entity_id, 
            "$or": [{"userId": None}, {"userId": {"$exists": False}}]
        }, {"_id": 0})
    if not expense_template:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    # Create immutable transaction record
    entry = {
        "id": str(uuid.uuid4()),
        "userId": user_id,
        "entityId": entity_id,
        "entityName": expense_template.get("expenseName", "Unknown"),
        "category": expense_template.get("category", "Other"),
        "amount": amount,
        "transactionDate": transaction_date,
        "notes": transaction.get("notes", ""),
        "source": transaction.get("source", "manual"),
        "isLocked": False,
        "createdAt": datetime.now(timezone.utc).isoformat()
    }
    await db.expense_transactions.insert_one(entry)
    
    # Update the expense template with last recorded info (for reference only)
    await db.expenses.update_one(
        {"id": entity_id, "userId": user_id},
        {"$set": {
            "lastPaidDate": transaction_date,
            "isPaid": True
        }}
    )
    
    return {"success": True, "transaction": {k: v for k, v in entry.items() if k != "_id"}}

@api_router.get("/expense-transactions")
async def get_expense_transactions(
    entity_id: str = None,
    expense_id: str = None,  # Alias for backwards compatibility
    start_date: str = None,
    end_date: str = None,
    category: str = None,
    request: Request = None
):
    """Get expense transactions with optional filters, sorted by date descending"""
    session_token = request.cookies.get("session_token")
    if not session_token:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    user_id = session.get("user_id")
    
    query = {"userId": user_id}
    
    # Support both entityId and expenseId for backwards compatibility
    source_id = entity_id or expense_id
    if source_id:
        query["$or"] = [{"entityId": source_id}, {"expenseId": source_id}]
    
    if category:
        query["category"] = category
    
    if start_date and end_date:
        query["transactionDate"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["transactionDate"] = {"$gte": start_date}
    elif end_date:
        query["transactionDate"] = {"$lte": end_date}
    
    transactions = await db.expense_transactions.find(query, {"_id": 0}).sort("transactionDate", -1).to_list(1000)
    
    return transactions

@api_router.get("/expense-transactions/history/{entity_id}")
async def get_expense_history(entity_id: str, request: Request):
    """
    Get complete transaction history for an expense, grouped by date.
    Used for the History Page view.
    """
    session_token = request.cookies.get("session_token")
    if not session_token:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    user_id = session.get("user_id")
    
    # Get expense template info (including legacy data without userId)
    expense_template = await db.expenses.find_one({"id": entity_id, "userId": user_id}, {"_id": 0})
    if not expense_template:
        # Also check for legacy data (userId is null/missing)
        expense_template = await db.expenses.find_one({
            "id": entity_id, 
            "$or": [{"userId": None}, {"userId": {"$exists": False}}]
        }, {"_id": 0})
    if not expense_template:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    # Get all transactions for this expense
    transactions = await db.expense_transactions.find(
        {"$or": [{"entityId": entity_id}, {"expenseId": entity_id}], "userId": user_id},
        {"_id": 0}
    ).sort("transactionDate", -1).to_list(1000)
    
    # Calculate totals
    total_amount = sum(t.get("amount", 0) for t in transactions)
    transaction_count = len(transactions)
    
    return {
        "expense": expense_template,
        "transactions": transactions,
        "summary": {
            "totalAmount": total_amount,
            "transactionCount": transaction_count,
            "averageAmount": total_amount / transaction_count if transaction_count > 0 else 0
        }
    }

@api_router.get("/expense-transactions/monthly-summary")
async def get_monthly_expense_summary(month: str = None, request: Request = None):
    """
    Get aggregated expense totals for a month.
    Used for Monthly Cash Flow calculation.
    Returns sum of all expense transactions for the specified month.
    """
    session_token = request.cookies.get("session_token")
    if not session_token:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    user_id = session.get("user_id")
    
    # Default to current month if not specified
    if not month:
        month = datetime.now().strftime("%Y-%m")
    
    # Get start and end dates for the month
    year, mon = month.split("-")
    start_date = f"{month}-01"
    # Get last day of month
    if int(mon) == 12:
        end_date = f"{int(year)+1}-01-01"
    else:
        end_date = f"{year}-{int(mon)+1:02d}-01"
    
    # Aggregate transactions for the month
    pipeline = [
        {
            "$match": {
                "userId": user_id,
                "transactionDate": {"$gte": start_date, "$lt": end_date}
            }
        },
        {
            "$group": {
                "_id": "$category",
                "totalAmount": {"$sum": "$amount"},
                "count": {"$sum": 1}
            }
        }
    ]
    
    results = await db.expense_transactions.aggregate(pipeline).to_list(100)
    
    # Calculate grand total
    grand_total = sum(r.get("totalAmount", 0) for r in results)
    
    return {
        "month": month,
        "actualExpenseTotal": grand_total,
        "byCategory": results
    }

@api_router.delete("/expense-transactions/{transaction_id}")
async def delete_expense_transaction(transaction_id: str, request: Request):
    """
    Delete an expense transaction (only if not locked).
    Transactions become locked 24 hours after creation.
    """
    session_token = request.cookies.get("session_token")
    if not session_token:
        raise HTTPException(status_code=401, detail="User not authenticated")
    
    session = await db.user_sessions.find_one({"session_token": session_token})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    user_id = session.get("user_id")
    
    # Find the transaction
    transaction = await db.expense_transactions.find_one(
        {"id": transaction_id, "userId": user_id},
        {"_id": 0}
    )
    
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Check if locked
    if transaction.get("isLocked"):
        raise HTTPException(
            status_code=403, 
            detail="This transaction is locked and cannot be deleted. Create an adjustment entry instead."
        )
    
    # Check if older than 24 hours
    created_at = transaction.get("createdAt")
    if created_at:
        created_time = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
        if datetime.now(timezone.utc) - created_time > timedelta(hours=24):
            # Lock the transaction
            await db.expense_transactions.update_one(
                {"id": transaction_id},
                {"$set": {"isLocked": True}}
            )
            raise HTTPException(
                status_code=403,
                detail="This transaction is now locked (older than 24 hours). Create an adjustment entry instead."
            )
    
    # Delete the transaction
    await db.expense_transactions.delete_one({"id": transaction_id, "userId": user_id})
    
    return {"success": True, "message": "Transaction deleted"}


app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Background scheduler state
scheduler_running = False

async def check_and_process_due_premiums():
    """
    Check for insurance premiums due today and auto-record them as expense transactions.
    This runs once per day as part of the scheduler.
    """
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        logger.info(f"Checking for due insurance premiums for date: {today}")
        
        # Find all insurances with autoCreateExpense enabled
        insurances = await db.insurances.find({
            "autoCreateExpense": True,
            "premiumPaymentDate": {"$exists": True, "$ne": None}
        }, {"_id": 0}).to_list(500)
        
        for insurance in insurances:
            user_id = insurance.get("userId")
            insurance_id = insurance.get("id")
            policy_name = insurance.get("policyName", "Insurance Premium")
            premium_amount = insurance.get("premiumAmount", 0)
            frequency = insurance.get("premiumFrequency", "Yearly")
            payment_date_str = insurance.get("premiumPaymentDate")
            end_date_str = insurance.get("endDate")
            premium_end_date_str = insurance.get("premiumEndDate")
            
            if not payment_date_str or not premium_amount:
                continue
            
            # Check if premium end date has passed
            if premium_end_date_str:
                premium_end = datetime.strptime(premium_end_date_str, "%Y-%m-%d")
                if datetime.now() > premium_end:
                    continue
            
            # Check if policy end date has passed
            if end_date_str:
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
                if datetime.now() > end_date:
                    continue
            
            # Calculate if today is a premium due date based on frequency
            base_date = datetime.strptime(payment_date_str, "%Y-%m-%d")
            current_date = datetime.now()
            is_due_today = False
            
            if frequency == "One-Time":
                is_due_today = payment_date_str == today
            elif frequency == "Monthly":
                # Check if today's day matches payment day
                is_due_today = base_date.day == current_date.day
            elif frequency == "Quarterly":
                # Check if this is a quarterly payment month and day matches
                months_diff = (current_date.year - base_date.year) * 12 + (current_date.month - base_date.month)
                is_due_today = (months_diff % 3 == 0) and (base_date.day == current_date.day)
            elif frequency == "Half-Yearly":
                months_diff = (current_date.year - base_date.year) * 12 + (current_date.month - base_date.month)
                is_due_today = (months_diff % 6 == 0) and (base_date.day == current_date.day)
            elif frequency == "Yearly":
                is_due_today = (base_date.month == current_date.month) and (base_date.day == current_date.day)
            
            if not is_due_today:
                continue
            
            # Check if we already recorded this premium today
            existing_transaction = await db.expense_transactions.find_one({
                "entityId": insurance_id,
                "transactionDate": today,
                "source": "auto_premium"
            })
            
            if existing_transaction:
                logger.debug(f"Premium already recorded today for {policy_name}")
                continue
            
            # Create expense transaction for the premium
            transaction = {
                "id": str(uuid.uuid4()),
                "userId": user_id,
                "entityId": insurance_id,
                "entityName": policy_name,
                "category": "Insurance",
                "amount": premium_amount,
                "transactionDate": today,
                "notes": f"Auto-recorded {frequency} premium for {policy_name}",
                "source": "auto_premium",
                "isLocked": False,
                "createdAt": datetime.now(timezone.utc).isoformat()
            }
            
            await db.expense_transactions.insert_one(transaction)
            logger.info(f"Auto-recorded premium transaction for {policy_name}: ₹{premium_amount}")
            
            # Create notification for the user
            if user_id:
                notification = {
                    "id": str(uuid.uuid4()),
                    "userId": user_id,
                    "title": f"Premium Recorded: {policy_name}",
                    "message": f"Your {frequency} premium of ₹{premium_amount:,.0f} for {policy_name} has been auto-recorded as an expense.",
                    "type": "premium_recorded",
                    "relatedInsuranceId": insurance_id,
                    "actionUrl": f"/insurance/{insurance_id}",
                    "isRead": False,
                    "createdAt": datetime.now(timezone.utc).isoformat()
                }
                await create_notification_and_cleanup(notification)
                
    except Exception as e:
        logger.error(f"Error processing due premiums: {str(e)}")


async def auto_record_fixed_income():
    """
    Auto-record income transactions for Fixed income sources based on their frequency.
    This runs once per day as part of the scheduler.
    """
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        logger.info(f"Checking for fixed income due today: {today}")
        
        # Find all fixed income sources
        fixed_incomes = await db.income_sources.find({
            "incomeType": "fixed"
        }, {"_id": 0}).to_list(500)
        
        for income in fixed_incomes:
            user_id = income.get("userId")
            income_id = income.get("id")
            income_name = income.get("name", "Income")
            expected_amount = income.get("expectedAmount", 0)
            frequency = income.get("frequency", "Monthly")
            selected_date = income.get("selectedDate")
            selected_day = income.get("selectedDay")
            selected_month = income.get("selectedMonth")
            custom_date = income.get("customDate")
            
            if not expected_amount:
                continue
            
            # Calculate if today is a payment due date based on frequency
            current_date = datetime.now()
            is_due_today = False
            
            if frequency == "Daily":
                is_due_today = True
            elif frequency == "Weekly":
                if selected_day:
                    days_of_week = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
                    is_due_today = days_of_week[current_date.weekday()] == selected_day or current_date.strftime("%A") == selected_day
            elif frequency == "Monthly":
                if selected_date:
                    try:
                        base_date = datetime.strptime(selected_date, "%Y-%m-%d") if "-" in selected_date else None
                        if base_date:
                            is_due_today = base_date.day == current_date.day
                        else:
                            is_due_today = int(selected_date) == current_date.day
                    except (ValueError, TypeError):
                        pass
            elif frequency == "Quarterly":
                if selected_date and selected_month:
                    try:
                        months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
                        base_month = months.index(selected_month) if selected_month in months else 0
                        base_date = datetime.strptime(selected_date, "%Y-%m-%d") if "-" in selected_date else None
                        base_day = base_date.day if base_date else int(selected_date)
                        
                        months_diff = (current_date.month - 1 - base_month) % 12
                        is_due_today = (months_diff % 3 == 0) and (base_day == current_date.day)
                    except (ValueError, TypeError, IndexError):
                        pass
            elif frequency == "Half-Yearly":
                if selected_date and selected_month:
                    try:
                        months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
                        base_month = months.index(selected_month) if selected_month in months else 0
                        base_date = datetime.strptime(selected_date, "%Y-%m-%d") if "-" in selected_date else None
                        base_day = base_date.day if base_date else int(selected_date)
                        
                        months_diff = (current_date.month - 1 - base_month) % 12
                        is_due_today = (months_diff % 6 == 0) and (base_day == current_date.day)
                    except (ValueError, TypeError, IndexError):
                        pass
            elif frequency == "Yearly":
                if selected_date and selected_month:
                    try:
                        months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
                        base_month = months.index(selected_month) if selected_month in months else 0
                        base_date = datetime.strptime(selected_date, "%Y-%m-%d") if "-" in selected_date else None
                        base_day = base_date.day if base_date else int(selected_date)
                        
                        is_due_today = (base_month == current_date.month - 1) and (base_day == current_date.day)
                    except (ValueError, TypeError, IndexError):
                        pass
            elif frequency == "Others":
                if custom_date:
                    is_due_today = custom_date == today
            
            if not is_due_today:
                continue
            
            # Check if we already recorded this income today
            existing_transaction = await db.income_transactions.find_one({
                "entityId": income_id,
                "transactionDate": today
            })
            
            if existing_transaction:
                logger.debug(f"Fixed income already recorded today for {income_name}")
                continue
            
            # Create income transaction
            transaction = {
                "id": str(uuid.uuid4()),
                "userId": user_id,
                "entityId": income_id,
                "entityName": income_name,
                "incomeAmount": expected_amount,
                "transactionDate": today,
                "type": "Fixed",
                "source": "auto_fixed",
                "notes": f"Auto-recorded {frequency} income",
                "isLocked": False,
                "createdAt": datetime.now(timezone.utc).isoformat()
            }
            
            await db.income_transactions.insert_one(transaction)
            logger.info(f"Auto-recorded fixed income for {income_name}: ₹{expected_amount}")
            
            # Create notification for the user
            if user_id:
                notification = {
                    "id": str(uuid.uuid4()),
                    "userId": user_id,
                    "title": f"Income Recorded: {income_name}",
                    "message": f"Your {frequency} income of ₹{expected_amount:,.0f} for {income_name} has been auto-recorded.",
                    "type": "auto_entry",
                    "relatedIncomeId": income_id,
                    "relatedIncomeName": income_name,
                    "expectedAmount": expected_amount,
                    "actionUrl": f"/income/{income.get('type', 'business').lower()}/{income_id}",
                    "isRead": False,
                    "createdAt": datetime.now(timezone.utc).isoformat()
                }
                
                await create_notification_and_cleanup(notification)
                logger.info(f"Created auto-entry notification for {income_name}")
                
    except Exception as e:
        logger.error(f"Error auto-recording fixed income: {str(e)}")


async def check_and_send_reminders():
    """
    Background task that runs every minute to check for income reminders.
    Sends notifications for variable income entries that have reminder times matching current time.
    Also checks for due insurance premiums once per day at midnight.
    """
    global scheduler_running
    scheduler_running = True
    logger.info("Background reminder scheduler started")
    
    last_premium_check_date = None
    
    while scheduler_running:
        try:
            now = datetime.now()
            current_time = now.strftime("%H:%M")
            today = now.strftime("%Y-%m-%d")
            
            # Check for due premiums once per day (at startup or when date changes)
            if last_premium_check_date != today:
                logger.info(f"Running daily checks for {today}")
                await check_and_process_due_premiums()
                await auto_record_fixed_income()
                last_premium_check_date = today
            
            logger.debug(f"Checking reminders for time: {current_time}")
            
            # Find variable income sources with reminders set for current time
            # that haven't been recorded today
            variable_incomes = await db.income_sources.find({
                "incomeType": "variable",
                "reminderTime": current_time,
                "lastEntryDate": {"$ne": today}
            }, {"_id": 0}).to_list(100)
            
            for source in variable_incomes:
                user_id = source.get("userId")
                source_id = source.get("id")
                source_name = source.get("name", "Income")
                source_type = source.get("type", "job").lower().replace(' ', '-')
                expected_amount = source.get("expectedAmount", 0)
                
                # Check if we already sent a reminder notification today for this source
                existing_notification = await db.notifications.find_one({
                    "userId": user_id,
                    "relatedIncomeId": source_id,
                    "type": "income_reminder",
                    "createdAt": {"$regex": f"^{today}"}
                })
                
                if existing_notification:
                    logger.debug(f"Reminder already sent today for {source_name}")
                    continue
                
                # Create notification
                action_url = f"/{source_type}-income/{source_id}"
                notification = {
                    "id": str(uuid.uuid4()),
                    "userId": user_id,
                    "title": f"Time to record {source_name}",
                    "message": f"Hi! It's time to record your {source_name} income. Expected: ₹{expected_amount:,.0f}" if expected_amount else f"Hi! It's time to record your {source_name} income.",
                    "type": "income_reminder",
                    "relatedIncomeId": source_id,
                    "relatedIncomeName": source_name,
                    "actionUrl": action_url,
                    "isRead": False,
                    "createdAt": datetime.now(timezone.utc).isoformat()
                }
                
                await create_notification_and_cleanup(notification)
                logger.info(f"Sent reminder notification for {source_name} to user {user_id}")
            
        except Exception as e:
            logger.error(f"Error in reminder scheduler: {str(e)}")
        
        # Wait 60 seconds before next check
        await asyncio.sleep(60)

# ============ ANALYTICS API ============
class AnalyticsSnapshot(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    userId: str
    month: int
    year: int
    netWorth: float = 0
    totalAssets: float = 0
    totalInvestments: float = 0
    totalLiabilities: float = 0
    liquidBalance: float = 0
    monthlyIncome: float = 0
    monthlyExpense: float = 0
    savingsRate: float = 0
    investmentGains: float = 0
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

@api_router.get("/analytics/snapshots")
async def get_analytics_snapshots(request: Request, months: int = 12):
    """Get historical monthly snapshots for analytics"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    
    # Get snapshots from database
    snapshots = await db.analytics_snapshots.find(
        {"userId": user_id},
        {"_id": 0}
    ).sort([("year", -1), ("month", -1)]).limit(months).to_list(months)
    
    return snapshots

@api_router.post("/analytics/snapshot")
async def create_analytics_snapshot(request: Request):
    """Create a monthly snapshot of current financial state"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    now = datetime.now(timezone.utc)
    current_month = now.month
    current_year = now.year
    
    # Check if snapshot already exists for this month
    existing = await db.analytics_snapshots.find_one({
        "userId": user_id,
        "month": current_month,
        "year": current_year
    })
    
    # Get current financial data
    user_filter = {"userId": user_id}
    
    assets, investments, accounts, loans, credit_cards, incomes, expenses = await asyncio.gather(
        db.assets.find(user_filter, {"_id": 0}).to_list(1000),
        db.investments.find(user_filter, {"_id": 0}).to_list(1000),
        db.accounts.find(user_filter, {"_id": 0}).to_list(1000),
        db.loans.find(user_filter, {"_id": 0}).to_list(1000),
        db.credit_cards.find(user_filter, {"_id": 0}).to_list(1000),
        db.income_sources.find(user_filter, {"_id": 0}).to_list(1000),
        db.expenses.find(user_filter, {"_id": 0}).to_list(1000)
    )
    
    total_assets = sum(a.get('currentValue', 0) for a in assets)
    total_investments = sum(i.get('currentValue', 0) for i in investments)
    investment_principal = sum(i.get('principal', 0) for i in investments)
    investment_gains = total_investments - investment_principal
    
    liquid_balance = sum(acc.get('currentBalance', 0) or acc.get('balance', 0) for acc in accounts)
    
    total_liabilities = sum(l.get('outstandingAmount', 0) for l in loans)
    total_liabilities += sum(c.get('currentOutstanding', 0) or c.get('outstandingAmount', 0) for c in credit_cards)
    
    net_worth = total_assets + total_investments + liquid_balance - total_liabilities
    
    # Calculate monthly income/expense
    monthly_income = sum(i.get('expectedAmount', 0) for i in incomes if i.get('frequency') == 'Monthly')
    monthly_income += sum(i.get('expectedAmount', 0) * 4 for i in incomes if i.get('frequency') == 'Weekly')
    monthly_income += sum(i.get('expectedAmount', 0) * 30 for i in incomes if i.get('frequency') == 'Daily')
    
    monthly_expense = sum(e.get('expectedAmount', 0) for e in expenses if e.get('frequency') == 'Monthly')
    monthly_expense += sum(e.get('expectedAmount', 0) * 4 for e in expenses if e.get('frequency') == 'Weekly')
    
    savings_rate = ((monthly_income - monthly_expense) / monthly_income * 100) if monthly_income > 0 else 0
    
    snapshot_data = {
        "id": str(uuid.uuid4()),
        "userId": user_id,
        "month": current_month,
        "year": current_year,
        "netWorth": net_worth,
        "totalAssets": total_assets,
        "totalInvestments": total_investments,
        "totalLiabilities": total_liabilities,
        "liquidBalance": liquid_balance,
        "monthlyIncome": monthly_income,
        "monthlyExpense": monthly_expense,
        "savingsRate": round(savings_rate, 1),
        "investmentGains": investment_gains,
        "createdAt": now
    }
    
    if existing:
        # Update existing snapshot
        await db.analytics_snapshots.update_one(
            {"userId": user_id, "month": current_month, "year": current_year},
            {"$set": snapshot_data}
        )
    else:
        # Create new snapshot
        await db.analytics_snapshots.insert_one(snapshot_data)
    
    return {"message": "Snapshot created", "snapshot": snapshot_data}

@api_router.get("/analytics/investment-performance")
async def get_investment_performance(request: Request):
    """Get investment performance breakdown"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    user_filter = {"userId": user_id}
    
    investments = await db.investments.find(user_filter, {"_id": 0}).to_list(1000)
    
    total_invested = sum(i.get('principal', 0) for i in investments)
    current_value = sum(i.get('currentValue', 0) for i in investments)
    total_gains = current_value - total_invested
    gain_percent = (total_gains / total_invested * 100) if total_invested > 0 else 0
    
    # Group by category
    by_category = {}
    for inv in investments:
        cat = inv.get('investmentCategory', 'Other')
        if cat not in by_category:
            by_category[cat] = {"invested": 0, "current": 0, "count": 0}
        by_category[cat]["invested"] += inv.get('principal', 0)
        by_category[cat]["current"] += inv.get('currentValue', 0)
        by_category[cat]["count"] += 1
    
    return {
        "totalInvested": total_invested,
        "currentValue": current_value,
        "totalGains": total_gains,
        "gainPercent": round(gain_percent, 2),
        "byCategory": by_category
    }

# ============ REPORTS API ============
from io import BytesIO
from fastapi.responses import StreamingResponse

@api_router.get("/reports/generate/{report_type}")
async def generate_report(
    request: Request, 
    report_type: str, 
    format: str = "pdf",
    from_date: Optional[str] = None,
    to_date: Optional[str] = None
):
    """Generate and download financial reports"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    user_name = user.get('name', 'User')
    user_filter = {"userId": user_id}
    
    # Parse dates
    try:
        start_date = datetime.strptime(from_date, "%Y-%m-%d") if from_date else datetime.now(timezone.utc).replace(day=1)
        end_date = datetime.strptime(to_date, "%Y-%m-%d") if to_date else datetime.now(timezone.utc)
    except:
        start_date = datetime.now(timezone.utc).replace(day=1)
        end_date = datetime.now(timezone.utc)
    
    # Fetch relevant data based on report type
    data = {}
    
    if report_type in ["income", "cashflow", "networth"]:
        data["incomes"] = await db.income_sources.find(user_filter, {"_id": 0}).to_list(1000)
    
    if report_type in ["expense", "cashflow", "networth"]:
        data["expenses"] = await db.expenses.find(user_filter, {"_id": 0}).to_list(1000)
    
    if report_type in ["loan", "networth"]:
        data["loans"] = await db.loans.find(user_filter, {"_id": 0}).to_list(1000)
    
    if report_type in ["investment", "networth"]:
        data["investments"] = await db.investments.find(user_filter, {"_id": 0}).to_list(1000)
    
    if report_type in ["networth"]:
        data["assets"] = await db.assets.find(user_filter, {"_id": 0}).to_list(1000)
        data["accounts"] = await db.accounts.find(user_filter, {"_id": 0}).to_list(1000)
        data["credit_cards"] = await db.credit_cards.find(user_filter, {"_id": 0}).to_list(1000)
    
    if report_type == "goal":
        data["goals"] = await db.goals.find(user_filter, {"_id": 0}).to_list(1000)
    
    # Generate report
    if format == "excel":
        return await generate_excel_report(report_type, data, user_name, start_date, end_date)
    else:
        return await generate_pdf_report(report_type, data, user_name, start_date, end_date)

async def generate_pdf_report(report_type: str, data: dict, user_name: str, start_date: datetime, end_date: datetime):
    """Generate PDF report using ReportLab"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=20)
    report_titles = {
        "income": "Income Report",
        "expense": "Expense Report",
        "cashflow": "Cash Flow Report",
        "loan": "Loan Report",
        "investment": "Investment Report",
        "networth": "Net Worth Report",
        "goal": "Goal Progress Report"
    }
    elements.append(Paragraph(report_titles.get(report_type, "Financial Report"), title_style))
    elements.append(Paragraph(f"Generated for: {user_name}", styles['Normal']))
    elements.append(Paragraph(f"Period: {start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Generate content based on report type
    if report_type == "income" and data.get("incomes"):
        elements.append(Paragraph("Income Sources", styles['Heading2']))
        table_data = [["Source", "Type", "Amount", "Frequency"]]
        total = 0
        for inc in data["incomes"]:
            amt = inc.get('expectedAmount', 0)
            total += amt
            table_data.append([
                inc.get('name', 'N/A'),
                inc.get('type', 'N/A').title(),
                f"₹{amt:,.0f}",
                inc.get('frequency', 'Monthly')
            ])
        table_data.append(["Total", "", f"₹{total:,.0f}", ""])
        
        table = Table(table_data, colWidths=[2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D1FAE5')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB'))
        ]))
        elements.append(table)
    
    elif report_type == "expense" and data.get("expenses"):
        elements.append(Paragraph("Expense Breakdown", styles['Heading2']))
        table_data = [["Expense", "Category", "Amount", "Frequency"]]
        total = 0
        for exp in data["expenses"]:
            amt = exp.get('expectedAmount', 0)
            total += amt
            table_data.append([
                exp.get('expenseName', 'N/A'),
                exp.get('category', 'N/A'),
                f"₹{amt:,.0f}",
                exp.get('frequency', 'Monthly')
            ])
        table_data.append(["Total", "", f"₹{total:,.0f}", ""])
        
        table = Table(table_data, colWidths=[2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EF4444')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEE2E2')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB'))
        ]))
        elements.append(table)
    
    elif report_type == "investment" and data.get("investments"):
        elements.append(Paragraph("Investment Portfolio", styles['Heading2']))
        table_data = [["Name", "Category", "Invested", "Current Value", "Gain/Loss"]]
        total_invested = 0
        total_current = 0
        for inv in data["investments"]:
            invested = inv.get('principal', 0)
            current = inv.get('currentValue', 0)
            gain = current - invested
            total_invested += invested
            total_current += current
            table_data.append([
                inv.get('name', 'N/A')[:20],
                inv.get('investmentCategory', 'N/A'),
                f"₹{invested:,.0f}",
                f"₹{current:,.0f}",
                f"₹{gain:,.0f}"
            ])
        table_data.append(["Total", "", f"₹{total_invested:,.0f}", f"₹{total_current:,.0f}", f"₹{total_current - total_invested:,.0f}"])
        
        table = Table(table_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1.1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EDE9FE')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB'))
        ]))
        elements.append(table)
    
    elif report_type == "loan" and data.get("loans"):
        elements.append(Paragraph("Loan Details", styles['Heading2']))
        table_data = [["Loan Name", "Type", "Outstanding", "EMI", "Interest Rate"]]
        total_outstanding = 0
        for loan in data["loans"]:
            outstanding = loan.get('outstandingAmount', 0)
            total_outstanding += outstanding
            table_data.append([
                loan.get('loanName', 'N/A')[:20],
                loan.get('loanType', 'N/A'),
                f"₹{outstanding:,.0f}",
                f"₹{loan.get('emiAmount', 0):,.0f}",
                f"{loan.get('interestRate', 0)}%"
            ])
        table_data.append(["Total Outstanding", "", f"₹{total_outstanding:,.0f}", "", ""])
        
        table = Table(table_data, colWidths=[1.4*inch, 1.1*inch, 1.1*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEF3C7')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB'))
        ]))
        elements.append(table)
    
    elif report_type == "networth":
        # Calculate totals
        total_assets = sum(a.get('currentValue', 0) for a in data.get("assets", []))
        total_investments = sum(i.get('currentValue', 0) for i in data.get("investments", []))
        liquid_balance = sum(a.get('currentBalance', 0) or a.get('balance', 0) for a in data.get("accounts", []))
        total_loans = sum(l.get('outstandingAmount', 0) for l in data.get("loans", []))
        total_cc = sum(c.get('currentOutstanding', 0) or c.get('outstandingAmount', 0) for c in data.get("credit_cards", []))
        total_liabilities = total_loans + total_cc
        net_worth = total_assets + total_investments + liquid_balance - total_liabilities
        
        elements.append(Paragraph("Net Worth Summary", styles['Heading2']))
        summary_data = [
            ["Category", "Amount"],
            ["Total Assets", f"₹{total_assets:,.0f}"],
            ["Total Investments", f"₹{total_investments:,.0f}"],
            ["Liquid Balance (Bank)", f"₹{liquid_balance:,.0f}"],
            ["Total Liabilities", f"₹{total_liabilities:,.0f}"],
            ["Net Worth", f"₹{net_worth:,.0f}"]
        ]
        
        table = Table(summary_data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#DBEAFE')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB'))
        ]))
        elements.append(table)
    
    elif report_type == "goal" and data.get("goals"):
        elements.append(Paragraph("Financial Goals Progress", styles['Heading2']))
        table_data = [["Goal", "Target", "Current", "Progress", "Target Date"]]
        for goal in data["goals"]:
            target = goal.get('targetAmount', 0)
            current = goal.get('currentAmount', 0)
            progress = (current / target * 100) if target > 0 else 0
            target_date = goal.get('targetDate', 'N/A')
            if isinstance(target_date, datetime):
                target_date = target_date.strftime('%d %b %Y')
            table_data.append([
                goal.get('goalName', 'N/A')[:20],
                f"₹{target:,.0f}",
                f"₹{current:,.0f}",
                f"{progress:.0f}%",
                target_date[:10] if target_date else 'N/A'
            ])
        
        table = Table(table_data, colWidths=[1.5*inch, 1*inch, 1*inch, 0.8*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EC4899')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB'))
        ]))
        elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"{report_type}_report_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

async def generate_excel_report(report_type: str, data: dict, user_name: str, start_date: datetime, end_date: datetime):
    """Generate Excel report using openpyxl"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws = wb.active
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    report_titles = {
        "income": "Income Report",
        "expense": "Expense Report",
        "cashflow": "Cash Flow Report",
        "loan": "Loan Report",
        "investment": "Investment Report",
        "networth": "Net Worth Report",
        "goal": "Goal Progress Report"
    }
    
    ws.title = report_titles.get(report_type, "Report")
    
    # Header
    ws['A1'] = report_titles.get(report_type, "Financial Report")
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Generated for: {user_name}"
    ws['A3'] = f"Period: {start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}"
    
    row = 5
    
    if report_type == "income" and data.get("incomes"):
        headers = ["Source", "Type", "Amount", "Frequency"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row += 1
        total = 0
        for inc in data["incomes"]:
            amt = inc.get('expectedAmount', 0)
            total += amt
            ws.cell(row=row, column=1, value=inc.get('name', 'N/A')).border = border
            ws.cell(row=row, column=2, value=inc.get('type', 'N/A').title()).border = border
            ws.cell(row=row, column=3, value=amt).border = border
            ws.cell(row=row, column=4, value=inc.get('frequency', 'Monthly')).border = border
            row += 1
        
        ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=row, column=3, value=total).font = Font(bold=True)
    
    elif report_type == "expense" and data.get("expenses"):
        header_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        headers = ["Expense", "Category", "Amount", "Frequency"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row += 1
        total = 0
        for exp in data["expenses"]:
            amt = exp.get('expectedAmount', 0)
            total += amt
            ws.cell(row=row, column=1, value=exp.get('expenseName', 'N/A')).border = border
            ws.cell(row=row, column=2, value=exp.get('category', 'N/A')).border = border
            ws.cell(row=row, column=3, value=amt).border = border
            ws.cell(row=row, column=4, value=exp.get('frequency', 'Monthly')).border = border
            row += 1
        
        ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=row, column=3, value=total).font = Font(bold=True)
    
    elif report_type == "investment" and data.get("investments"):
        header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        headers = ["Name", "Category", "Invested", "Current Value", "Gain/Loss"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row += 1
        total_invested = 0
        total_current = 0
        for inv in data["investments"]:
            invested = inv.get('principal', 0)
            current = inv.get('currentValue', 0)
            total_invested += invested
            total_current += current
            ws.cell(row=row, column=1, value=inv.get('name', 'N/A')).border = border
            ws.cell(row=row, column=2, value=inv.get('investmentCategory', 'N/A')).border = border
            ws.cell(row=row, column=3, value=invested).border = border
            ws.cell(row=row, column=4, value=current).border = border
            ws.cell(row=row, column=5, value=current - invested).border = border
            row += 1
        
        ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_invested).font = Font(bold=True)
        ws.cell(row=row, column=4, value=total_current).font = Font(bold=True)
        ws.cell(row=row, column=5, value=total_current - total_invested).font = Font(bold=True)
    
    elif report_type == "loan" and data.get("loans"):
        header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        headers = ["Loan Name", "Type", "Outstanding", "EMI", "Interest Rate"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row += 1
        total = 0
        for loan in data["loans"]:
            outstanding = loan.get('outstandingAmount', 0)
            total += outstanding
            ws.cell(row=row, column=1, value=loan.get('loanName', 'N/A')).border = border
            ws.cell(row=row, column=2, value=loan.get('loanType', 'N/A')).border = border
            ws.cell(row=row, column=3, value=outstanding).border = border
            ws.cell(row=row, column=4, value=loan.get('emiAmount', 0)).border = border
            ws.cell(row=row, column=5, value=f"{loan.get('interestRate', 0)}%").border = border
            row += 1
        
        ws.cell(row=row, column=1, value="Total Outstanding").font = Font(bold=True)
        ws.cell(row=row, column=3, value=total).font = Font(bold=True)
    
    elif report_type == "networth":
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        
        total_assets = sum(a.get('currentValue', 0) for a in data.get("assets", []))
        total_investments = sum(i.get('currentValue', 0) for i in data.get("investments", []))
        liquid_balance = sum(a.get('currentBalance', 0) or a.get('balance', 0) for a in data.get("accounts", []))
        total_loans = sum(l.get('outstandingAmount', 0) for l in data.get("loans", []))
        total_cc = sum(c.get('currentOutstanding', 0) or c.get('outstandingAmount', 0) for c in data.get("credit_cards", []))
        net_worth = total_assets + total_investments + liquid_balance - total_loans - total_cc
        
        headers = ["Category", "Amount"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row += 1
        items = [
            ("Total Assets", total_assets),
            ("Total Investments", total_investments),
            ("Liquid Balance", liquid_balance),
            ("Loans Outstanding", -total_loans),
            ("Credit Card Due", -total_cc),
            ("Net Worth", net_worth)
        ]
        for item, value in items:
            ws.cell(row=row, column=1, value=item).border = border
            ws.cell(row=row, column=2, value=value).border = border
            row += 1
        
        ws.cell(row=row-1, column=1).font = Font(bold=True)
        ws.cell(row=row-1, column=2).font = Font(bold=True)
    
    elif report_type == "goal" and data.get("goals"):
        header_fill = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
        headers = ["Goal", "Target", "Current", "Progress %", "Target Date"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row += 1
        for goal in data["goals"]:
            target = goal.get('targetAmount', 0)
            current = goal.get('currentAmount', 0)
            progress = (current / target * 100) if target > 0 else 0
            target_date = goal.get('targetDate', 'N/A')
            if isinstance(target_date, datetime):
                target_date = target_date.strftime('%d %b %Y')
            
            ws.cell(row=row, column=1, value=goal.get('goalName', 'N/A')).border = border
            ws.cell(row=row, column=2, value=target).border = border
            ws.cell(row=row, column=3, value=current).border = border
            ws.cell(row=row, column=4, value=f"{progress:.0f}%").border = border
            ws.cell(row=row, column=5, value=str(target_date)[:10] if target_date else 'N/A').border = border
            row += 1
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 30)
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"{report_type}_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============ SETTINGS APIs ============
class NotificationPreferences(BaseModel):
    emailNotifications: bool = True
    pushNotifications: bool = True
    smsNotifications: bool = False
    incomeReminders: bool = True
    expenseReminders: bool = True
    billReminders: bool = True
    goalReminders: bool = True
    weeklyDigest: bool = True
    monthlyReport: bool = True

class UserPreferences(BaseModel):
    theme: str = "light"
    accentColor: str = "#10B981"
    currency: str = "INR"
    dateFormat: str = "DD/MM/YYYY"
    language: str = "en"

@api_router.get("/settings/notifications")
async def get_notification_settings(request: Request):
    """Get user notification preferences"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    
    prefs = await db.user_preferences.find_one({"userId": user_id, "type": "notifications"}, {"_id": 0})
    
    if not prefs:
        # Return defaults
        return NotificationPreferences().model_dump()
    
    return prefs.get("settings", NotificationPreferences().model_dump())

@api_router.put("/settings/notifications")
async def update_notification_settings(request: Request, prefs: NotificationPreferences):
    """Update user notification preferences"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    
    await db.user_preferences.update_one(
        {"userId": user_id, "type": "notifications"},
        {"$set": {"userId": user_id, "type": "notifications", "settings": prefs.model_dump(), "updatedAt": datetime.now(timezone.utc)}},
        upsert=True
    )
    
    return {"message": "Notification preferences updated", "settings": prefs.model_dump()}

@api_router.get("/settings/preferences")
async def get_user_preferences(request: Request):
    """Get user app preferences"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    
    prefs = await db.user_preferences.find_one({"userId": user_id, "type": "app"}, {"_id": 0})
    
    if not prefs:
        return UserPreferences().model_dump()
    
    return prefs.get("settings", UserPreferences().model_dump())

@api_router.put("/settings/preferences")
async def update_user_preferences(request: Request, prefs: UserPreferences):
    """Update user app preferences"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    
    await db.user_preferences.update_one(
        {"userId": user_id, "type": "app"},
        {"$set": {"userId": user_id, "type": "app", "settings": prefs.model_dump(), "updatedAt": datetime.now(timezone.utc)}},
        upsert=True
    )
    
    return {"message": "Preferences updated", "settings": prefs.model_dump()}

@api_router.get("/settings/data-export")
async def export_user_data(request: Request):
    """Export all user data as JSON"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    user_filter = {"userId": user_id}
    
    # Gather all user data
    data = {
        "user": {
            "email": user.get('email'),
            "name": user.get('name'),
            "exportedAt": datetime.now(timezone.utc).isoformat()
        },
        "assets": await db.assets.find(user_filter, {"_id": 0}).to_list(1000),
        "investments": await db.investments.find(user_filter, {"_id": 0}).to_list(1000),
        "accounts": await db.accounts.find(user_filter, {"_id": 0}).to_list(1000),
        "loans": await db.loans.find(user_filter, {"_id": 0}).to_list(1000),
        "credit_cards": await db.credit_cards.find(user_filter, {"_id": 0}).to_list(1000),
        "income_sources": await db.income_sources.find(user_filter, {"_id": 0}).to_list(1000),
        "expenses": await db.expenses.find(user_filter, {"_id": 0}).to_list(1000),
        "goals": await db.goals.find(user_filter, {"_id": 0}).to_list(1000),
        "insurances": await db.insurances.find(user_filter, {"_id": 0}).to_list(1000)
    }
    
    import json
    json_data = json.dumps(data, default=str, indent=2)
    buffer = BytesIO(json_data.encode())
    
    filename = f"moneyssutra_data_export_{datetime.now().strftime('%Y%m%d')}.json"
    return StreamingResponse(
        buffer,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.delete("/settings/delete-account")
async def delete_user_account(request: Request):
    """Delete user account and all associated data"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    user_id = user.get('user_id')
    user_filter = {"userId": user_id}
    
    # Delete all user data from all collections
    await asyncio.gather(
        db.assets.delete_many(user_filter),
        db.investments.delete_many(user_filter),
        db.accounts.delete_many(user_filter),
        db.loans.delete_many(user_filter),
        db.credit_cards.delete_many(user_filter),
        db.income_sources.delete_many(user_filter),
        db.other_income.delete_many(user_filter),
        db.expenses.delete_many(user_filter),
        db.goals.delete_many(user_filter),
        db.insurances.delete_many(user_filter),
        db.user_preferences.delete_many(user_filter),
        db.analytics_snapshots.delete_many(user_filter),
        db.notifications.delete_many(user_filter),
        db.income_transactions.delete_many(user_filter),
        db.expense_transactions.delete_many(user_filter),
        db.user_sessions.delete_many({"user_id": user_id}),
        db.users.delete_one({"user_id": user_id})
    )
    
    return {"message": "Account and all data deleted successfully"}

@app.on_event("startup")
async def startup_db_client():
    """Create database indexes and start background scheduler"""
    try:
        # Create indexes for user isolation queries
        await db.assets.create_index("userId")
        await db.assets.create_index([("userId", 1), ("name", 1)])  # Compound index for name lookup
        await db.investments.create_index("userId")
        await db.investments.create_index([("userId", 1), ("name", 1)])
        await db.loans.create_index("userId")
        await db.loans.create_index([("userId", 1), ("loanName", 1)])
        await db.accounts.create_index("userId")
        await db.accounts.create_index([("userId", 1), ("accountName", 1)])
        await db.credit_cards.create_index("userId")
        await db.income_sources.create_index("userId")
        await db.income_sources.create_index([("userId", 1), ("name", 1)])  # For name uniqueness check
        await db.income_sources.create_index([("userId", 1), ("type", 1)])  # For type filtering
        await db.other_income.create_index("userId")
        await db.other_income.create_index([("userId", 1), ("incomeName", 1)])
        await db.expenses.create_index("userId")
        await db.expenses.create_index([("userId", 1), ("expenseName", 1)])
        await db.expenses.create_index([("userId", 1), ("category", 1)])  # For category filtering
        await db.goals.create_index("userId")
        await db.insurances.create_index("userId")
        await db.insurances.create_index([("userId", 1), ("policyName", 1)])
        await db.insurances.create_index([("userId", 1), ("insuranceType", 1)])
        await db.user_sessions.create_index("session_token")
        await db.users.create_index("user_id")
        await db.users.create_index("email")
        await db.notifications.create_index([("userId", 1), ("createdAt", -1)])
        await db.notifications.create_index([("userId", 1), ("relatedIncomeId", 1)])  # For dismissing by entity
        # Transaction collections indexes
        await db.income_transactions.create_index([("userId", 1), ("transactionDate", -1)])
        await db.income_transactions.create_index([("entityId", 1)])
        await db.income_transactions.create_index([("entityId", 1), ("transactionDate", -1)])  # For latest entry
        await db.expense_transactions.create_index([("userId", 1), ("transactionDate", -1)])
        await db.expense_transactions.create_index([("entityId", 1)])
        await db.expense_transactions.create_index([("entityId", 1), ("transactionDate", -1)])  # For latest entry
        logger.info("Database indexes created successfully")
        
        # Start background reminder scheduler
        asyncio.create_task(check_and_send_reminders())
        logger.info("Background reminder scheduler task created")
        
    except Exception as e:
        logger.warning(f"Startup warning: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    global scheduler_running
    scheduler_running = False
    client.close()
    logger.info("Background scheduler stopped and database connection closed")